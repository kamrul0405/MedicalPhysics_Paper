"""v200: Does per-patient UODSL λ correlate with molecular features
(IDH1, MGMT) on MU-Glioma-Post n=102? — round 38 (CPU).

Senior-Nature-reviewer-driven biological validation of Layer 2:
round 34/37 established λ is patient-intrinsic. The next critical
question: does λ correlate with KNOWN BIOLOGICAL FEATURES
(IDH1 mutation status, MGMT methylation status)?

If λ differs by IDH or MGMT subgroup → λ has a biological mechanism,
not just statistical persistence. This would link the UODSL physics
parameter directly to molecular oncology — a major Nature/Cell-level
contribution.

Method (MU-Glioma-Post, n=102 with valid per-patient λ):
  - Load MU clinical xlsx (IDH1, MGMT, Age)
  - Load per-patient λ from v198
  - Statistical tests:
    * Mann-Whitney U: λ_IDH-mut vs λ_IDH-wt
    * Mann-Whitney U: λ_MGMT-methylated vs λ_MGMT-unmethylated
    * Spearman: λ vs Age
    * Subgroup means/medians
  - Multiple-testing: Bonferroni for 3 primary tests

Outputs:
  Nature_project/05_results/v200_lambda_vs_molecular.json
  Nature_project/05_results/v200_lambda_vs_molecular.csv
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import distance_transform_edt
from scipy.stats import mannwhitneyu, spearmanr

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_XLSX = Path(r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v200_lambda_vs_molecular.json"
OUT_CSV = RESULTS / "v200_lambda_vs_molecular.csv"

DISTANCE_BINS = np.arange(1, 25)


def fit_lambda(mask, outgrowth):
    m = mask.astype(bool)
    out = outgrowth.astype(bool)
    if m.sum() == 0 or out.sum() == 0:
        return float("nan"), float("nan"), 0
    d = distance_transform_edt(~m)
    d_int = np.round(d).astype(int)
    d_arr, p_arr, n_arr = [], [], []
    for b in DISTANCE_BINS:
        sel = (d_int == b)
        if sel.sum() < 5:
            continue
        n_total = int(sel.sum())
        n_out = int((sel & out).sum())
        d_arr.append(b)
        p_arr.append(n_out / n_total)
        n_arr.append(n_total)
    d_arr = np.array(d_arr, dtype=float)
    p_arr = np.array(p_arr, dtype=float)
    n_arr = np.array(n_arr, dtype=float)
    valid = p_arr > 0
    if valid.sum() < 3:
        return float("nan"), float("nan"), int(valid.sum())
    d_v = d_arr[valid]
    p_v = p_arr[valid]
    n_v = n_arr[valid]
    log_p = np.log(p_v)
    w = np.sqrt(n_v)
    sw = np.sum(w)
    sw_d = np.sum(w * d_v)
    sw_logp = np.sum(w * log_p)
    sw_dd = np.sum(w * d_v * d_v)
    sw_d_logp = np.sum(w * d_v * log_p)
    denom = sw * sw_dd - sw_d ** 2
    if denom == 0:
        return float("nan"), float("nan"), int(valid.sum())
    slope = (sw * sw_d_logp - sw_d * sw_logp) / denom
    intercept = (sw_logp - slope * sw_d) / sw
    lam = float(-1.0 / slope) if slope < 0 else float("inf")
    pred = intercept + slope * d_v
    ss_res = float(np.sum(w * (log_p - pred) ** 2))
    ss_tot = float(np.sum(w * (log_p - np.mean(log_p)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return lam, r2, int(valid.sum())


def main():
    print("=" * 78, flush=True)
    print("v200 LAMBDA vs MOLECULAR FEATURES (MU-Glioma-Post)", flush=True)
    print("=" * 78, flush=True)

    # Load MU clinical
    wb = openpyxl.load_workbook(CLINICAL_XLSX, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    grade_col = header.index("Grade of Primary Brain Tumor")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")

    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = float(row[age_col]) if row[age_col] is not None else None
            grade = float(row[grade_col]) if row[grade_col] is not None else None
            idh1 = (float(row[idh1_col]) if row[idh1_col] is not None
                      else None)
            mgmt = (float(row[mgmt_col]) if row[mgmt_col] is not None
                      else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "grade": grade, "idh1": idh1, "mgmt": mgmt,
        }
    print(f"  Loaded clinical for {len(clinical)} MU patients", flush=True)

    # Compute per-patient lambda for MU
    print("\nComputing per-patient lambda for MU-Glioma-Post...",
          flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        fr = f.parent / f.name.replace("_b.npy", "_r.npy")
        if not fr.exists():
            continue
        fu = (np.load(fr) > 0).astype(np.float32)
        outgrowth = (fu.astype(bool) & ~m.astype(bool)).astype(np.float32)
        if m.sum() == 0 or outgrowth.sum() == 0:
            continue
        lam, r2, n_pts = fit_lambda(m, outgrowth)
        valid = (not np.isnan(lam)) and (not np.isinf(lam)) and \
                lam > 0 and lam < 200 and r2 > 0.5 and n_pts >= 4
        if not valid:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid,
            "lambda": float(lam),
            "r2": float(r2),
            "age": c["age"],
            "grade": c["grade"],
            "idh1": c["idh1"],
            "mgmt": c["mgmt"],
        })
    print(f"  {len(rows)} MU patients with valid lambda + clinical",
          flush=True)

    # ---------- Test 1: lambda vs IDH1 ----------
    print("\n=== Test 1: λ vs IDH1 mutation ===", flush=True)
    idh_groups = {0: [], 1: []}
    for r in rows:
        if r["idh1"] is not None and r["idh1"] in (0, 1):
            idh_groups[int(r["idh1"])].append(r["lambda"])
    print(f"  IDH1 mutant (1):    n = {len(idh_groups[1])}  "
          f"mean λ = {np.mean(idh_groups[1]):.3f}  median = "
          f"{np.median(idh_groups[1]):.3f}", flush=True)
    print(f"  IDH1 wild-type (0): n = {len(idh_groups[0])}  "
          f"mean λ = {np.mean(idh_groups[0]):.3f}  median = "
          f"{np.median(idh_groups[0]):.3f}", flush=True)
    if len(idh_groups[0]) >= 5 and len(idh_groups[1]) >= 5:
        u, p_idh = mannwhitneyu(idh_groups[1], idh_groups[0],
                                  alternative="two-sided")
        print(f"  Mann-Whitney U: U = {u:.0f}, p = {p_idh:.4f}", flush=True)
    else:
        u, p_idh = float("nan"), float("nan")
        print(f"  Insufficient sample for IDH1 test", flush=True)

    # ---------- Test 2: lambda vs MGMT methylation ----------
    print("\n=== Test 2: λ vs MGMT methylation ===", flush=True)
    mgmt_groups = {0: [], 1: []}
    for r in rows:
        if r["mgmt"] is not None and r["mgmt"] in (0, 1):
            mgmt_groups[int(r["mgmt"])].append(r["lambda"])
    print(f"  MGMT methylated (1):     n = {len(mgmt_groups[1])}  "
          f"mean λ = {np.mean(mgmt_groups[1]):.3f}  median = "
          f"{np.median(mgmt_groups[1]):.3f}", flush=True)
    print(f"  MGMT unmethylated (0):   n = {len(mgmt_groups[0])}  "
          f"mean λ = {np.mean(mgmt_groups[0]):.3f}  median = "
          f"{np.median(mgmt_groups[0]):.3f}", flush=True)
    if len(mgmt_groups[0]) >= 5 and len(mgmt_groups[1]) >= 5:
        u_m, p_mgmt = mannwhitneyu(mgmt_groups[1], mgmt_groups[0],
                                     alternative="two-sided")
        print(f"  Mann-Whitney U: U = {u_m:.0f}, p = {p_mgmt:.4f}",
              flush=True)
    else:
        u_m, p_mgmt = float("nan"), float("nan")
        print(f"  Insufficient sample for MGMT test", flush=True)

    # ---------- Test 3: lambda vs Age ----------
    print("\n=== Test 3: λ vs Age (Spearman) ===", flush=True)
    ages = [r["age"] for r in rows if r["age"] is not None]
    lams_age = [r["lambda"] for r in rows if r["age"] is not None]
    if len(ages) >= 5:
        rho_age, p_age = spearmanr(ages, lams_age)
        print(f"  n = {len(ages)}  rho = {rho_age:+.4f}  p = {p_age:.4f}",
              flush=True)
    else:
        rho_age, p_age = float("nan"), float("nan")

    # Bonferroni correction for 3 primary tests
    p_values = [p_idh, p_mgmt, p_age]
    valid_ps = [p for p in p_values if not np.isnan(p)]
    n_tests = len(valid_ps)
    print(f"\n=== Multiple-testing correction (Bonferroni, "
          f"n_tests = {n_tests}) ===", flush=True)
    p_idh_bf = (p_idh * n_tests) if not np.isnan(p_idh) else float("nan")
    p_mgmt_bf = (p_mgmt * n_tests) if not np.isnan(p_mgmt) else float("nan")
    p_age_bf = (p_age * n_tests) if not np.isnan(p_age) else float("nan")
    print(f"  IDH1: p = {p_idh:.4f} → Bonferroni p = "
          f"{min(p_idh_bf, 1.0):.4f}", flush=True)
    print(f"  MGMT: p = {p_mgmt:.4f} → Bonferroni p = "
          f"{min(p_mgmt_bf, 1.0):.4f}", flush=True)
    print(f"  Age:  p = {p_age:.4f} → Bonferroni p = "
          f"{min(p_age_bf, 1.0):.4f}", flush=True)

    # ---------- Cross-tabulation ----------
    print("\n=== Cross-tabulation: λ by (IDH1, MGMT) subgroups ===",
          flush=True)
    subgroups = {}
    for r in rows:
        if r["idh1"] is None or r["mgmt"] is None:
            continue
        if r["idh1"] not in (0, 1) or r["mgmt"] not in (0, 1):
            continue
        key = f"IDH={int(r['idh1'])}_MGMT={int(r['mgmt'])}"
        subgroups.setdefault(key, []).append(r["lambda"])
    for key in sorted(subgroups.keys()):
        ls = subgroups[key]
        print(f"  {key:25s}  n = {len(ls):3d}  "
              f"mean = {np.mean(ls):.3f}  median = {np.median(ls):.3f}",
              flush=True)

    # ---------- Save ----------
    out = {
        "version": "v200",
        "experiment": ("Per-patient lambda vs molecular features (IDH1, "
                       "MGMT) on MU-Glioma-Post"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "n_total": len(rows),
        "test_1_idh1": {
            "n_mut": len(idh_groups[1]),
            "n_wt": len(idh_groups[0]),
            "mean_lambda_mut": (float(np.mean(idh_groups[1]))
                                  if idh_groups[1] else None),
            "mean_lambda_wt": (float(np.mean(idh_groups[0]))
                                 if idh_groups[0] else None),
            "median_lambda_mut": (float(np.median(idh_groups[1]))
                                    if idh_groups[1] else None),
            "median_lambda_wt": (float(np.median(idh_groups[0]))
                                   if idh_groups[0] else None),
            "mannwhitney_u": float(u) if not np.isnan(u) else None,
            "p_value_raw": float(p_idh) if not np.isnan(p_idh) else None,
            "p_value_bonferroni": (float(min(p_idh_bf, 1.0))
                                     if not np.isnan(p_idh_bf) else None),
        },
        "test_2_mgmt": {
            "n_methylated": len(mgmt_groups[1]),
            "n_unmethylated": len(mgmt_groups[0]),
            "mean_lambda_methylated": (float(np.mean(mgmt_groups[1]))
                                          if mgmt_groups[1] else None),
            "mean_lambda_unmethylated": (float(np.mean(mgmt_groups[0]))
                                            if mgmt_groups[0] else None),
            "median_lambda_methylated": (float(np.median(mgmt_groups[1]))
                                            if mgmt_groups[1] else None),
            "median_lambda_unmethylated": (float(np.median(mgmt_groups[0]))
                                              if mgmt_groups[0] else None),
            "mannwhitney_u": float(u_m) if not np.isnan(u_m) else None,
            "p_value_raw": float(p_mgmt) if not np.isnan(p_mgmt) else None,
            "p_value_bonferroni": (float(min(p_mgmt_bf, 1.0))
                                     if not np.isnan(p_mgmt_bf) else None),
        },
        "test_3_age": {
            "n": len(ages),
            "spearman_rho": float(rho_age) if not np.isnan(rho_age) else None,
            "p_value_raw": float(p_age) if not np.isnan(p_age) else None,
            "p_value_bonferroni": (float(min(p_age_bf, 1.0))
                                     if not np.isnan(p_age_bf) else None),
        },
        "subgroup_means": {k: float(np.mean(v)) for k, v in subgroups.items()},
        "subgroup_n": {k: len(v) for k, v in subgroups.items()},
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)

    if rows:
        with OUT_CSV.open("w", newline="") as fp:
            w = csv.DictWriter(fp, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        print(f"Saved {OUT_CSV}", flush=True)


if __name__ == "__main__":
    main()
