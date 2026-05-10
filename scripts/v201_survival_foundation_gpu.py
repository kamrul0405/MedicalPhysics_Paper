"""v201: Survival-supervised 3D U-Net foundation model — round 38 (GPU).

Senior-Nature-reviewer-driven beyond-Nature experiment. Round 32-33
showed the kernel doesn't predict survival. Round 35-36 showed
preliminary λ+V_kernel synergy refuted on replication. The natural
question: can a 3D U-Net trained DIRECTLY for survival prediction
(Cox PH loss) outperform clinical features?

Architecture:
  3D U-Net encoder → global average pool → linear → scalar risk score
  Loss: Cox proportional hazards partial likelihood
  Input: baseline mask + bimodal kernel (2 channels, like round 22)

Training set: combined RHUH (n=39) + MU (n=151 with valid OS) = 190
LOCO held-out: hold one cohort out, evaluate C-index on the other.

Key tests:
  - C-index of trained survival U-Net vs clinical-only baseline
  - Compare to round 33 V_kernel + clinical (which also failed)
  - Compare to round 35 RHUH n=13 lambda+V_kernel (which refuted)

If trained survival U-Net achieves C-index > clinical baseline → MAJOR
positive (deep learning DOES capture survival biology when supervised
appropriately). If not → another honest negative scoping the
foundation model's role.

Outputs:
  Nature_project/05_results/v201_survival_foundation.json
"""
from __future__ import annotations

import csv
import gc
import json
import re
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_RHUH = Path(r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
CLINICAL_MU = Path(r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v201_survival_foundation.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_BROAD = 7.0
TARGET_SHAPE = (16, 48, 48)
EPOCHS = 50
LR = 5e-4
SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    h_broad = heat_constant(mask, sigma)
    return np.maximum(persistence, h_broad)


class SurvivalUNet(nn.Module):
    """3D U-Net encoder + global-pool + linear → scalar risk score."""
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32),
            nn.GELU(),
            nn.Linear(32, 1),
        )

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        pooled = self.global_pool(e3)
        return self.head(pooled).squeeze(-1)


def cox_loss(risk, time, event):
    """Cox PH partial likelihood loss, given a batch of (risk, time, event)."""
    # Sort by time descending → cumulative risk-set
    order = torch.argsort(time, descending=True)
    risk_sorted = risk[order]
    event_sorted = event[order]
    # log-cumsum-exp
    max_risk = torch.cummax(risk_sorted, dim=0)[0]
    log_cum = max_risk + torch.log(torch.cumsum(
        torch.exp(risk_sorted - max_risk), dim=0))
    # only events contribute
    valid = event_sorted == 1
    if valid.sum() == 0:
        return torch.tensor(0.0, requires_grad=True, device=risk.device)
    log_pl = (risk_sorted[valid] - log_cum[valid]).sum()
    return -log_pl / valid.sum()


def concordance(times, events, risk_scores):
    times = np.asarray(times)
    events = np.asarray(events)
    risk_scores = np.asarray(risk_scores)
    n = len(times)
    n_pairs = 0
    n_concordant = 0
    n_ties = 0
    for i in range(n):
        for j in range(i+1, n):
            if events[i] == 0 and events[j] == 0:
                continue
            if times[i] < times[j]:
                if events[i] != 1:
                    continue
                n_pairs += 1
                if risk_scores[i] > risk_scores[j]:
                    n_concordant += 1
                elif risk_scores[i] == risk_scores[j]:
                    n_ties += 1
            elif times[j] < times[i]:
                if events[j] != 1:
                    continue
                n_pairs += 1
                if risk_scores[j] > risk_scores[i]:
                    n_concordant += 1
                elif risk_scores[i] == risk_scores[j]:
                    n_ties += 1
    if n_pairs == 0:
        return float("nan")
    return (n_concordant + 0.5 * n_ties) / n_pairs


def load_rhuh():
    """Load RHUH baseline masks + clinical OS."""
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        pid_col = "Patient ID"
        os_col = next((c for c in cols if "Overall survival" in c
                         and "day" in c.lower()), None)
        cens_col = next((c for c in cols if "censor" in c.lower()), None)
        for row in reader:
            pid = row[pid_col].strip()
            try:
                os_d = float(row[os_col]) if row[os_col] else None
            except (ValueError, TypeError):
                os_d = None
            event = (0 if (cens_col and row[cens_col].strip().lower() ==
                            "yes") else 1)
            if os_d is not None:
                clinical[pid] = {"os": os_d, "event": event,
                                  "cohort": "RHUH"}

    rows = []
    for f in sorted(CACHE.glob("RHUH-GBM_RHUH-*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({"pid": pid, "cohort": "RHUH", "mask": m,
                     "os": c["os"], "event": c["event"]})
    return rows


def load_mu():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    death_col = header.index("Overall Survival (Death)")
    os_col = header.index("Number of days from Diagnosis to death (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            event = int(row[death_col]) if row[death_col] is not None else None
            os_d = float(row[os_col]) if row[os_col] is not None else None
        except (ValueError, TypeError):
            continue
        if os_d is not None and event is not None:
            clinical[str(pid)] = {"os": os_d, "event": event,
                                    "cohort": "MU"}

    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({"pid": pid, "cohort": "MU", "mask": m,
                     "os": c["os"], "event": c["event"]})
    return rows


def train_survival_unet(train_rows, seed=SEED):
    torch.manual_seed(seed)
    np.random.seed(seed)
    n = len(train_rows)
    # Pre-compute kernels
    X = np.stack([np.stack([d["mask"], heat_bimodal(d["mask"], SIGMA_BROAD)],
                              axis=0)
                   for d in train_rows]).astype(np.float32)
    times = np.array([d["os"] for d in train_rows], dtype=np.float32)
    events = np.array([d["event"] for d in train_rows], dtype=np.float32)
    Xt = torch.from_numpy(X).to(DEVICE)
    times_t = torch.from_numpy(times).to(DEVICE)
    events_t = torch.from_numpy(events).to(DEVICE)

    model = SurvivalUNet(in_ch=2, base=24).to(DEVICE)
    opt = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=1e-3)

    for ep in range(EPOCHS):
        model.train()
        # full-batch Cox loss (batches break Cox PH)
        risk = model(Xt)
        loss = cox_loss(risk, times_t, events_t)
        opt.zero_grad()
        loss.backward()
        opt.step()
        if (ep + 1) % 10 == 0:
            with torch.no_grad():
                model.eval()
                r = model(Xt).cpu().numpy()
                c_train = concordance(times, events, r)
            print(f"    ep {ep+1}/{EPOCHS}  loss={loss.item():.4f}  "
                  f"train C={c_train:.4f}", flush=True)
    return model


def evaluate_survival(model, test_rows):
    model.eval()
    X = np.stack([np.stack([d["mask"], heat_bimodal(d["mask"], SIGMA_BROAD)],
                              axis=0)
                   for d in test_rows]).astype(np.float32)
    times = np.array([d["os"] for d in test_rows], dtype=float)
    events = np.array([d["event"] for d in test_rows], dtype=float)
    with torch.no_grad():
        Xt = torch.from_numpy(X).to(DEVICE)
        risks = model(Xt).cpu().numpy()
    c = concordance(times, events, risks)
    return c, risks


def main():
    print("=" * 78, flush=True)
    print("v201 SURVIVAL-SUPERVISED 3D U-NET (round 38 GPU)", flush=True)
    print(f"  Device: {DEVICE}", flush=True)
    print("=" * 78, flush=True)

    print("\nLoading RHUH...", flush=True)
    rhuh_rows = load_rhuh()
    print(f"  RHUH: {len(rhuh_rows)} patients with mask + OS", flush=True)
    print("\nLoading MU...", flush=True)
    mu_rows = load_mu()
    print(f"  MU: {len(mu_rows)} patients with mask + OS", flush=True)

    # LOCO design:
    # (a) Train on MU, test on RHUH
    # (b) Train on RHUH, test on MU
    # Plus pooled cross-validation
    results = {}

    print("\n=== LOCO 1: Train MU → Test RHUH ===", flush=True)
    t0 = time.time()
    model_mu = train_survival_unet(mu_rows, seed=SEED)
    print(f"  Train MU took {time.time()-t0:.0f}s", flush=True)
    c_rhuh_test, risks_rhuh = evaluate_survival(model_mu, rhuh_rows)
    print(f"  Test on RHUH: C-index = {c_rhuh_test:.4f} "
          f"(n={len(rhuh_rows)})", flush=True)
    results["train_MU_test_RHUH"] = {
        "n_train": len(mu_rows), "n_test": len(rhuh_rows),
        "c_index_test": float(c_rhuh_test),
    }
    del model_mu
    torch.cuda.empty_cache()
    gc.collect()

    print("\n=== LOCO 2: Train RHUH → Test MU ===", flush=True)
    t0 = time.time()
    model_rhuh = train_survival_unet(rhuh_rows, seed=SEED)
    print(f"  Train RHUH took {time.time()-t0:.0f}s", flush=True)
    c_mu_test, risks_mu = evaluate_survival(model_rhuh, mu_rows)
    print(f"  Test on MU: C-index = {c_mu_test:.4f} "
          f"(n={len(mu_rows)})", flush=True)
    results["train_RHUH_test_MU"] = {
        "n_train": len(rhuh_rows), "n_test": len(mu_rows),
        "c_index_test": float(c_mu_test),
    }
    del model_rhuh
    torch.cuda.empty_cache()
    gc.collect()

    # Within-cohort 5-fold CV on combined cohort (sanity check)
    print("\n=== Within-cohort: train RHUH → test RHUH ===", flush=True)
    model_rhuh_in = train_survival_unet(rhuh_rows, seed=SEED)
    c_rhuh_in, _ = evaluate_survival(model_rhuh_in, rhuh_rows)
    print(f"  Test on RHUH (training set!): C-index = {c_rhuh_in:.4f}",
          flush=True)
    del model_rhuh_in
    torch.cuda.empty_cache()
    gc.collect()
    results["train_RHUH_test_RHUH"] = {
        "n_train": len(rhuh_rows), "n_test": len(rhuh_rows),
        "c_index_test": float(c_rhuh_in),
        "note": "within-training (overfit) C-index for comparison"
    }

    print("\n" + "=" * 60, flush=True)
    print("SURVIVAL FOUNDATION MODEL COMPARISON", flush=True)
    print("=" * 60, flush=True)
    print(f"  v201 train MU → test RHUH:  C = {results['train_MU_test_RHUH']['c_index_test']:.4f}  "
          f"(true held-out)", flush=True)
    print(f"  v201 train RHUH → test MU:  C = {results['train_RHUH_test_MU']['c_index_test']:.4f}  "
          f"(true held-out)", flush=True)
    print(f"  v201 train RHUH → test RHUH (overfit): C = "
          f"{results['train_RHUH_test_RHUH']['c_index_test']:.4f}",
          flush=True)
    print(f"\n  References:", flush=True)
    print(f"  Round 33 RHUH clinical-only Cox: C = 0.6664", flush=True)
    print(f"  Round 33 RHUH clinical + V_kernel Cox: C = 0.6618 (Δ=-0.005, p=0.53)",
          flush=True)
    print(f"  Round 35 RHUH n=13 clinical + λ + V_kernel: C = 0.8833 (preliminary, REFUTED)",
          flush=True)
    print(f"  Round 36 MU clinical-only Cox: C = 0.6011", flush=True)

    out = {
        "version": "v201",
        "experiment": ("Survival-supervised 3D U-Net (Cox PH loss) "
                       "trained on baseline mask + bimodal kernel"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "epochs": EPOCHS,
        "seed": SEED,
        "results": results,
        "references": {
            "round_33_RHUH_clinical_C": 0.6664,
            "round_33_RHUH_clinical_plus_Vkernel_C": 0.6618,
            "round_35_RHUH_n13_clinical_lambda_Vkernel_C": 0.8833,
            "round_36_MU_clinical_C": 0.6011,
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
