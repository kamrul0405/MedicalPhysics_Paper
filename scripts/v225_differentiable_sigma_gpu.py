"""v225: DIFFERENTIABLE-σ end-to-end model — round 50 (GPU).

Truly novel beyond-NMI method: make σ in the bimodal kernel a
LEARNABLE parameter via differentiable Gaussian filtering.

Round 47 v218 used FIXED σ ∈ {2,3,4,5}. Round 50 v225 makes K σ
values learnable: σ_k = softplus(θ_k) for k=1..K. The network
LEARNS its own optimal σ values from PFS labels.

Architecture:
  Input: binary mask (B, 1, D, H, W)
  Differentiable bimodal kernel layer:
    For k = 1..K:
      σ_k = softplus(θ_k)
      G_k(x) = Gaussian filter of mask with σ_k (differentiable)
      K_k(x) = max(mask, G_k / G_k.max())
      V_k = sum( (K_k >= 0.5) & ~mask )  [continuous via sigmoid]
  Concatenate V_1..V_K with clinical features → MLP classifier

Trains end-to-end on binary 365-d PFS. Compare to v218 fixed-σ
multi-σ logistic.

Outputs:
  Nature_project/05_results/v225_differentiable_sigma.json
"""
from __future__ import annotations

import gc
import json
import math
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v225_differentiable_sigma.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
EPOCHS = 60
LR = 5e-3
SEED = 42
N_KERNELS = 4  # K differentiable σ values


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


def gaussian_3d_kernel(sigma, kernel_size):
    """Build separable Gaussian kernel."""
    half = kernel_size // 2
    x = torch.arange(-half, half + 1,
                       dtype=torch.float32, device=sigma.device)
    g = torch.exp(-0.5 * (x / sigma) ** 2)
    g = g / g.sum()
    return g  # 1D Gaussian


def differentiable_gaussian_3d(volume, sigma, kernel_size=15):
    """Apply 3D Gaussian filter via separable convolution.
    volume: (B, 1, D, H, W); sigma: scalar tensor."""
    g = gaussian_3d_kernel(sigma, kernel_size)
    pad = kernel_size // 2
    # Conv along D
    g_d = g.view(1, 1, kernel_size, 1, 1)
    v = F.conv3d(volume, g_d, padding=(pad, 0, 0))
    # Conv along H
    g_h = g.view(1, 1, 1, kernel_size, 1)
    v = F.conv3d(v, g_h, padding=(0, pad, 0))
    # Conv along W
    g_w = g.view(1, 1, 1, 1, kernel_size)
    v = F.conv3d(v, g_w, padding=(0, 0, pad))
    return v


class DifferentiableKernelLayer(nn.Module):
    """K differentiable bimodal kernels with learnable σ.
    Output: K-dim feature per patient (V_k integrated)."""

    def __init__(self, n_kernels=N_KERNELS,
                  init_sigmas=(1.5, 2.5, 3.5, 5.0)):
        super().__init__()
        # Parameterize σ_k = softplus(θ_k) so σ > 0
        # Initialize so softplus(θ_k) ≈ init_sigmas[k]
        # softplus(x) = log(1 + exp(x)); inverse: log(exp(x) - 1)
        # for x > 0
        init_thetas = [math.log(math.exp(s) - 1)
                        for s in init_sigmas[:n_kernels]]
        self.theta = nn.Parameter(
            torch.tensor(init_thetas, dtype=torch.float32))

    def forward(self, mask, kernel_size=11):
        """mask: (B, 1, D, H, W) binary 0/1.
        Returns (B, K) integrated outgrowth-volume per σ."""
        sigmas = F.softplus(self.theta)
        B = mask.size(0)
        outputs = []
        for k in range(len(sigmas)):
            s = sigmas[k]
            # Gaussian-filter the mask
            G = differentiable_gaussian_3d(mask, s,
                                              kernel_size=
                                              kernel_size)
            # Normalize by max per-sample
            G_max = G.amax(dim=(2, 3, 4), keepdim=True
                              ).clamp(min=1e-6)
            G_norm = G / G_max
            # Bimodal kernel: max(mask, G_norm)
            K = torch.maximum(mask, G_norm)
            # V_k = soft count of voxels where K >= 0.5 AND
            # mask < 0.5
            # Use sigmoid for differentiability
            inside_kernel = torch.sigmoid(20.0 * (K - 0.5))
            outside_mask = torch.sigmoid(20.0 * (0.5 - mask))
            v_k = (inside_kernel
                    * outside_mask).sum(dim=(2, 3, 4))
            outputs.append(v_k)  # (B, 1)
        return torch.cat(outputs, dim=-1), sigmas


class DifferentiableSigmaModel(nn.Module):
    """Differentiable-σ kernel + clinical features + MLP."""

    def __init__(self, n_kernels=N_KERNELS, n_clin=3):
        super().__init__()
        self.kernel_layer = DifferentiableKernelLayer(
            n_kernels=n_kernels)
        self.head = nn.Sequential(
            nn.Linear(n_kernels + n_clin, 16), nn.GELU(),
            nn.Dropout(0.2),
            nn.Linear(16, 1))

    def forward(self, mask, clinical):
        v_k, sigmas = self.kernel_layer(mask)
        # Log-transform V_k for scale
        v_k_log = torch.log1p(v_k)
        # Standardize per-batch
        v_k_z = (v_k_log - v_k_log.mean(dim=0, keepdim=True)
                  ) / (v_k_log.std(dim=0, keepdim=True) + 1e-3)
        x = torch.cat([clinical, v_k_z], dim=-1)
        return self.head(x).squeeze(-1), sigmas


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": prog, "pfs_days": pfs_d,
            "idh1": idh1, "mgmt": mgmt,
        }
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None or c["progress"] is None
                or c["pfs_days"] is None):
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        data.append({"pid": pid, "mask": m, "y": int(y),
                     "age": c["age"], "idh1": c["idh1"],
                     "mgmt": c["mgmt"]})
    return data


def main():
    print("=" * 78, flush=True)
    print(f"v225 DIFFERENTIABLE-σ MODEL (round 50 GPU) "
          f"device={DEVICE}", flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    mu_data = load_mu_data()
    print(f"  MU labelled: n={len(mu_data)}", flush=True)
    y_all = np.array([d["y"] for d in mu_data], dtype=np.float32)
    folds = stratified_kfold(y_all, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}", flush=True)

    # Stack masks and clinical
    masks_np = np.stack([d["mask"] for d in mu_data]).astype(
        np.float32)[:, None, :, :, :]  # (N, 1, D, H, W)
    clin_np = np.stack(
        [[d["age"], d["idh1"], d["mgmt"]] for d in mu_data]
    ).astype(np.float32)
    # Standardize clinical (full-data stats; we'll re-standardize
    # per fold)
    print(f"  masks shape: {masks_np.shape}, clinical shape: "
          f"{clin_np.shape}", flush=True)

    # ---- 5-fold CV ----
    n = len(mu_data)
    oof_scores = np.zeros(n, dtype=np.float32)
    oof_assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    learned_sigmas_per_fold = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        n_pos = int(y_all[train_idx].sum())
        n_neg = len(train_idx) - n_pos
        if n_pos < 2 or n_neg < 2:
            continue
        masks_tr = masks_np[train_idx]
        masks_te = masks_np[test_idx]
        clin_tr_raw = clin_np[train_idx]
        clin_te_raw = clin_np[test_idx]
        # Standardize clinical
        clin_mean = clin_tr_raw.mean(axis=0)
        clin_std = clin_tr_raw.std(axis=0)
        clin_std[clin_std == 0] = 1
        clin_tr = (clin_tr_raw - clin_mean) / clin_std
        clin_te = (clin_te_raw - clin_mean) / clin_std
        ytr = y_all[train_idx]
        yte = y_all[test_idx]

        masks_tr_t = torch.from_numpy(masks_tr).to(DEVICE)
        masks_te_t = torch.from_numpy(masks_te).to(DEVICE)
        clin_tr_t = torch.from_numpy(
            clin_tr.astype(np.float32)).to(DEVICE)
        clin_te_t = torch.from_numpy(
            clin_te.astype(np.float32)).to(DEVICE)
        ytr_t = torch.from_numpy(ytr).to(DEVICE)

        torch.manual_seed(SEED + fold_i * 100)
        np.random.seed(SEED + fold_i * 100)
        model = DifferentiableSigmaModel(
            n_kernels=N_KERNELS, n_clin=3).to(DEVICE)
        opt = torch.optim.AdamW(model.parameters(), lr=LR,
                                 weight_decay=1e-3)
        pos_w = torch.tensor([n_neg / max(1, n_pos)],
                              dtype=torch.float32,
                              device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        bs = 4
        for ep in range(EPOCHS):
            model.train()
            perm = np.random.permutation(len(train_idx))
            ep_loss = 0.0
            for i in range(0, len(train_idx), bs):
                idx = perm[i:i+bs]
                logits, sigmas = model(masks_tr_t[idx],
                                          clin_tr_t[idx])
                loss = criterion(logits, ytr_t[idx])
                opt.zero_grad()
                loss.backward()
                opt.step()
                ep_loss += loss.item() * len(idx)
            if (ep + 1) % 15 == 0 or ep == 0:
                with torch.no_grad():
                    s_now = F.softplus(
                        model.kernel_layer.theta).cpu().numpy()
                print(f"    fold {fold_i+1} ep {ep+1}: loss="
                      f"{ep_loss/len(train_idx):.4f}, "
                      f"σ={s_now.tolist()}", flush=True)

        # Evaluate
        model.eval()
        with torch.no_grad():
            logits_te, sigmas_te = model(masks_te_t,
                                            clin_te_t)
            scores = logits_te.cpu().numpy()
        oof_scores[test_idx] = scores
        oof_assigned[test_idx] = True
        a = auroc(scores, yte)
        fold_aucs.append(float(a))
        s_final = sigmas_te.cpu().numpy()
        learned_sigmas_per_fold.append(s_final.tolist())
        print(f"  Fold {fold_i+1}: AUC={a:.4f}, learned σ="
              f"{s_final.tolist()}", flush=True)
        del model, opt, masks_tr_t, masks_te_t, clin_tr_t, \
            clin_te_t, ytr_t
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()

    pooled_auc = auroc(oof_scores[oof_assigned],
                        y_all[oof_assigned])
    print(f"\n  v225 differentiable-σ pooled OOF AUC = "
          f"{pooled_auc:.4f}", flush=True)
    print(f"  per-fold mean = {np.mean(fold_aucs):.4f}",
          flush=True)
    sigma_means = np.mean(learned_sigmas_per_fold, axis=0)
    sigma_stds = np.std(learned_sigmas_per_fold, axis=0)
    print(f"\n  Average learned σ across 5 folds:", flush=True)
    for k, (m, s) in enumerate(zip(sigma_means, sigma_stds)):
        print(f"    σ_{k+1}: {m:.3f} ± {s:.3f}", flush=True)

    # Compare to v218 fixed-σ multi-σ
    print(f"\n=== COMPARISON ===", flush=True)
    print(f"  v218 fixed-σ multi-σ logistic (σ=2,3,4,5) "
          f"in-sample AUC: 0.815", flush=True)
    print(f"  v223 multi-σ logistic 5-fold OOF AUC: 0.708",
          flush=True)
    print(f"  v225 DIFFERENTIABLE-σ end-to-end 5-fold OOF "
          f"AUC: {pooled_auc:.4f}", flush=True)

    out = {
        "version": "v225",
        "experiment": ("Differentiable-σ end-to-end model: "
                       "learnable σ in bimodal kernel"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "epochs_per_fold": EPOCHS,
        "n_kernels": N_KERNELS,
        "init_sigmas": [1.5, 2.5, 3.5, 5.0],
        "n_mu": len(mu_data),
        "fold_aucs": fold_aucs,
        "fold_mean_auc": float(np.mean(fold_aucs)),
        "fold_std_auc": float(np.std(fold_aucs)),
        "pooled_oof_auc": float(pooled_auc),
        "learned_sigmas_per_fold": learned_sigmas_per_fold,
        "learned_sigma_means_across_folds": [
            float(x) for x in sigma_means],
        "learned_sigma_stds_across_folds": [
            float(x) for x in sigma_stds],
        "comparison": {
            "v218_fixed_multi_sigma_in_sample": 0.815,
            "v223_multi_sigma_5fold_OOF": 0.708,
            "v225_differentiable_sigma_5fold_OOF":
                float(pooled_auc),
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
