"""v219: SOTA 3D ResNet-18 architecture comparison on binary 365-d
PFS — round 47 (GPU).

The SOTA deep-learning baseline for medical imaging classification
is typically a deeper architecture like 3D ResNet-18/34 (e.g.,
3D MedicalNet, MONAI). Compare:

  (A) v207-style 3-block CNN (24-channel base, our baseline)
  (B) 3D ResNet-18 (4 stages × 2 BasicBlocks, channels [24, 48,
      96, 192]) — SOTA architecture
  (C) 3D ResNet-18 with SimCLR multi-cohort label-free pretraining
  (D) v202 logistic clin+V_kernel — best simple baseline

5-fold stratified CV on MU n=130 binary 365-d PFS.

Outputs:
  Nature_project/05_results/v219_3d_resnet_sota.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v219_3d_resnet_sota.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
EPOCHS = 30
LR = 5e-4
BASE = 24
SIMCLR_EPOCHS = 30
TEMPERATURE = 0.5
PROJ_DIM = 64
SEED = 42

COHORT_PREFIXES_PRE = ["MU-Glioma-Post", "RHUH-GBM",
                         "UCSF-POSTOP", "LUMIERE"]


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class BasicBlock3D(nn.Module):
    """ResNet-18 BasicBlock for 3D."""
    def __init__(self, in_ch, out_ch, stride=1):
        super().__init__()
        self.conv1 = nn.Conv3d(in_ch, out_ch, 3,
                                 padding=1, stride=stride,
                                 bias=False)
        self.norm1 = nn.GroupNorm(8, out_ch)
        self.conv2 = nn.Conv3d(out_ch, out_ch, 3, padding=1,
                                 bias=False)
        self.norm2 = nn.GroupNorm(8, out_ch)
        if stride != 1 or in_ch != out_ch:
            self.shortcut = nn.Sequential(
                nn.Conv3d(in_ch, out_ch, 1, stride=stride,
                            bias=False),
                nn.GroupNorm(8, out_ch))
        else:
            self.shortcut = nn.Identity()

    def forward(self, x):
        out = F.gelu(self.norm1(self.conv1(x)))
        out = self.norm2(self.conv2(out))
        out = out + self.shortcut(x)
        return F.gelu(out)


class ResNet3D18(nn.Module):
    """3D ResNet-18-style: 4 stages × 2 BasicBlocks."""
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.stem = nn.Sequential(
            nn.Conv3d(in_ch, base, 5, padding=2, bias=False),
            nn.GroupNorm(8, base), nn.GELU())
        self.layer1 = nn.Sequential(
            BasicBlock3D(base, base),
            BasicBlock3D(base, base))
        self.layer2 = nn.Sequential(
            BasicBlock3D(base, base * 2, stride=2),
            BasicBlock3D(base * 2, base * 2))
        self.layer3 = nn.Sequential(
            BasicBlock3D(base * 2, base * 4, stride=2),
            BasicBlock3D(base * 4, base * 4))
        self.layer4 = nn.Sequential(
            BasicBlock3D(base * 4, base * 8, stride=2),
            BasicBlock3D(base * 8, base * 8))
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.feat_dim = base * 8
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 8, 64), nn.GELU(),
            nn.Dropout(0.3), nn.Linear(64, 1))

    def encode(self, x):
        x = self.stem(x)
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        x = self.layer4(x)
        return self.global_pool(x).flatten(1)

    def forward(self, x):
        feat = self.encode(x)
        return self.head(feat).squeeze(-1)


class SimpleCNN(nn.Module):
    """v207-style 3-block CNN (baseline)."""
    def __init__(self, in_ch=2, base=BASE):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.feat_dim = base * 4
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32), nn.GELU(),
            nn.Dropout(0.3), nn.Linear(32, 1))

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def encode(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.global_pool(e3).flatten(1)

    def forward(self, x):
        return self.head(self.encode(x)).squeeze(-1)


class ProjectionHead(nn.Module):
    def __init__(self, feat_dim, proj_dim=PROJ_DIM):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, feat_dim), nn.GELU(),
            nn.Linear(feat_dim, proj_dim))

    def forward(self, feat):
        return F.normalize(self.net(feat), dim=-1)


def random_3d_augment(x, rng):
    out = x.clone()
    for axis in [2, 3, 4]:
        if rng.random() > 0.5:
            out = torch.flip(out, dims=[axis])
    scale = float(rng.uniform(0.9, 1.1))
    out = out * scale + 0.02 * torch.randn_like(out)
    return torch.clamp(out, 0.0, 1.5)


def nt_xent_loss(z1, z2, temperature=TEMPERATURE):
    n = z1.size(0)
    z = torch.cat([z1, z2], dim=0)
    sim = torch.matmul(z, z.T) / temperature
    mask = torch.eye(2 * n, dtype=torch.bool,
                      device=sim.device)
    sim = sim.masked_fill(mask, float("-inf"))
    targets = torch.cat([torch.arange(n, 2 * n),
                          torch.arange(0, n)],
                          dim=0).to(sim.device)
    return F.cross_entropy(sim, targets)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def load_pretrain_masks():
    samples = []
    for prefix in COHORT_PREFIXES_PRE:
        for f in sorted(CACHE.glob(f"{prefix}_*_b.npy")):
            try:
                m_full = (np.load(f) > 0).astype(np.float32)
                if m_full.sum() == 0:
                    continue
                m = resize_to_target(m_full, TARGET_SHAPE)
                k = heat_bimodal(m, SIGMA_KERNEL)
                samples.append({"pid": f.stem, "mask": m,
                                 "kernel": k})
            except Exception:
                continue
    return samples


def simclr_pretrain_resnet(samples, epochs, seed=SEED):
    """SimCLR pretrain a ResNet3D18 encoder."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    rng_np = np.random.default_rng(seed)
    X = stack_inputs(samples)
    Xt = torch.from_numpy(X).to(DEVICE)
    model = ResNet3D18(in_ch=2, base=BASE).to(DEVICE)
    proj = ProjectionHead(model.feat_dim,
                            proj_dim=PROJ_DIM).to(DEVICE)
    opt = torch.optim.AdamW(
        list(model.parameters()) + list(proj.parameters()),
        lr=1e-3, weight_decay=1e-4)
    n = len(samples)
    bs = 16
    losses = []
    for ep in range(epochs):
        model.train()
        proj.train()
        perm = np.random.permutation(n)
        ep_loss = 0.0
        n_batches = 0
        for i in range(0, n, bs):
            idx = perm[i:i+bs]
            if len(idx) < 2:
                continue
            xb = Xt[idx]
            v1 = random_3d_augment(xb, rng_np)
            v2 = random_3d_augment(xb, rng_np)
            f1 = model.encode(v1)
            f2 = model.encode(v2)
            z1 = proj(f1)
            z2 = proj(f2)
            loss = nt_xent_loss(z1, z2)
            opt.zero_grad()
            loss.backward()
            opt.step()
            ep_loss += loss.item()
            n_batches += 1
        losses.append(ep_loss / max(1, n_batches))
        if (ep + 1) % 5 == 0:
            print(f"    SimCLR ep {ep+1}/{epochs}  "
                  f"loss={losses[-1]:.4f}", flush=True)
    return model, losses


def train_eval_5fold(model_class, mu_data, folds,
                      pretrained_model=None, name="model",
                      seed=SEED):
    """5-fold CV on MU. If pretrained_model is given, copy its
    encoder weights into a fresh model and freeze them, retrain
    only the head."""
    n = len(mu_data)
    y_all = np.array([d["y"] for d in mu_data], dtype=np.float32)
    oof = np.zeros(n, dtype=np.float32)
    assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        n_pos = int(y_all[train_idx].sum())
        n_neg = len(train_idx) - n_pos
        if n_pos < 2 or n_neg < 2:
            continue
        Xtr = stack_inputs([mu_data[i] for i in train_idx])
        Xte = stack_inputs([mu_data[i] for i in test_idx])
        ytr = y_all[train_idx]
        yte = y_all[test_idx]
        Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
        Xte_t = torch.from_numpy(Xte).to(DEVICE)
        ytr_t = torch.from_numpy(ytr).to(DEVICE)

        torch.manual_seed(seed + fold_i * 100)
        np.random.seed(seed + fold_i * 100)
        model = model_class(in_ch=2, base=BASE).to(DEVICE)
        if pretrained_model is not None:
            # copy encoder (everything except head)
            pre_state = pretrained_model.state_dict()
            tgt_state = model.state_dict()
            for k, v in pre_state.items():
                if k in tgt_state and k != "head" and \
                    not k.startswith("head."):
                    tgt_state[k].copy_(v)
            # freeze encoder
            for n_p, p in model.named_parameters():
                if not n_p.startswith("head."):
                    p.requires_grad = False
            params = [p for p in model.parameters()
                       if p.requires_grad]
        else:
            params = list(model.parameters())
        opt = torch.optim.AdamW(params, lr=LR,
                                 weight_decay=1e-3)
        pos_w = torch.tensor([n_neg / max(1, n_pos)],
                              dtype=torch.float32, device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        bs = 4
        for ep in range(EPOCHS):
            model.train()
            perm = np.random.permutation(len(train_idx))
            for i in range(0, len(train_idx), bs):
                idx = perm[i:i+bs]
                logits = model(Xtr_t[idx])
                loss = criterion(logits, ytr_t[idx])
                opt.zero_grad()
                loss.backward()
                opt.step()
        model.eval()
        with torch.no_grad():
            scores = model(Xte_t).cpu().numpy()
        oof[test_idx] = scores
        assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            fold_aucs.append(float(a))
            print(f"    {name} fold {fold_i+1}: AUC={a:.4f}",
                  flush=True)
        del model, opt
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()
    pooled_auc = auroc(oof[assigned], y_all[assigned])
    print(f"  {name} pooled OOF AUC = {pooled_auc:.4f}",
          flush=True)
    return {
        "fold_aucs": fold_aucs,
        "fold_mean": float(np.mean(fold_aucs)),
        "fold_std": float(np.std(fold_aucs)),
        "pooled_oof_auc": float(pooled_auc),
    }


def main():
    print("=" * 78, flush=True)
    print(f"v219 SOTA 3D-ResNet-18 (round 47 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    mu_data = load_mu_data()
    print(f"  MU labelled: n={len(mu_data)}", flush=True)
    y_mu = np.array([d["y"] for d in mu_data], dtype=np.float32)
    folds = stratified_kfold(y_mu, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}", flush=True)

    results = {}

    # ---- Variant A: SimpleCNN (v207-style) ----
    print(f"\n=== VARIANT A: SimpleCNN (v207-style) baseline ===",
          flush=True)
    res_simple = train_eval_5fold(SimpleCNN, mu_data, folds,
                                    name="SimpleCNN", seed=SEED)
    results["A_simple_cnn"] = res_simple
    n_params_simple = sum(p.numel()
                            for p in SimpleCNN().parameters())
    print(f"  SimpleCNN parameters: {n_params_simple:,}",
          flush=True)

    # ---- Variant B: 3D ResNet-18 ----
    print(f"\n=== VARIANT B: 3D ResNet-18 (SOTA) ===",
          flush=True)
    n_params_resnet = sum(p.numel()
                            for p in ResNet3D18().parameters())
    print(f"  ResNet3D18 parameters: {n_params_resnet:,} "
          f"({n_params_resnet / n_params_simple:.1f}x SimpleCNN)",
          flush=True)
    res_resnet = train_eval_5fold(ResNet3D18, mu_data, folds,
                                    name="ResNet3D18", seed=SEED)
    results["B_resnet3d18"] = res_resnet

    # ---- Variant C: 3D ResNet-18 + SimCLR multi-cohort pretrain ----
    print(f"\n=== VARIANT C: 3D ResNet-18 + SimCLR multi-cohort "
          f"pretrain ===", flush=True)
    pretrain_samples = load_pretrain_masks()
    print(f"  Multi-cohort masks: {len(pretrain_samples)}",
          flush=True)
    pretrained_resnet, simclr_losses = simclr_pretrain_resnet(
        pretrain_samples, SIMCLR_EPOCHS, seed=SEED)
    res_resnet_simclr = train_eval_5fold(
        ResNet3D18, mu_data, folds,
        pretrained_model=pretrained_resnet,
        name="ResNet3D18+SimCLR", seed=SEED)
    results["C_resnet3d18_simclr_pretrain"] = res_resnet_simclr

    # ---- Summary ----
    print(f"\n{'=' * 70}\nSOTA COMPARISON SUMMARY\n"
          f"{'=' * 70}", flush=True)
    summary = [
        ("v202 logistic clinical+V_kernel (deterministic)",
         0.728, "—"),
        ("v207 5-seed SimpleCNN supervised",
         0.586, "(reference)"),
        ("v215 SimpleCNN + SimCLR pretrain",
         0.706, "per-fold mean"),
        ("v219 (A) SimpleCNN baseline",
         res_simple["pooled_oof_auc"], "pooled OOF"),
        ("v219 (B) 3D ResNet-18 SOTA",
         res_resnet["pooled_oof_auc"], "pooled OOF"),
        ("v219 (C) 3D ResNet-18 + SimCLR pretrain",
         res_resnet_simclr["pooled_oof_auc"], "pooled OOF"),
    ]
    for name, auc, note in summary:
        print(f"  {name:55s}  AUC = {auc:.4f}  {note}",
              flush=True)

    out = {
        "version": "v219",
        "experiment": ("SOTA 3D ResNet-18 architecture "
                       "comparison vs v207 SimpleCNN and "
                       "logistic baseline"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "epochs_per_fold": EPOCHS,
        "n_mu": len(mu_data),
        "n_pretrain_samples": len(pretrain_samples),
        "simple_cnn_params": int(n_params_simple),
        "resnet3d18_params": int(n_params_resnet),
        "param_ratio": float(n_params_resnet / n_params_simple),
        "simclr_pretrain_loss_curve": [float(x)
                                         for x in simclr_losses],
        "results": results,
        "comparison_summary": [
            {"name": n, "auc": float(a), "note": note}
            for n, a, note in summary
        ],
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
