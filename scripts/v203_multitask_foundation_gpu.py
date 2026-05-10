"""v203: Multi-task foundation model (outgrowth + survival) — round 39 (GPU).

Senior-Nature-reviewer-driven experiment. Round 38 v201 showed that
training a 3D U-Net with Cox loss alone fails cross-cohort (C ~ 0.45).
But the foundation model trained for OUTGROWTH (round 22) generalizes
across cohorts (round 27 established kernel-only also generalizes).

Hypothesis: training with MULTI-TASK loss (outgrowth + survival)
might regularize the survival head via the outgrowth-task supervision,
enabling cross-cohort generalization.

Architecture:
  Shared 3D U-Net encoder
    -> Outgrowth head (decoder, voxel-wise sigmoid)
    -> Survival head (global pool + linear, scalar risk)
  Loss = alpha * focal_dice_loss(outgrowth) + beta * cox_loss(risk)

Training: pooled RHUH + MU + the 5 trained cohorts (UCSF, MU, RHUH,
LUMIERE, PROTEAS) where outgrowth labels are available.
For survival head: only RHUH + MU patients have OS data.

Method: train multi-task model; evaluate cross-cohort C-index on
RHUH and MU LOCO. Compare to round 38 v201 (single-task survival,
C ~ 0.45-0.49).

If multi-task helps → outgrowth supervision regularizes survival
generalization → publishable joint biomarker recipe.
If not → another honest negative confirming kernel/UODSL not
clinically prognostic.

Outputs:
  Nature_project/05_results/v203_multitask_foundation.json
"""
from __future__ import annotations

import csv
import gc
import json
import re
import shutil
import tempfile
import time
import warnings
import zipfile
from pathlib import Path

import nibabel as nib
import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_RHUH = Path(r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
CLINICAL_MU = Path(r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
DATA_ZIP = Path(r"C:\Users\kamru\Downloads\Datasets\PKG - PROTEAS-brain-mets-zenodo-17253793.zip")
OUT_JSON = RESULTS / "v203_multitask_foundation.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_BROAD = 7.0
TARGET_SHAPE = (16, 48, 48)
EPOCHS = 30
LR = 5e-4
SEED = 42
ALPHA_OUTGROWTH = 1.0
BETA_SURVIVAL = 0.5


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors, order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class MultiTaskUNet(nn.Module):
    """3D U-Net encoder + outgrowth decoder + survival head."""
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.dec2 = self._block(base * 4 + base * 2, base * 2)
        self.dec1 = self._block(base * 2 + base, base)
        self.out = nn.Conv3d(base, 1, 1)
        self.pool = nn.MaxPool3d(2)
        self.up = nn.Upsample(scale_factor=2, mode="trilinear",
                                align_corners=False)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.surv_head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32),
            nn.GELU(),
            nn.Linear(32, 1),
        )

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        # Outgrowth decoder
        d2 = self.dec2(torch.cat([self.up(e3), e2], dim=1))
        d1 = self.dec1(torch.cat([self.up(d2), e1], dim=1))
        outgrowth_logits = self.out(d1)
        # Survival head from bottleneck e3
        risk = self.surv_head(self.global_pool(e3)).squeeze(-1)
        return outgrowth_logits, risk


def focal_dice_loss(logits, target, alpha=0.95, gamma=2.0, smooth=1e-5):
    p = torch.sigmoid(logits)
    p_t = p * target + (1 - p) * (1 - target)
    alpha_t = alpha * target + (1 - alpha) * (1 - target)
    focal = -alpha_t * (1 - p_t) ** gamma * torch.log(
        p_t.clamp(1e-7, 1 - 1e-7))
    return focal.mean() + (1 - (2 * (p * target).sum() + smooth) /
                              (p.sum() + target.sum() + smooth))


def cox_loss(risk, time, event):
    order = torch.argsort(time, descending=True)
    risk_sorted = risk[order]
    event_sorted = event[order]
    max_risk = torch.cummax(risk_sorted, dim=0)[0]
    log_cum = max_risk + torch.log(torch.cumsum(
        torch.exp(risk_sorted - max_risk), dim=0))
    valid = event_sorted == 1
    if valid.sum() == 0:
        return torch.tensor(0.0, requires_grad=True, device=risk.device)
    log_pl = (risk_sorted[valid] - log_cum[valid]).sum()
    return -log_pl / valid.sum()


def concordance(times, events, risk_scores):
    times = np.asarray(times)
    events = np.asarray(events)
    risk_scores = np.asarray(risk_scores)
    n = len(times)
    n_pairs = 0
    n_concordant = 0
    n_ties = 0
    for i in range(n):
        for j in range(i+1, n):
            if events[i] == 0 and events[j] == 0:
                continue
            if times[i] < times[j]:
                if events[i] != 1:
                    continue
                n_pairs += 1
                if risk_scores[i] > risk_scores[j]:
                    n_concordant += 1
                elif risk_scores[i] == risk_scores[j]:
                    n_ties += 1
            elif times[j] < times[i]:
                if events[j] != 1:
                    continue
                n_pairs += 1
                if risk_scores[j] > risk_scores[i]:
                    n_concordant += 1
                elif risk_scores[i] == risk_scores[j]:
                    n_ties += 1
    if n_pairs == 0:
        return float("nan")
    return (n_concordant + 0.5 * n_ties) / n_pairs


def load_glioma_cohort(cohort):
    rows = []
    for f in sorted(CACHE.glob(f"{cohort}_*_b.npy")):
        pid = f.stem.replace(f"{cohort}_", "").replace("_b", "")
        fr = f.parent / f.name.replace("_b.npy", "_r.npy")
        if not fr.exists():
            continue
        m = (np.load(f) > 0).astype(np.float32)
        t = (np.load(fr) > 0).astype(np.float32)
        if m.sum() == 0 or t.sum() == 0:
            continue
        outgrowth = (t.astype(bool) & ~m.astype(bool)).astype(np.float32)
        rows.append({"pid": pid, "cohort": cohort,
                     "mask": m, "outgrowth": outgrowth,
                     "kernel": heat_bimodal(m, SIGMA_BROAD)})
    return rows


def load_rhuh_with_os():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        os_col = next((c for c in cols if "Overall survival" in c
                         and "day" in c.lower()), None)
        cens_col = next((c for c in cols if "censor" in c.lower()), None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                os_d = float(row[os_col]) if row[os_col] else None
            except (ValueError, TypeError):
                continue
            event = (0 if row[cens_col].strip().lower() == "yes" else 1)
            if os_d is not None:
                clinical[pid] = {"os": os_d, "event": event}

    rows = load_glioma_cohort("RHUH-GBM")
    enriched = []
    for r in rows:
        if r["pid"] in clinical:
            r["os"] = clinical[r["pid"]]["os"]
            r["event"] = clinical[r["pid"]]["event"]
            enriched.append(r)
    return enriched


def load_mu_with_os():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    death_col = header.index("Overall Survival (Death)")
    os_col = header.index("Number of days from Diagnosis to death (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            event = int(row[death_col]) if row[death_col] is not None else None
            os_d = float(row[os_col]) if row[os_col] is not None else None
        except (ValueError, TypeError):
            continue
        if event is not None and os_d is not None:
            clinical[str(pid)] = {"os": os_d, "event": event}

    rows = load_glioma_cohort("MU-Glioma-Post")
    enriched = []
    for r in rows:
        if r["pid"] in clinical:
            r["os"] = clinical[r["pid"]]["os"]
            r["event"] = clinical[r["pid"]]["event"]
            enriched.append(r)
    return enriched


def train_multitask(train_data_outgrowth, train_data_survival, seed=SEED):
    """Co-train: outgrowth on full set; survival on subset with OS."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    # Outgrowth tensors
    n_o = len(train_data_outgrowth)
    Xo = np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                    for d in train_data_outgrowth]).astype(np.float32)
    Yo = np.stack([d["outgrowth"]
                    for d in train_data_outgrowth]).astype(np.float32)[:, None]
    Xo_t = torch.from_numpy(Xo).to(DEVICE)
    Yo_t = torch.from_numpy(Yo).to(DEVICE)
    # Survival tensors
    n_s = len(train_data_survival)
    Xs = np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                    for d in train_data_survival]).astype(np.float32)
    times_s = np.array([d["os"] for d in train_data_survival],
                         dtype=np.float32)
    events_s = np.array([d["event"] for d in train_data_survival],
                          dtype=np.float32)
    Xs_t = torch.from_numpy(Xs).to(DEVICE)
    times_s_t = torch.from_numpy(times_s).to(DEVICE)
    events_s_t = torch.from_numpy(events_s).to(DEVICE)

    model = MultiTaskUNet(in_ch=2, base=24).to(DEVICE)
    opt = torch.optim.AdamW(model.parameters(), lr=LR, weight_decay=1e-3)

    for ep in range(EPOCHS):
        model.train()
        # Outgrowth pass: minibatch through full set
        perm = np.random.permutation(n_o)
        bs = 4
        loss_out_total = 0.0
        for i in range(0, n_o, bs):
            idx = perm[i:i+bs]
            xb = Xo_t[idx]
            yb = Yo_t[idx]
            out_logits, _ = model(xb)
            loss_out = focal_dice_loss(out_logits, yb)
            opt.zero_grad()
            (ALPHA_OUTGROWTH * loss_out).backward()
            opt.step()
            loss_out_total += loss_out.item() * len(idx)
        loss_out_total /= n_o

        # Survival pass: full-batch Cox loss
        _, risk = model(Xs_t)
        loss_surv = cox_loss(risk, times_s_t, events_s_t)
        opt.zero_grad()
        (BETA_SURVIVAL * loss_surv).backward()
        opt.step()

        if (ep + 1) % 5 == 0:
            with torch.no_grad():
                model.eval()
                _, r = model(Xs_t)
                c_train = concordance(times_s, events_s, r.cpu().numpy())
            print(f"    ep {ep+1}/{EPOCHS}  out_loss={loss_out_total:.4f}  "
                  f"surv_loss={loss_surv.item():.4f}  train C={c_train:.4f}",
                  flush=True)
    return model


def evaluate_survival_head(model, test_rows):
    model.eval()
    X = np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                   for d in test_rows]).astype(np.float32)
    times = np.array([d["os"] for d in test_rows], dtype=float)
    events = np.array([d["event"] for d in test_rows], dtype=float)
    with torch.no_grad():
        Xt = torch.from_numpy(X).to(DEVICE)
        _, risks = model(Xt)
        risks = risks.cpu().numpy()
    return concordance(times, events, risks)


def main():
    print("=" * 78, flush=True)
    print("v203 MULTI-TASK FOUNDATION (round 39 GPU)", flush=True)
    print(f"  Device: {DEVICE}", flush=True)
    print(f"  Loss weights: alpha_out={ALPHA_OUTGROWTH}, beta_surv={BETA_SURVIVAL}",
          flush=True)
    print("=" * 78, flush=True)

    # Outgrowth supervision: all trained glioma cohorts (UCSF, MU, RHUH,
    # LUMIERE) — use cached PROTEAS only if needed; here keep simple.
    print("\nLoading outgrowth supervision cohorts...", flush=True)
    outgrowth_train = []
    for c in ["UCSF-POSTOP", "MU-Glioma-Post", "RHUH-GBM", "LUMIERE"]:
        rows = load_glioma_cohort(c)
        outgrowth_train.extend(rows)
        print(f"  {c}: {len(rows)} patients", flush=True)
    print(f"  Total outgrowth-supervised: {len(outgrowth_train)} patients",
          flush=True)

    # Survival supervision: RHUH + MU with OS
    print("\nLoading survival supervision cohorts...", flush=True)
    rhuh_with_os = load_rhuh_with_os()
    mu_with_os = load_mu_with_os()
    print(f"  RHUH with OS: {len(rhuh_with_os)} patients", flush=True)
    print(f"  MU with OS: {len(mu_with_os)} patients", flush=True)

    # ---------- LOCO 1: train MU+others, test RHUH ----------
    print("\n=== LOCO 1: Train (outgrowth: all 4 cohorts; survival: MU only) "
          "→ Test RHUH ===", flush=True)
    surv_train_1 = mu_with_os
    out_train_1 = [r for r in outgrowth_train if r["cohort"] != "RHUH-GBM"]
    print(f"  Outgrowth train: {len(out_train_1)}; "
          f"Survival train: {len(surv_train_1)}", flush=True)
    t0 = time.time()
    model_1 = train_multitask(out_train_1, surv_train_1, seed=SEED)
    print(f"  Train took {time.time()-t0:.0f}s", flush=True)
    c_rhuh = evaluate_survival_head(model_1, rhuh_with_os)
    print(f"  Test on RHUH (held-out for survival): C = {c_rhuh:.4f}",
          flush=True)
    del model_1
    torch.cuda.empty_cache()
    gc.collect()

    # ---------- LOCO 2: train RHUH+others, test MU ----------
    print("\n=== LOCO 2: Train (outgrowth: all 4 cohorts; survival: RHUH only) "
          "→ Test MU ===", flush=True)
    surv_train_2 = rhuh_with_os
    out_train_2 = [r for r in outgrowth_train if r["cohort"] != "MU-Glioma-Post"]
    print(f"  Outgrowth train: {len(out_train_2)}; "
          f"Survival train: {len(surv_train_2)}", flush=True)
    t0 = time.time()
    model_2 = train_multitask(out_train_2, surv_train_2, seed=SEED)
    print(f"  Train took {time.time()-t0:.0f}s", flush=True)
    c_mu = evaluate_survival_head(model_2, mu_with_os)
    print(f"  Test on MU (held-out for survival): C = {c_mu:.4f}",
          flush=True)
    del model_2
    torch.cuda.empty_cache()
    gc.collect()

    print("\n" + "=" * 60, flush=True)
    print("MULTI-TASK FOUNDATION SURVIVAL COMPARISON", flush=True)
    print("=" * 60, flush=True)
    print(f"  v203 multi-task train MU → test RHUH: C = {c_rhuh:.4f}",
          flush=True)
    print(f"  v203 multi-task train RHUH → test MU: C = {c_mu:.4f}",
          flush=True)
    print(f"\n  Reference (round 38 v201 single-task):", flush=True)
    print(f"    train MU → test RHUH: C = 0.4516", flush=True)
    print(f"    train RHUH → test MU: C = 0.4897", flush=True)
    print(f"\n  Reference (clinical-only Cox):", flush=True)
    print(f"    RHUH: C = 0.6664; MU: C = 0.6011", flush=True)

    out = {
        "version": "v203",
        "experiment": ("Multi-task foundation model (outgrowth + survival) "
                       "with shared 3D U-Net encoder"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "epochs": EPOCHS,
        "alpha_outgrowth": ALPHA_OUTGROWTH,
        "beta_survival": BETA_SURVIVAL,
        "loco_1_train_MU_test_RHUH": {
            "n_outgrowth_train": len(out_train_1),
            "n_survival_train": len(surv_train_1),
            "n_test_RHUH": len(rhuh_with_os),
            "c_index_test_RHUH": float(c_rhuh),
        },
        "loco_2_train_RHUH_test_MU": {
            "n_outgrowth_train": len(out_train_2),
            "n_survival_train": len(surv_train_2),
            "n_test_MU": len(mu_with_os),
            "c_index_test_MU": float(c_mu),
        },
        "references": {
            "v201_single_task_train_MU_test_RHUH": 0.4516,
            "v201_single_task_train_RHUH_test_MU": 0.4897,
            "clinical_cox_RHUH": 0.6664,
            "clinical_cox_MU": 0.6011,
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
