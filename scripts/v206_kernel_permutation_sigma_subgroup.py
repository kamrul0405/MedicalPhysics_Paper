"""v206: Permutation test + sigma-sweep + IDH/MGMT subgroup analysis on
the binary 365-day PFS task — round 41 (CPU).

Round 40 v204 established the kernel-as-PFS-screen claim with 4
evidence levels (window, NB, calibration, irreducibility). For
Nature/Lancet, three additional pieces of empirical grounding are
mandatory:

  1. PERMUTATION TEST: is the +0.108 Delta AUC signal random?
     Shuffle V_kernel column 1000x, recompute Delta AUC. If the
     observed Delta is below the 95th percentile of the null,
     the signal is real.
  2. SIGMA SWEEP: was sigma=3 cherry-picked? Sweep sigma in
     {1, 2, 3, 4, 5, 7, 10}; report Delta AUC + bootstrap CI per
     sigma. If the curve is broad (multiple sigmas significant),
     the choice is robust.
  3. SUBGROUP ANALYSIS: does the kernel work across IDH/MGMT
     subgroups (regulatory requirement)? Per-subgroup logistic
     Delta AUC + 95% CI.

If all three confirm, the kernel-as-PFS-screen claim is bullet-proof
for Nature/Lancet.

Outputs:
  Nature_project/05_results/v206_kernel_permutation_sigma_subgroup.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_XLSX = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v206_kernel_permutation_sigma_subgroup.json"

HORIZON = 365
SIGMA_LIST = [1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]
SIGMA_PRIMARY = 3.0
N_PERMUTATIONS = 1000
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def build_X(rows, feats):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0)
    stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return np.column_stack([np.ones(len(arr_z)), arr_z])


def fit_eval_pair(rows_lab, feats_clin, feats_clin_vk):
    cc = [l for l in rows_lab if all(
        l[f] is not None and not (isinstance(l[f], float)
                                  and np.isnan(l[f]))
        for f in feats_clin_vk)]
    if len(cc) < 20:
        return None
    y = np.array([l["y"] for l in cc], dtype=float)
    n_pos = int(y.sum())
    if n_pos < 5 or len(y) - n_pos < 5:
        return None
    X_clin = build_X(cc, feats_clin)
    X_full = build_X(cc, feats_clin_vk)
    bc = logistic_fit(X_clin, y)
    bf = logistic_fit(X_full, y)
    return {
        "n": len(cc), "n_pos": n_pos,
        "auc_clin": auroc(X_clin @ bc, y),
        "auc_full": auroc(X_full @ bf, y),
        "y": y,
    }


def label_binary(rows, X):
    out = []
    for r in rows:
        pfs = r["pfs_days"]
        prog = r["progress"]
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def main():
    print("=" * 78, flush=True)
    print("v206 PERMUTATION + SIGMA-SWEEP + SUBGROUP (round 41 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # ---- Load clinical ----
    wb = openpyxl.load_workbook(CLINICAL_XLSX, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress, "pfs_days": pfs_d,
            "idh1": idh1, "mgmt": mgmt,
        }
    print(f"  Loaded clinical for {len(clinical)} MU patients",
          flush=True)

    # ---- Compute V_kernel for ALL sigmas per patient ----
    print("\nComputing V_kernel across sigma sweep per patient...",
          flush=True)
    rows = []
    pids = sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy"))
    for f in pids:
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        v_k_per_sigma = {f"v_kernel_s{int(sg)}":
                          kernel_outgrowth_volume(m, sg)
                          for sg in SIGMA_LIST}
        rows.append({
            "pid": pid,
            **v_k_per_sigma,
            "baseline_volume": int(m.sum()),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    print(f"  {len(rows)} MU patients", flush=True)
    base = [r for r in rows
            if r["pfs_days"] is not None
            and r["progress"] is not None
            and r["age"] is not None
            and r["idh1"] is not None
            and r["mgmt"] is not None]
    labelled = label_binary(base, HORIZON)
    n_pos = sum(1 for l in labelled if l["y"] == 1)
    n_neg = len(labelled) - n_pos
    print(f"  At H={HORIZON}: {len(labelled)} labelled "
          f"({n_pos} pos, {n_neg} neg)", flush=True)

    # ============================================================
    # PART 1: SIGMA-SWEEP + bootstrap CI per sigma
    # ============================================================
    print("\n=== PART 1: sigma sweep ===", flush=True)
    feats_clin = ["age", "idh1", "mgmt"]
    sigma_results = {}
    for sg in SIGMA_LIST:
        feat_vk = f"v_kernel_s{int(sg)}"
        feats_full = feats_clin + [feat_vk]
        full = fit_eval_pair(labelled, feats_clin, feats_full)
        if full is None:
            continue
        delta_pt = full["auc_full"] - full["auc_clin"]

        # Bootstrap CI on Delta AUC at this sigma
        deltas = []
        n = len(labelled)
        for _ in range(N_BOOTSTRAPS):
            idx = rng.integers(0, n, size=n)
            br = [labelled[i] for i in idx]
            res = fit_eval_pair(br, feats_clin, feats_full)
            if res is not None:
                deltas.append(res["auc_full"] - res["auc_clin"])
        deltas = np.array(deltas)
        sigma_results[int(sg)] = {
            "auc_clin": float(full["auc_clin"]),
            "auc_full": float(full["auc_full"]),
            "delta_AUC_point": float(delta_pt),
            "delta_AUC_bootstrap_mean": float(deltas.mean()),
            "delta_AUC_95_CI": [float(np.percentile(deltas, 2.5)),
                                  float(np.percentile(deltas, 97.5))],
            "p_one_sided": float((deltas <= 0).mean()),
        }
        print(f"  sigma={sg:.0f}: AUC_clin={full['auc_clin']:.4f}, "
              f"AUC_full={full['auc_full']:.4f}, "
              f"Delta={delta_pt:+.4f} "
              f"[CI {sigma_results[int(sg)]['delta_AUC_95_CI'][0]:+.4f}, "
              f"{sigma_results[int(sg)]['delta_AUC_95_CI'][1]:+.4f}], "
              f"P(<=0)={sigma_results[int(sg)]['p_one_sided']:.3f}",
              flush=True)

    # ============================================================
    # PART 2: PERMUTATION TEST at sigma=3 (PRIMARY)
    # ============================================================
    print(f"\n=== PART 2: permutation test at sigma={SIGMA_PRIMARY} "
          f"({N_PERMUTATIONS} shuffles) ===",
          flush=True)
    feat_vk_primary = f"v_kernel_s{int(SIGMA_PRIMARY)}"
    feats_full_primary = feats_clin + [feat_vk_primary]
    full_primary = fit_eval_pair(labelled, feats_clin,
                                   feats_full_primary)
    delta_observed = full_primary["auc_full"] - full_primary["auc_clin"]
    print(f"  Observed Delta AUC = {delta_observed:+.4f}", flush=True)

    # Build complete-case sample
    cc = [l for l in labelled if all(
        l[f] is not None and not (isinstance(l[f], float)
                                  and np.isnan(l[f]))
        for f in feats_full_primary)]
    null_deltas = []
    for it in range(N_PERMUTATIONS):
        # Shuffle the V_kernel column independently
        vk_vals = [c[feat_vk_primary] for c in cc]
        perm_idx = rng.permutation(len(vk_vals))
        cc_perm = [{**c, feat_vk_primary: vk_vals[perm_idx[i]]}
                    for i, c in enumerate(cc)]
        res = fit_eval_pair(cc_perm, feats_clin, feats_full_primary)
        if res is not None:
            null_deltas.append(res["auc_full"] - res["auc_clin"])
        if (it + 1) % 200 == 0:
            print(f"    {it+1}/{N_PERMUTATIONS} done", flush=True)
    null_deltas = np.array(null_deltas)
    p_perm = float((null_deltas >= delta_observed).mean())
    null_p95 = float(np.percentile(null_deltas, 95))
    null_p99 = float(np.percentile(null_deltas, 99))
    print(f"  Null Delta AUC: mean={null_deltas.mean():+.4f}, "
          f"95%={null_p95:+.4f}, 99%={null_p99:+.4f}", flush=True)
    print(f"  Permutation p-value (one-sided): {p_perm:.4f} "
          f"({int(p_perm * N_PERMUTATIONS)} of "
          f"{len(null_deltas)} null shuffles >= observed)",
          flush=True)

    # ============================================================
    # PART 3: SUBGROUP ANALYSIS (IDH, MGMT)
    # ============================================================
    print(f"\n=== PART 3: subgroup analysis at sigma={SIGMA_PRIMARY} ===",
          flush=True)
    subgroup_results = {}
    for sub_name, key, vals in [
        ("IDH-WT", "idh1", [0]),
        ("IDH-mut", "idh1", [1]),
        ("MGMT-unmeth", "mgmt", [0]),
        ("MGMT-meth", "mgmt", [1]),
    ]:
        sub_lab = [l for l in labelled if l[key] in vals]
        n_sub = len(sub_lab)
        n_pos_sub = sum(1 for l in sub_lab if l["y"] == 1)
        n_neg_sub = n_sub - n_pos_sub
        if n_sub < 20 or n_pos_sub < 5 or n_neg_sub < 5:
            print(f"  {sub_name}: n={n_sub}, "
                  f"{n_pos_sub} pos, {n_neg_sub} neg "
                  f"-> insufficient, skipped",
                  flush=True)
            subgroup_results[sub_name] = {
                "n": n_sub, "n_pos": n_pos_sub,
                "n_neg": n_neg_sub, "skipped": True,
            }
            continue
        full_sub = fit_eval_pair(sub_lab, feats_clin,
                                   feats_full_primary)
        if full_sub is None:
            continue
        delta_sub = full_sub["auc_full"] - full_sub["auc_clin"]
        # Bootstrap CI on subgroup
        sub_deltas = []
        for _ in range(N_BOOTSTRAPS):
            idx = rng.integers(0, n_sub, size=n_sub)
            br = [sub_lab[i] for i in idx]
            res = fit_eval_pair(br, feats_clin, feats_full_primary)
            if res is not None:
                sub_deltas.append(
                    res["auc_full"] - res["auc_clin"])
        sub_deltas = np.array(sub_deltas)
        ci_low = float(np.percentile(sub_deltas, 2.5))
        ci_high = float(np.percentile(sub_deltas, 97.5))
        p_one = float((sub_deltas <= 0).mean())
        print(f"  {sub_name}: n={n_sub} ({n_pos_sub} pos, "
              f"{n_neg_sub} neg)  AUC_clin={full_sub['auc_clin']:.4f}  "
              f"AUC_full={full_sub['auc_full']:.4f}  "
              f"Delta={delta_sub:+.4f} [CI {ci_low:+.4f}, "
              f"{ci_high:+.4f}], P(<=0)={p_one:.3f}", flush=True)
        subgroup_results[sub_name] = {
            "n": n_sub, "n_pos": n_pos_sub, "n_neg": n_neg_sub,
            "auc_clin": float(full_sub["auc_clin"]),
            "auc_full": float(full_sub["auc_full"]),
            "delta_AUC_point": float(delta_sub),
            "delta_AUC_95_CI": [ci_low, ci_high],
            "p_one_sided": p_one,
            "skipped": False,
        }

    out = {
        "version": "v206",
        "experiment": ("Permutation test + sigma sweep + IDH/MGMT "
                       "subgroup analysis on binary 365-d PFS"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "sigma_list": SIGMA_LIST,
        "sigma_primary": SIGMA_PRIMARY,
        "n_permutations": N_PERMUTATIONS,
        "n_bootstraps": N_BOOTSTRAPS,
        "n_total": len(labelled),
        "n_pos": n_pos,
        "n_neg": n_neg,
        "sigma_sweep": sigma_results,
        "permutation_test": {
            "delta_observed": float(delta_observed),
            "n_null": len(null_deltas),
            "null_mean": float(null_deltas.mean()),
            "null_95th_percentile": null_p95,
            "null_99th_percentile": null_p99,
            "p_value_one_sided": p_perm,
        },
        "subgroup_analysis": subgroup_results,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
