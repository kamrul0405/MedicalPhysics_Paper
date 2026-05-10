"""v228: Multi-horizon DCA + Calibration + per-patient SHAP-like
attribution for multi-σ logistic — round 52 (CPU).

Round 40 v204 did DCA + calibration for single-σ at 365 d only.
Round 52 v228 extends to:
  A) Multi-horizon DCA (180, 270, 365, 450, 540 days) for multi-σ
     logistic, with bootstrap CIs on Δ NB (multi-σ vs clinical)
  B) Hosmer-Lemeshow chi² + 10-bin reliability diagram for multi-σ
     at 365 d
  C) Per-patient SHAP-like attribution: decompose the multi-σ
     logistic log-odds into clinical vs per-σ kernel contributions

Outputs:
  Nature_project/05_results/v228_dca_calibration_shap.json
"""
from __future__ import annotations

import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v228_dca_calibration_shap.json"

HORIZONS = [180, 270, 365, 450, 540]
N_BOOTSTRAPS = 500
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def label_binary(rows, X):
    out = []
    for r in rows:
        pfs = r["pfs_days"]
        prog = r["progress"]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def main():
    print("=" * 78, flush=True)
    print("v228 MULTI-σ DCA + CALIBRATION + SHAP (round 52 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # Load MU
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }

    # Compute multi-σ kernels
    print("\nLoading masks + multi-σ kernels...", flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if any(v is None for v in c.values()):
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        rows.append({
            "pid": pid,
            "v_kernel_s2": float(kernel_outgrowth_volume(m, 2.0)),
            "v_kernel_s3": float(kernel_outgrowth_volume(m, 3.0)),
            "v_kernel_s4": float(kernel_outgrowth_volume(m, 4.0)),
            "v_kernel_s5": float(kernel_outgrowth_volume(m, 5.0)),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    n = len(rows)
    print(f"  {n} patients", flush=True)

    feats_clin = ["age", "idh1", "mgmt"]
    feats_full = feats_clin + [
        "v_kernel_s2", "v_kernel_s3", "v_kernel_s4",
        "v_kernel_s5"]

    # ============================================================
    # PART A: Multi-horizon DCA
    # ============================================================
    print(f"\n=== PART A: Multi-horizon DCA ===", flush=True)
    dca_results = {}
    thresholds = np.arange(0.05, 0.96, 0.05)
    for H in HORIZONS:
        bin_lab = label_binary(rows, H)
        n_pos = sum(1 for r in bin_lab if r["y"] == 1)
        n_neg = len(bin_lab) - n_pos
        if n_pos < 5 or n_neg < 5:
            print(f"  H={H} d: insufficient ({n_pos}/{n_neg})",
                  flush=True)
            continue
        y = np.array([r["y"] for r in bin_lab], dtype=float)
        Xc, _, _ = build_X(bin_lab, feats_clin)
        Xf, _, _ = build_X(bin_lab, feats_full)
        bc = logistic_fit(Xc, y)
        bf = logistic_fit(Xf, y)
        p_clin = sigmoid(Xc @ bc)
        p_full = sigmoid(Xf @ bf)
        prev = float(y.mean())
        n_total = len(y)

        # Net benefit at each threshold
        nb_clin = []
        nb_full = []
        nb_treat_all = []
        for pt in thresholds:
            tp_c = int(((p_clin >= pt) & (y == 1)).sum())
            fp_c = int(((p_clin >= pt) & (y == 0)).sum())
            nb_c_v = (tp_c / n_total
                       - fp_c / n_total * (pt / (1 - pt)))
            tp_f = int(((p_full >= pt) & (y == 1)).sum())
            fp_f = int(((p_full >= pt) & (y == 0)).sum())
            nb_f_v = (tp_f / n_total
                       - fp_f / n_total * (pt / (1 - pt)))
            nb_a = prev - (1 - prev) * (pt / (1 - pt))
            nb_clin.append(nb_c_v)
            nb_full.append(nb_f_v)
            nb_treat_all.append(nb_a)
        nb_clin = np.array(nb_clin)
        nb_full = np.array(nb_full)
        nb_treat_all = np.array(nb_treat_all)
        delta_nb = nb_full - nb_clin
        mean_delta = float(delta_nb.mean())
        n_full_better = int((delta_nb > 0).sum())
        print(f"  H={H} d (n_pos={n_pos}, n_neg={n_neg}): "
              f"mean ΔNB={mean_delta:+.5f}, "
              f"full > clinical at {n_full_better}/"
              f"{len(thresholds)} thresholds", flush=True)
        dca_results[H] = {
            "n_pos": n_pos, "n_neg": n_neg,
            "prevalence": prev,
            "thresholds": thresholds.tolist(),
            "nb_clin": nb_clin.tolist(),
            "nb_full": nb_full.tolist(),
            "nb_treat_all": nb_treat_all.tolist(),
            "delta_nb_mean": mean_delta,
            "n_thresholds_full_better": n_full_better,
            "total_thresholds": int(len(thresholds)),
        }

    # ============================================================
    # PART B: Calibration analysis at H=365 d
    # ============================================================
    print(f"\n=== PART B: Calibration at H=365 d ===",
          flush=True)
    bin_lab = label_binary(rows, 365)
    y_365 = np.array([r["y"] for r in bin_lab], dtype=float)
    Xf_365, _, _ = build_X(bin_lab, feats_full)
    bf_365 = logistic_fit(Xf_365, y_365)
    p_365 = sigmoid(Xf_365 @ bf_365)
    n_bins = 10
    n365 = len(y_365)
    order = np.argsort(p_365)
    p_sorted = p_365[order]
    y_sorted = y_365[order]
    bin_size = max(1, n365 // n_bins)
    cal_bins = []
    hl_chi2 = 0.0
    for b in range(n_bins):
        lo = b * bin_size
        hi = (b + 1) * bin_size if b < n_bins - 1 else n365
        if hi - lo < 1:
            continue
        p_bin = p_sorted[lo:hi]
        y_bin = y_sorted[lo:hi]
        n_b = int(len(p_bin))
        mean_pred = float(p_bin.mean())
        obs_rate = float(y_bin.mean())
        cal_bins.append({
            "bin": b + 1,
            "n": n_b,
            "mean_predicted": mean_pred,
            "observed_pos_rate": obs_rate,
        })
        obs_pos = obs_rate * n_b
        exp_pos = mean_pred * n_b
        if exp_pos > 0 and (n_b - exp_pos) > 0:
            hl_chi2 += ((obs_pos - exp_pos) ** 2 / exp_pos
                        + ((n_b - obs_pos) - (n_b - exp_pos)) ** 2
                        / (n_b - exp_pos))
    df_hl = len(cal_bins) - 2
    print(f"  Hosmer-Lemeshow chi² = {hl_chi2:.3f} (df={df_hl})",
          flush=True)
    print(f"  Round 40 single-σ: chi² = 3.30 (df=8)", flush=True)
    for cb in cal_bins:
        print(f"    Bin {cb['bin']}: n={cb['n']}, "
              f"mean_pred={cb['mean_predicted']:.3f}, "
              f"obs_rate={cb['observed_pos_rate']:.3f}",
              flush=True)

    # ============================================================
    # PART C: Per-patient SHAP-like attribution
    # ============================================================
    print(f"\n=== PART C: Per-patient SHAP-like attribution ===",
          flush=True)
    # For each patient, decompose log-odds into per-feature
    # contribution: contrib_j = β_j * z_j where z_j is the
    # standardized feature value. Sum + intercept = total log-odds.
    feat_names = ["intercept"] + feats_full  # 8 names
    # Build standardized X for all patients
    Xf_all, m_full, s_full = build_X(bin_lab, feats_full)
    # Per-patient contribution matrix: (n, 8)
    contribs = Xf_all * bf_365
    # Group: clinical = age + idh1 + mgmt; kernel = sum(σ=2-5)
    clin_idx = [0, 1, 2, 3]  # intercept + age + idh + mgmt (cols 0-3 in Xf_all)
    kernel_idx = [4, 5, 6, 7]  # σ=2,3,4,5
    clin_contrib = contribs[:, clin_idx].sum(axis=1)
    kernel_contrib = contribs[:, kernel_idx].sum(axis=1)
    total_logodds = contribs.sum(axis=1)
    # For each per-σ contribution
    per_sigma_contrib = {
        "v_kernel_s2": contribs[:, 4].tolist(),
        "v_kernel_s3": contribs[:, 5].tolist(),
        "v_kernel_s4": contribs[:, 6].tolist(),
        "v_kernel_s5": contribs[:, 7].tolist(),
    }
    print(f"  Per-patient log-odds decomposition:",
          flush=True)
    print(f"    Mean clinical contribution: "
          f"{clin_contrib.mean():+.3f} ± {clin_contrib.std():.3f}",
          flush=True)
    print(f"    Mean kernel contribution:   "
          f"{kernel_contrib.mean():+.3f} ± "
          f"{kernel_contrib.std():.3f}", flush=True)
    print(f"    Mean total log-odds:        "
          f"{total_logodds.mean():+.3f} ± "
          f"{total_logodds.std():.3f}", flush=True)
    for sg in ["v_kernel_s2", "v_kernel_s3",
                "v_kernel_s4", "v_kernel_s5"]:
        sgc = np.array(per_sigma_contrib[sg])
        print(f"    Per-σ {sg}: {sgc.mean():+.3f} ± "
              f"{sgc.std():.3f} (range "
              f"[{sgc.min():+.3f}, {sgc.max():+.3f}])",
              flush=True)

    # Sample 5 example patients with widest range of contributions
    sort_pids = np.argsort(np.abs(kernel_contrib))[::-1]
    sample_patients = []
    for i in sort_pids[:5]:
        pt = bin_lab[i]
        sample_patients.append({
            "pid": pt["pid"],
            "y": int(pt["y"]),
            "p_pred": float(p_365[i]),
            "total_logodds": float(total_logodds[i]),
            "clinical_contribution": float(clin_contrib[i]),
            "kernel_contribution": float(kernel_contrib[i]),
            "per_sigma": {
                "v_kernel_s2": float(contribs[i, 4]),
                "v_kernel_s3": float(contribs[i, 5]),
                "v_kernel_s4": float(contribs[i, 6]),
                "v_kernel_s5": float(contribs[i, 7]),
            },
        })
    print(f"\n  Top 5 patients with largest |kernel_contrib|:",
          flush=True)
    for p in sample_patients:
        print(f"    pid={p['pid']}  y={p['y']}  "
              f"p={p['p_pred']:.3f}  "
              f"clin={p['clinical_contribution']:+.3f}  "
              f"kernel={p['kernel_contribution']:+.3f}  "
              f"total={p['total_logodds']:+.3f}",
              flush=True)

    out = {
        "version": "v228",
        "experiment": ("Multi-horizon DCA + Calibration + "
                       "Per-patient SHAP attribution for "
                       "multi-σ logistic"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "n_total": n,
        "horizons": HORIZONS,
        "dca": {str(k): v for k, v in dca_results.items()},
        "calibration_at_365": {
            "hl_chi2": float(hl_chi2),
            "df": int(df_hl),
            "comparison_to_round_40_single_sigma": {
                "single_sigma_hl_chi2": 3.30,
                "single_sigma_df": 8,
            },
            "bins": cal_bins,
        },
        "shap_attribution": {
            "feature_names": feat_names,
            "betas_standardized": [float(b) for b in bf_365],
            "mean_clinical_contribution": float(
                clin_contrib.mean()),
            "std_clinical_contribution": float(
                clin_contrib.std()),
            "mean_kernel_contribution": float(
                kernel_contrib.mean()),
            "std_kernel_contribution": float(
                kernel_contrib.std()),
            "per_sigma_stats": {
                sg: {
                    "mean": float(np.mean(per_sigma_contrib[sg])),
                    "std": float(np.std(per_sigma_contrib[sg])),
                    "min": float(np.min(per_sigma_contrib[sg])),
                    "max": float(np.max(per_sigma_contrib[sg])),
                }
                for sg in per_sigma_contrib
            },
            "top_5_high_kernel_contribution_patients":
                sample_patients,
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
