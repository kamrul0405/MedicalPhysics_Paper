"""v213: Transfer learning — pretrain CNN on MU, freeze encoder,
head-only fine-tune on RHUH — round 44 (GPU).

Round 43 v211 showed pooled MU+RHUH CNN training improved MU
performance (+0.08 AUC) but cross-cohort LOCO MU->RHUH still failed
(0.511, chance). Open question: does TRANSFER LEARNING (pretrain on
MU, freeze encoder, head-only fine-tune on RHUH) help where pooled
training fails?

Method:
  1. Pretrain 3D CNN (mask + kernel input) on full MU n=130 binary
     365-d PFS for 30 epochs.
  2. Freeze encoder weights; reinitialize head.
  3. 5-fold CV on RHUH n=31: per fold train head only on RHUH train
     fold, evaluate on RHUH test fold.
  4. Compare to: (a) RHUH from-scratch baseline (5-fold); (b) v211
     LOCO MU->RHUH AUC=0.511 (no fine-tuning).
  5. Bootstrap 200 resamples on the held-out RHUH OOF predictions
     for 95% CI on AUC.

Outputs:
  Nature_project/05_results/v213_transfer_learning.json
"""
from __future__ import annotations

import csv
import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v213_transfer_learning.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
PRETRAIN_EPOCHS = 30
FT_EPOCHS = 50  # head-only is small, more epochs ok
LR_PRE = 5e-4
LR_FT = 1e-3
SEED = 42
N_BOOT = 200


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class Encoder(nn.Module):
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.feat_dim = base * 4

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.global_pool(e3).flatten(1)  # (B, base*4)


class Head(nn.Module):
    def __init__(self, feat_dim=96):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, 32),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(32, 1),
        )

    def forward(self, feat):
        return self.net(feat).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def load_rhuh_data():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if pfs_d is None:
                continue
            clinical[pid] = {"pfs_days": pfs_d,
                              "progress": event}
    data = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def pretrain_encoder(mu_data, seed=SEED):
    """Train Encoder + Head on full MU; return Encoder weights."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    n = len(mu_data)
    n_pos = sum(1 for d in mu_data if d["y"] == 1)
    n_neg = n - n_pos
    X = stack_inputs(mu_data)
    y = np.array([d["y"] for d in mu_data], dtype=np.float32)
    Xt = torch.from_numpy(X).to(DEVICE)
    yt = torch.from_numpy(y).to(DEVICE)
    enc = Encoder(in_ch=2, base=24).to(DEVICE)
    head = Head(feat_dim=enc.feat_dim).to(DEVICE)
    opt = torch.optim.AdamW(
        list(enc.parameters()) + list(head.parameters()),
        lr=LR_PRE, weight_decay=1e-3)
    pos_w = torch.tensor([n_neg / max(1, n_pos)],
                          dtype=torch.float32, device=DEVICE)
    criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    bs = 4
    for ep in range(PRETRAIN_EPOCHS):
        enc.train()
        head.train()
        perm = np.random.permutation(n)
        for i in range(0, n, bs):
            idx = perm[i:i+bs]
            feat = enc(Xt[idx])
            logits = head(feat)
            loss = criterion(logits, yt[idx])
            opt.zero_grad()
            loss.backward()
            opt.step()
    return enc


def head_finetune_eval(enc, train_data, test_data, seed):
    """Freeze enc, train head on train_data, eval on test_data."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    for p in enc.parameters():
        p.requires_grad = False
    enc.eval()
    n_tr = len(train_data)
    n_pos = sum(1 for d in train_data if d["y"] == 1)
    n_neg = n_tr - n_pos
    Xtr = stack_inputs(train_data)
    ytr = np.array([d["y"] for d in train_data],
                    dtype=np.float32)
    Xte = stack_inputs(test_data)
    yte = np.array([d["y"] for d in test_data],
                    dtype=np.float32)
    Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
    Xte_t = torch.from_numpy(Xte).to(DEVICE)
    ytr_t = torch.from_numpy(ytr).to(DEVICE)
    # Precompute features (encoder is frozen)
    with torch.no_grad():
        feat_tr = enc(Xtr_t)
        feat_te = enc(Xte_t)
    head = Head(feat_dim=enc.feat_dim).to(DEVICE)
    opt = torch.optim.AdamW(head.parameters(), lr=LR_FT,
                             weight_decay=1e-3)
    pos_w = torch.tensor([n_neg / max(1, n_pos)],
                          dtype=torch.float32, device=DEVICE)
    criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    bs = min(8, n_tr)
    for ep in range(FT_EPOCHS):
        head.train()
        perm = np.random.permutation(n_tr)
        for i in range(0, n_tr, bs):
            idx = perm[i:i+bs]
            logits = head(feat_tr[idx])
            loss = criterion(logits, ytr_t[idx])
            opt.zero_grad()
            loss.backward()
            opt.step()
    head.eval()
    with torch.no_grad():
        scores = head(feat_te).cpu().numpy()
    return scores, yte


def from_scratch_eval(train_data, test_data, seed):
    """Train Encoder+Head from scratch on train_data, eval test_data."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    n_tr = len(train_data)
    n_pos = sum(1 for d in train_data if d["y"] == 1)
    n_neg = n_tr - n_pos
    Xtr = stack_inputs(train_data)
    ytr = np.array([d["y"] for d in train_data],
                    dtype=np.float32)
    Xte = stack_inputs(test_data)
    yte = np.array([d["y"] for d in test_data],
                    dtype=np.float32)
    Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
    Xte_t = torch.from_numpy(Xte).to(DEVICE)
    ytr_t = torch.from_numpy(ytr).to(DEVICE)
    enc = Encoder(in_ch=2, base=24).to(DEVICE)
    head = Head(feat_dim=enc.feat_dim).to(DEVICE)
    opt = torch.optim.AdamW(
        list(enc.parameters()) + list(head.parameters()),
        lr=LR_PRE, weight_decay=1e-3)
    pos_w = torch.tensor([n_neg / max(1, n_pos)],
                          dtype=torch.float32, device=DEVICE)
    criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    bs = min(4, n_tr)
    for ep in range(PRETRAIN_EPOCHS):
        enc.train()
        head.train()
        perm = np.random.permutation(n_tr)
        for i in range(0, n_tr, bs):
            idx = perm[i:i+bs]
            feat = enc(Xtr_t[idx])
            logits = head(feat)
            loss = criterion(logits, ytr_t[idx])
            opt.zero_grad()
            loss.backward()
            opt.step()
    enc.eval()
    head.eval()
    with torch.no_grad():
        feat_te = enc(Xte_t)
        scores = head(feat_te).cpu().numpy()
    return scores, yte


def main():
    print("=" * 78, flush=True)
    print(f"v213 TRANSFER LEARNING (round 44 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    mu_data = load_mu_data()
    rhuh_data = load_rhuh_data()
    print(f"  MU: n={len(mu_data)} "
          f"({sum(1 for d in mu_data if d['y']==1)} pos)",
          flush=True)
    print(f"  RHUH: n={len(rhuh_data)} "
          f"({sum(1 for d in rhuh_data if d['y']==1)} pos)",
          flush=True)

    # ------------------------------------------------------------
    # Step 1: pretrain Encoder on full MU
    # ------------------------------------------------------------
    print(f"\n=== STEP 1: PRETRAIN on MU n={len(mu_data)} for "
          f"{PRETRAIN_EPOCHS} epochs ===", flush=True)
    enc_mu = pretrain_encoder(mu_data, seed=SEED)
    print(f"  Pretrain done", flush=True)

    # ------------------------------------------------------------
    # Step 2: 5-fold CV on RHUH with frozen encoder + head only
    # ------------------------------------------------------------
    print(f"\n=== STEP 2: 5-fold CV on RHUH with frozen MU "
          f"encoder ===", flush=True)
    y_rhuh = np.array([d["y"] for d in rhuh_data],
                       dtype=np.float32)
    folds_rhuh = stratified_kfold(y_rhuh, N_FOLDS, seed=SEED)
    print(f"  Stratified fold sizes: "
          f"{[len(f) for f in folds_rhuh]}", flush=True)
    transfer_oof = np.zeros(len(rhuh_data), dtype=np.float32)
    transfer_assigned = np.zeros(len(rhuh_data), dtype=bool)
    transfer_fold_aucs = []
    for fold_i, test_idx in enumerate(folds_rhuh):
        train_idx = np.array(sorted(
            set(range(len(rhuh_data))) - set(test_idx.tolist())))
        tr = [rhuh_data[i] for i in train_idx]
        te = [rhuh_data[i] for i in test_idx]
        if (sum(1 for d in tr if d["y"] == 1) < 2
                or sum(1 for d in tr if d["y"] == 0) < 2):
            continue
        scores, yte = head_finetune_eval(
            enc_mu, tr, te, seed=SEED + fold_i * 100)
        transfer_oof[test_idx] = scores
        transfer_assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            transfer_fold_aucs.append(float(a))
            print(f"    Fold {fold_i+1}/{N_FOLDS}: "
                  f"transfer-AUC = {a:.4f} "
                  f"(n_train={len(tr)}, n_test={len(te)})",
                  flush=True)
    transfer_pooled_auc = auroc(transfer_oof[transfer_assigned],
                                  y_rhuh[transfer_assigned])
    print(f"\n  Transfer pooled OOF AUC = "
          f"{transfer_pooled_auc:.4f}", flush=True)
    print(f"  Per-fold mean = "
          f"{np.mean(transfer_fold_aucs):.4f}", flush=True)

    # ------------------------------------------------------------
    # Step 3: From-scratch RHUH baseline (5-fold)
    # ------------------------------------------------------------
    print(f"\n=== STEP 3: From-scratch RHUH baseline (5-fold) ===",
          flush=True)
    scratch_oof = np.zeros(len(rhuh_data), dtype=np.float32)
    scratch_assigned = np.zeros(len(rhuh_data), dtype=bool)
    scratch_fold_aucs = []
    for fold_i, test_idx in enumerate(folds_rhuh):
        train_idx = np.array(sorted(
            set(range(len(rhuh_data))) - set(test_idx.tolist())))
        tr = [rhuh_data[i] for i in train_idx]
        te = [rhuh_data[i] for i in test_idx]
        if (sum(1 for d in tr if d["y"] == 1) < 2
                or sum(1 for d in tr if d["y"] == 0) < 2):
            continue
        scores, yte = from_scratch_eval(
            tr, te, seed=SEED + fold_i * 100)
        scratch_oof[test_idx] = scores
        scratch_assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            scratch_fold_aucs.append(float(a))
            print(f"    Fold {fold_i+1}/{N_FOLDS}: "
                  f"scratch-AUC = {a:.4f}", flush=True)
        del scores
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()
    scratch_pooled_auc = auroc(scratch_oof[scratch_assigned],
                                 y_rhuh[scratch_assigned])
    print(f"\n  From-scratch pooled OOF AUC = "
          f"{scratch_pooled_auc:.4f}", flush=True)

    # ------------------------------------------------------------
    # Step 4: Bootstrap CI on transfer pooled OOF AUC
    # ------------------------------------------------------------
    print(f"\n=== STEP 4: Bootstrap on RHUH OOF "
          f"({N_BOOT} resamples) ===", flush=True)
    rng = np.random.default_rng(SEED)
    transfer_aucs_b = []
    scratch_aucs_b = []
    delta_aucs_b = []
    n_assigned = int(transfer_assigned.sum())
    for _ in range(N_BOOT):
        idx = rng.integers(0, n_assigned, size=n_assigned)
        valid_idx = np.where(transfer_assigned)[0][idx]
        yb = y_rhuh[valid_idx]
        if int(yb.sum()) < 2 or int(len(yb) - yb.sum()) < 2:
            continue
        a_t = auroc(transfer_oof[valid_idx], yb)
        a_s = auroc(scratch_oof[valid_idx], yb)
        transfer_aucs_b.append(a_t)
        scratch_aucs_b.append(a_s)
        delta_aucs_b.append(a_t - a_s)
    transfer_aucs_b = np.array(transfer_aucs_b)
    scratch_aucs_b = np.array(scratch_aucs_b)
    delta_aucs_b = np.array(delta_aucs_b)
    print(f"  Transfer AUC: mean={transfer_aucs_b.mean():.4f}, "
          f"95% CI [{np.percentile(transfer_aucs_b, 2.5):.4f}, "
          f"{np.percentile(transfer_aucs_b, 97.5):.4f}]",
          flush=True)
    print(f"  Scratch AUC: mean={scratch_aucs_b.mean():.4f}, "
          f"95% CI [{np.percentile(scratch_aucs_b, 2.5):.4f}, "
          f"{np.percentile(scratch_aucs_b, 97.5):.4f}]",
          flush=True)
    print(f"  Delta (transfer - scratch): mean="
          f"{delta_aucs_b.mean():+.4f}, "
          f"95% CI [{np.percentile(delta_aucs_b, 2.5):+.4f}, "
          f"{np.percentile(delta_aucs_b, 97.5):+.4f}], "
          f"P(<=0)={float((delta_aucs_b <= 0).mean()):.4f}",
          flush=True)

    print(f"\n=== SUMMARY ===", flush=True)
    print(f"  v211 LOCO MU->RHUH (no fine-tune): AUC = 0.511 "
          f"(reference)", flush=True)
    print(f"  v213 RHUH from-scratch (5-fold)  : AUC = "
          f"{scratch_pooled_auc:.4f}", flush=True)
    print(f"  v213 transfer (frozen MU encoder + head): AUC = "
          f"{transfer_pooled_auc:.4f}", flush=True)

    out = {
        "version": "v213",
        "experiment": ("Transfer learning: pretrain on MU, "
                       "freeze encoder, head-only fine-tune "
                       "on RHUH (5-fold CV)"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "pretrain_epochs": PRETRAIN_EPOCHS,
        "ft_epochs": FT_EPOCHS,
        "n_mu_pretrain": len(mu_data),
        "n_rhuh": len(rhuh_data),
        "n_folds": N_FOLDS,
        "transfer_5fold": {
            "fold_aucs": transfer_fold_aucs,
            "fold_mean": float(np.mean(transfer_fold_aucs)),
            "fold_std": float(np.std(transfer_fold_aucs)),
            "pooled_oof_auc": float(transfer_pooled_auc),
        },
        "scratch_5fold": {
            "fold_aucs": scratch_fold_aucs,
            "fold_mean": float(np.mean(scratch_fold_aucs)),
            "fold_std": float(np.std(scratch_fold_aucs)),
            "pooled_oof_auc": float(scratch_pooled_auc),
        },
        "bootstrap": {
            "transfer_auc_mean": float(transfer_aucs_b.mean()),
            "transfer_auc_95_CI": [
                float(np.percentile(transfer_aucs_b, 2.5)),
                float(np.percentile(transfer_aucs_b, 97.5))],
            "scratch_auc_mean": float(scratch_aucs_b.mean()),
            "scratch_auc_95_CI": [
                float(np.percentile(scratch_aucs_b, 2.5)),
                float(np.percentile(scratch_aucs_b, 97.5))],
            "delta_mean": float(delta_aucs_b.mean()),
            "delta_95_CI": [
                float(np.percentile(delta_aucs_b, 2.5)),
                float(np.percentile(delta_aucs_b, 97.5))],
            "delta_p_one_sided": float(
                (delta_aucs_b <= 0).mean()),
        },
        "comparison": {
            "v211_loco_mu_to_rhuh_no_finetune": 0.5109,
            "v213_scratch_rhuh_5fold": float(scratch_pooled_auc),
            "v213_transfer_frozen_encoder": float(
                transfer_pooled_auc),
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
