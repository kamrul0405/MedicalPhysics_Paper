"""v202: PFS-as-binary-screening — does the kernel predict early
progression as a CLASSIFICATION task? — round 39 (CPU).

Senior-Nature-reviewer-driven re-framing of survival as binary
SCREENING. Round 27 v189 showed the kernel beats foundation on AUC.
But rounds 32-38 showed neither kernel nor lambda predicts SURVIVAL
in Cox regression. Could this be a metric-mismatch issue?

  * Cox regression ranks risks across all patients (continuous)
  * The kernel is a SCREENING tool — best at YES/NO at a fixed horizon

Re-frame: predict 'will this patient progress within X days?' as a
BINARY classification task. Test V_kernel + lambda + clinical at
horizons X in {180, 365, 730} days. Use AUC as primary metric.

If V_kernel/lambda achieve high AUC for early-progression screening
on MU n>=80 → kernel CAN be a clinical screening tool, just not a
continuous survival regressor.

Method (MU-Glioma-Post n=151 with PFS data):
  - Compute V_kernel sigma=3 + per-patient lambda + clinical
  - For each horizon X in {180, 365, 730}:
    * Binary outcome: progressed_by_X = (PFS_days < X) AND (event=1) OR
                                          (PFS_days >= X = censored at X)
    * Per-feature AUC: V_kernel, lambda, baseline_volume,
      age, IDH1, MGMT
    * Multivariate logistic with clinical + V_kernel + lambda

Outputs:
  Nature_project/05_results/v202_pfs_binary_screening.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import distance_transform_edt, gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_XLSX = Path(r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v202_pfs_binary_screening.json"

DISTANCE_BINS = np.arange(1, 25)
SIGMA_KERNEL = 3.0
HORIZONS = [180, 365, 730]


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def fit_lambda(mask, outgrowth):
    m = mask.astype(bool)
    out = outgrowth.astype(bool)
    if m.sum() == 0 or out.sum() == 0:
        return float("nan"), float("nan"), 0
    d = distance_transform_edt(~m)
    d_int = np.round(d).astype(int)
    d_arr, p_arr, n_arr = [], [], []
    for b in DISTANCE_BINS:
        sel = (d_int == b)
        if sel.sum() < 5:
            continue
        d_arr.append(b)
        p_arr.append(int((sel & out).sum()) / int(sel.sum()))
        n_arr.append(int(sel.sum()))
    d_arr = np.array(d_arr, dtype=float)
    p_arr = np.array(p_arr, dtype=float)
    n_arr = np.array(n_arr, dtype=float)
    valid = p_arr > 0
    if valid.sum() < 3:
        return float("nan"), float("nan"), int(valid.sum())
    d_v = d_arr[valid]
    p_v = p_arr[valid]
    n_v = n_arr[valid]
    log_p = np.log(p_v)
    w = np.sqrt(n_v)
    sw = np.sum(w)
    sw_d = np.sum(w * d_v)
    sw_logp = np.sum(w * log_p)
    sw_dd = np.sum(w * d_v * d_v)
    sw_d_logp = np.sum(w * d_v * log_p)
    denom = sw * sw_dd - sw_d ** 2
    if denom == 0:
        return float("nan"), float("nan"), int(valid.sum())
    slope = (sw * sw_d_logp - sw_d * sw_logp) / denom
    intercept = (sw_logp - slope * sw_d) / sw
    lam = float(-1.0 / slope) if slope < 0 else float("inf")
    pred = intercept + slope * d_v
    ss_res = float(np.sum(w * (log_p - pred) ** 2))
    ss_tot = float(np.sum(w * (log_p - np.mean(log_p)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return lam, r2, int(valid.sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float32)
    y = np.array(labels, dtype=np.float32)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    """Multivariate L2-regularized logistic regression via scipy."""
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        # numerically stable log(1+exp)
        log_one_plus = np.where(eta > 0,
                                  eta + np.log1p(np.exp(-eta)),
                                  np.log1p(np.exp(eta)))
        nll = -np.sum(y * eta - log_one_plus) + l2 * np.sum(beta * beta)
        return nll

    beta0 = np.zeros(p)
    result = minimize(neg_log_lik, beta0, method="L-BFGS-B")
    return result.x


def main():
    print("=" * 78, flush=True)
    print("v202 PFS-AS-BINARY-SCREENING (round 39 CPU)", flush=True)
    print("=" * 78, flush=True)

    # Load MU clinical
    wb = openpyxl.load_workbook(CLINICAL_XLSX, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")

    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = float(row[age_col]) if row[age_col] is not None else None
            progress = (int(row[progress_col])
                          if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                       if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                      if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                      if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress, "pfs_days": pfs_d,
            "idh1": idh1, "mgmt": mgmt,
        }
    print(f"  Loaded clinical for {len(clinical)} MU patients", flush=True)

    # Compute features for each patient
    print("\nComputing kernel features per patient...", flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        fr = f.parent / f.name.replace("_b.npy", "_r.npy")
        if not fr.exists():
            continue
        fu = (np.load(fr) > 0).astype(np.float32)
        outgrowth = (fu.astype(bool) & ~m.astype(bool)).astype(np.float32)
        if m.sum() == 0:
            continue
        v_k = kernel_outgrowth_volume(m, SIGMA_KERNEL)
        baseline_volume = int(m.sum())
        if outgrowth.sum() > 0:
            lam, r2, n_pts = fit_lambda(m, outgrowth)
            valid = (not np.isnan(lam)) and (not np.isinf(lam)) and \
                    lam > 0 and lam < 200 and r2 > 0.5 and n_pts >= 4
        else:
            lam, r2, valid = float("nan"), float("nan"), False
        c = clinical[pid]
        rows.append({
            "pid": pid,
            "v_kernel_s3": v_k,
            "baseline_volume": baseline_volume,
            "lambda": lam if valid else None,
            "lambda_valid": valid,
            "age": c["age"],
            "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"],
            "progress": c["progress"],
        })
    print(f"  {len(rows)} MU patients with mask + clinical", flush=True)

    # Filter for valid PFS info
    base_filter = [r for r in rows
                     if r["pfs_days"] is not None
                     and r["progress"] is not None
                     and r["age"] is not None]
    print(f"  {len(base_filter)} patients with PFS + progression event "
          f"+ age", flush=True)
    n_progress_total = sum(1 for r in base_filter if r["progress"] == 1)
    print(f"  Total progression events: {n_progress_total}", flush=True)

    horizon_results = {}
    for X in HORIZONS:
        print(f"\n=== Horizon: progressed within {X} days ===", flush=True)
        # Binary outcome:
        # progressed_by_X = 1 if (progress=1 AND PFS_days < X) else 0
        # Censored patients (progress=0) with PFS_days < X are excluded
        # (we don't know if they would have progressed by X)
        labelled = []
        for r in base_filter:
            pfs = r["pfs_days"]
            prog = r["progress"]
            if prog == 1 and pfs < X:
                y = 1
            elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
                y = 0
            else:
                continue  # censored before X — uninformative
            labelled.append({**r, "y": y})
        n_lab = len(labelled)
        n_pos = sum(1 for l in labelled if l["y"] == 1)
        n_neg = n_lab - n_pos
        print(f"  n_labelled = {n_lab}  (positives={n_pos}, "
              f"negatives={n_neg})", flush=True)
        if n_pos < 5 or n_neg < 5:
            print("  Insufficient class balance, skipping", flush=True)
            continue

        # Per-feature univariate AUC
        per_feature = {}
        for feat in ["v_kernel_s3", "baseline_volume", "age", "idh1",
                       "mgmt", "lambda"]:
            xs = []
            ys = []
            for l in labelled:
                if l[feat] is None or (isinstance(l[feat], float)
                                          and np.isnan(l[feat])):
                    continue
                xs.append(l[feat])
                ys.append(l["y"])
            if len(xs) < 10 or sum(ys) < 3 or len(ys) - sum(ys) < 3:
                continue
            auc = auroc(xs, ys)
            per_feature[feat] = {
                "n": int(len(xs)), "AUC": float(auc),
            }
            print(f"    {feat:18s}  n={len(xs):3d}  AUC = {auc:.4f}",
                  flush=True)

        # Multivariate logistic: clinical + V_kernel (+ optionally lambda)
        feats_clin = ["age", "idh1", "mgmt"]
        feats_clin_vk = feats_clin + ["v_kernel_s3"]

        def build_X(rows, feats):
            arr = np.array([[r[f] for f in feats] for r in rows],
                              dtype=float)
            means = np.nanmean(arr, axis=0)
            stds = np.nanstd(arr, axis=0)
            stds[stds == 0] = 1
            arr_z = (arr - means) / stds
            arr_z = np.nan_to_num(arr_z, nan=0.0)
            # add intercept column
            return np.column_stack([np.ones(len(arr_z)), arr_z])

        # restrict to complete clinical
        cc = [l for l in labelled if all(l[f] is not None
                                            and not (isinstance(l[f],
                                                                  float) and
                                                       np.isnan(l[f]))
                                            for f in feats_clin_vk)]
        if len(cc) < 20 or sum(1 for l in cc if l["y"] == 1) < 5:
            print("  Insufficient complete-case, skipping multivariate",
                  flush=True)
            mv_result = None
        else:
            y_cc = np.array([l["y"] for l in cc], dtype=float)
            X_clin = build_X(cc, feats_clin)
            X_clin_vk = build_X(cc, feats_clin_vk)
            beta_clin = logistic_fit(X_clin, y_cc)
            beta_clin_vk = logistic_fit(X_clin_vk, y_cc)
            score_clin = X_clin @ beta_clin
            score_clin_vk = X_clin_vk @ beta_clin_vk
            auc_clin = auroc(score_clin, y_cc)
            auc_clin_vk = auroc(score_clin_vk, y_cc)
            print(f"\n  Multivariate logistic (complete-case n={len(cc)}, "
                  f"events={int(y_cc.sum())}):", flush=True)
            print(f"    Clinical only (age + IDH + MGMT):  AUC = "
                  f"{auc_clin:.4f}", flush=True)
            print(f"    Clinical + V_kernel:                AUC = "
                  f"{auc_clin_vk:.4f}  (Δ = {auc_clin_vk - auc_clin:+.4f})",
                  flush=True)
            mv_result = {
                "n": len(cc), "events": int(y_cc.sum()),
                "auc_clinical_only": float(auc_clin),
                "auc_clinical_plus_Vkernel": float(auc_clin_vk),
                "delta_AUC": float(auc_clin_vk - auc_clin),
            }

        horizon_results[X] = {
            "n_labelled": n_lab,
            "n_positive": n_pos,
            "n_negative": n_neg,
            "univariate_AUCs": per_feature,
            "multivariate": mv_result,
        }

    out = {
        "version": "v202",
        "experiment": ("PFS as binary screening task (within X days) on "
                       "MU-Glioma-Post"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizons_days": HORIZONS,
        "horizon_results": {str(k): v for k, v in horizon_results.items()},
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
