"""v223: SimCLR-features + multi-σ HYBRID logistic — round 49 (GPU).

Round 47 v218 multi-σ logistic = AUC=0.815 (7 features). Round 45
v215 SimCLR-pretrained encoder yields useful representations. Open
question: does CONCATENATING the 96-dim SimCLR feature vector with
the 7 multi-σ features give incremental lift?

Method:
  1. SimCLR pretrain encoder on 509 multi-cohort masks (label-free,
     30 epochs). Freeze encoder.
  2. Extract 96-dim feature vector per MU patient.
  3. Build feature matrices:
     - logistic-baseline: 7 multi-σ features (clinical + V_k(σ=2..5))
     - SimCLR-only: 96 features (encoder output)
     - HYBRID: 96 + 7 = 103 features
  4. 5-fold stratified CV on MU n=130 binary 365-d PFS for each.
  5. Bootstrap CI on Δ AUC.

Tests if learnable representations add value to handcrafted multi-σ
kernel features (and vice versa).

Outputs:
  Nature_project/05_results/v223_simclr_hybrid_logistic.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import gaussian_filter, zoom
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v223_simclr_hybrid_logistic.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
SIMCLR_EPOCHS = 30
TEMPERATURE = 0.5
PROJ_DIM = 64
BASE = 24
SEED = 42
N_BOOT = 500

COHORT_PREFIXES_PRE = ["MU-Glioma-Post", "RHUH-GBM",
                         "UCSF-POSTOP", "LUMIERE"]


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class Encoder(nn.Module):
    def __init__(self, in_ch=2, base=BASE):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.feat_dim = base * 4

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.global_pool(e3).flatten(1)


class ProjectionHead(nn.Module):
    def __init__(self, feat_dim, proj_dim=PROJ_DIM):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, feat_dim), nn.GELU(),
            nn.Linear(feat_dim, proj_dim))

    def forward(self, feat):
        return F.normalize(self.net(feat), dim=-1)


def random_3d_augment(x, rng):
    out = x.clone()
    for axis in [2, 3, 4]:
        if rng.random() > 0.5:
            out = torch.flip(out, dims=[axis])
    scale = float(rng.uniform(0.9, 1.1))
    out = out * scale + 0.02 * torch.randn_like(out)
    return torch.clamp(out, 0.0, 1.5)


def nt_xent_loss(z1, z2, temperature=TEMPERATURE):
    n = z1.size(0)
    z = torch.cat([z1, z2], dim=0)
    sim = torch.matmul(z, z.T) / temperature
    mask = torch.eye(2 * n, dtype=torch.bool,
                      device=sim.device)
    sim = sim.masked_fill(mask, float("-inf"))
    targets = torch.cat([torch.arange(n, 2 * n),
                          torch.arange(0, n)],
                          dim=0).to(sim.device)
    return F.cross_entropy(sim, targets)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def standardize(X, means=None, stds=None):
    if means is None:
        means = np.nanmean(X, axis=0)
    if stds is None:
        stds = np.nanstd(X, axis=0)
        stds[stds == 0] = 1
    Xz = (X - means) / stds
    return np.column_stack([np.ones(len(Xz)), Xz]), means, stds


def load_pretrain_masks():
    samples = []
    for prefix in COHORT_PREFIXES_PRE:
        for f in sorted(CACHE.glob(f"{prefix}_*_b.npy")):
            try:
                m_full = (np.load(f) > 0).astype(np.float32)
                if m_full.sum() == 0:
                    continue
                m = resize_to_target(m_full, TARGET_SHAPE)
                k = heat_bimodal(m, SIGMA_KERNEL)
                samples.append({"pid": f.stem, "mask": m,
                                 "kernel": k})
            except Exception:
                continue
    return samples


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": prog, "pfs_days": pfs_d,
            "idh1": idh1, "mgmt": mgmt,
        }
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["progress"] is None or c["pfs_days"] is None
                or c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None):
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        # Multi-σ kernel volumes (from full-resolution mask)
        m_full_b = (np.load(f) > 0).astype(np.float32)
        vk2 = kernel_outgrowth_volume(m_full_b, 2.0)
        vk3 = kernel_outgrowth_volume(m_full_b, 3.0)
        vk4 = kernel_outgrowth_volume(m_full_b, 4.0)
        vk5 = kernel_outgrowth_volume(m_full_b, 5.0)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y), "age": c["age"],
                     "idh1": c["idh1"], "mgmt": c["mgmt"],
                     "v_k2": float(vk2), "v_k3": float(vk3),
                     "v_k4": float(vk4), "v_k5": float(vk5)})
    return data


def main():
    print("=" * 78, flush=True)
    print(f"v223 SimCLR HYBRID LOGISTIC (round 49 GPU) "
          f"device={DEVICE}", flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    # Load data
    print(f"\n=== STEP 1: Load data ===", flush=True)
    pretrain_samples = load_pretrain_masks()
    print(f"  Multi-cohort masks: {len(pretrain_samples)}",
          flush=True)
    mu_data = load_mu_data()
    print(f"  MU labelled (with multi-σ feats): "
          f"{len(mu_data)}", flush=True)
    y_mu = np.array([d["y"] for d in mu_data], dtype=np.float32)
    print(f"  Pos: {int(y_mu.sum())}, neg: "
          f"{int(len(y_mu) - y_mu.sum())}", flush=True)

    # ---- SimCLR pretrain ----
    print(f"\n=== STEP 2: SimCLR pretrain "
          f"({SIMCLR_EPOCHS} epochs) ===", flush=True)
    torch.manual_seed(SEED)
    np.random.seed(SEED)
    rng_np = np.random.default_rng(SEED)
    X_pre = stack_inputs(pretrain_samples)
    Xt_pre = torch.from_numpy(X_pre).to(DEVICE)
    enc = Encoder(in_ch=2, base=BASE).to(DEVICE)
    proj = ProjectionHead(enc.feat_dim,
                            proj_dim=PROJ_DIM).to(DEVICE)
    opt = torch.optim.AdamW(
        list(enc.parameters()) + list(proj.parameters()),
        lr=1e-3, weight_decay=1e-4)
    n_pre = len(pretrain_samples)
    bs = 16
    losses = []
    for ep in range(SIMCLR_EPOCHS):
        enc.train()
        proj.train()
        perm = np.random.permutation(n_pre)
        ep_loss = 0.0
        n_b = 0
        for i in range(0, n_pre, bs):
            idx = perm[i:i+bs]
            if len(idx) < 2:
                continue
            xb = Xt_pre[idx]
            v1 = random_3d_augment(xb, rng_np)
            v2 = random_3d_augment(xb, rng_np)
            z1 = proj(enc(v1))
            z2 = proj(enc(v2))
            loss = nt_xent_loss(z1, z2)
            opt.zero_grad()
            loss.backward()
            opt.step()
            ep_loss += loss.item()
            n_b += 1
        losses.append(ep_loss / max(1, n_b))
        if (ep + 1) % 5 == 0 or ep == 0:
            print(f"    ep {ep+1}: loss={losses[-1]:.4f}",
                  flush=True)
    print(f"  Pretrain done. Final loss = {losses[-1]:.4f}",
          flush=True)

    # ---- Extract SimCLR features for MU ----
    print(f"\n=== STEP 3: Extract SimCLR features ===",
          flush=True)
    for p in enc.parameters():
        p.requires_grad = False
    enc.eval()
    X_mu_in = stack_inputs(mu_data)
    Xt_mu = torch.from_numpy(X_mu_in).to(DEVICE)
    with torch.no_grad():
        feat_simclr = enc(Xt_mu).cpu().numpy()
    print(f"  SimCLR features shape: {feat_simclr.shape}",
          flush=True)

    # ---- Build feature sets ----
    feats_logistic_arr = np.array(
        [[d["age"], d["idh1"], d["mgmt"],
          d["v_k2"], d["v_k3"], d["v_k4"], d["v_k5"]]
         for d in mu_data], dtype=float)
    print(f"  Multi-σ logistic features shape: "
          f"{feats_logistic_arr.shape}", flush=True)

    # ---- 5-fold CV: 3 variants ----
    print(f"\n=== STEP 4: 5-fold CV ===", flush=True)
    folds = stratified_kfold(y_mu, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}", flush=True)

    n = len(mu_data)
    variants = {
        "multi_sigma_only": feats_logistic_arr,
        "simclr_only": feat_simclr,
        "hybrid_simclr_plus_multi_sigma": np.concatenate(
            [feat_simclr, feats_logistic_arr], axis=1),
    }
    results = {}
    for name, X_full in variants.items():
        print(f"\n  --- variant: {name} (n_feats="
              f"{X_full.shape[1]}) ---", flush=True)
        oof = np.zeros(n, dtype=np.float32)
        assigned = np.zeros(n, dtype=bool)
        fold_aucs = []
        for fold_i, test_idx in enumerate(folds):
            train_idx = np.array(sorted(
                set(range(n)) - set(test_idx.tolist())))
            n_tr_pos = int(y_mu[train_idx].sum())
            n_tr_neg = len(train_idx) - n_tr_pos
            if n_tr_pos < 2 or n_tr_neg < 2:
                continue
            Xtr_raw = X_full[train_idx]
            Xte_raw = X_full[test_idx]
            ytr = y_mu[train_idx]
            yte = y_mu[test_idx]
            # Standardize using train stats
            Xtr, m, s = standardize(Xtr_raw)
            Xte, _, _ = standardize(Xte_raw, means=m, stds=s)
            beta = logistic_fit(Xtr, ytr)
            scores = Xte @ beta
            oof[test_idx] = scores
            assigned[test_idx] = True
            if int(yte.sum()) >= 1 and int(
                    len(yte) - yte.sum()) >= 1:
                a = auroc(scores, yte)
                fold_aucs.append(float(a))
                print(f"    fold {fold_i+1}: AUC={a:.4f}",
                      flush=True)
        pooled_auc = auroc(oof[assigned], y_mu[assigned])
        print(f"  {name} pooled OOF AUC = {pooled_auc:.4f}",
              flush=True)
        results[name] = {
            "n_features": int(X_full.shape[1]),
            "fold_aucs": fold_aucs,
            "fold_mean": float(np.mean(fold_aucs)),
            "pooled_oof_auc": float(pooled_auc),
            "oof_scores": oof.tolist(),
            "assigned": assigned.tolist(),
        }

    # ---- Bootstrap pairwise comparisons ----
    print(f"\n=== STEP 5: Bootstrap pairwise Δ AUC ===",
          flush=True)
    rng = np.random.default_rng(SEED)
    n_assigned = int(np.array(
        results["multi_sigma_only"]["assigned"]).sum())
    deltas = {}
    pairs = [
        ("hybrid_minus_multi_sigma",
         "hybrid_simclr_plus_multi_sigma",
         "multi_sigma_only"),
        ("hybrid_minus_simclr",
         "hybrid_simclr_plus_multi_sigma",
         "simclr_only"),
        ("multi_sigma_minus_simclr", "multi_sigma_only",
         "simclr_only"),
    ]
    for pair_name, name_a, name_b in pairs:
        oof_a = np.array(results[name_a]["oof_scores"])
        oof_b = np.array(results[name_b]["oof_scores"])
        bs = []
        for _ in range(N_BOOT):
            idx = rng.integers(0, n_assigned, size=n_assigned)
            valid_idx = np.where(np.array(
                results[name_a]["assigned"]))[0][idx]
            yb = y_mu[valid_idx]
            if int(yb.sum()) < 3 or int(len(yb) - yb.sum()) < 3:
                continue
            a_a = auroc(oof_a[valid_idx], yb)
            a_b = auroc(oof_b[valid_idx], yb)
            bs.append(a_a - a_b)
        bs = np.array(bs)
        if len(bs) > 0:
            ci_lo = float(np.percentile(bs, 2.5))
            ci_hi = float(np.percentile(bs, 97.5))
            p_one = float((bs <= 0).mean())
            print(f"  Δ {pair_name}: mean="
                  f"{bs.mean():+.4f}, 95% CI ["
                  f"{ci_lo:+.4f}, {ci_hi:+.4f}], "
                  f"P<=0={p_one:.4f}", flush=True)
            deltas[pair_name] = {
                "mean": float(bs.mean()),
                "95_CI": [ci_lo, ci_hi],
                "p_one_sided": p_one,
            }

    # ---- Summary ----
    print(f"\n=== SUMMARY ===", flush=True)
    print(f"  v218 logistic clin+V_k σ=3 (4 feats): AUC=0.728",
          flush=True)
    print(f"  v218 logistic clin+V_k multi-σ (7 feats): "
          f"AUC=0.815", flush=True)
    print(f"  v223 multi_sigma_only logistic 5-fold OOF "
          f"({results['multi_sigma_only']['n_features']} feats): "
          f"{results['multi_sigma_only']['pooled_oof_auc']:.4f}",
          flush=True)
    print(f"  v223 simclr_only logistic 5-fold OOF "
          f"({results['simclr_only']['n_features']} feats): "
          f"{results['simclr_only']['pooled_oof_auc']:.4f}",
          flush=True)
    print(f"  v223 HYBRID logistic 5-fold OOF "
          f"({results['hybrid_simclr_plus_multi_sigma']['n_features']} "
          f"feats): "
          f"{results['hybrid_simclr_plus_multi_sigma']['pooled_oof_auc']:.4f}",
          flush=True)

    out = {
        "version": "v223",
        "experiment": ("SimCLR-feature + multi-σ HYBRID "
                       "logistic vs each individually"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_pretrain_samples": len(pretrain_samples),
        "n_mu": len(mu_data),
        "n_folds": N_FOLDS,
        "simclr_pretrain_loss": [float(x) for x in losses],
        "results": {
            k: {kk: vv for kk, vv in v.items()
                if kk not in ("oof_scores", "assigned")}
            for k, v in results.items()
        },
        "bootstrap_deltas": deltas,
        "comparison_to_prior_rounds": {
            "v202_logistic_clin_Vk_s3": 0.728,
            "v218_logistic_clin_Vk_multi_sigma": 0.815,
            "v215_simclr_per_fold_mean": 0.706,
            "v219_3d_resnet_18": 0.568,
            "v221_3d_vit": 0.599,
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
