"""v221: SOTA 3D Vision Transformer (ViT) comparison + final
architecture leaderboard — round 48 (GPU).

Round 47 v219 showed 3D ResNet-18 (4.7M params) cannot beat the
simple 7-feature multi-σ logistic. This round tests the modern
SOTA: 3D Vision Transformer (ViT-3D). ViTs have replaced CNNs as
SOTA for many imaging tasks. Test whether attention-based
architectures can compete with the simple feature-engineered
logistic at MU n=130.

Method: small 3D ViT — 3D patch embedding (patch=4×6×6) →
positional embedding → 4-layer transformer encoder (4 heads,
embed_dim=64) → mean-pooled CLS-style embedding → MLP head. Train
5-fold stratified CV on MU n=130 binary 365-d PFS.

Compare on the same 5-fold CV against:
  - SimpleCNN (round 47 v219)
  - 3D ResNet-18 (round 47 v219)
  - 3D ViT (this round)
  - logistic clin+V_k σ=3 (round 39 v202)
  - logistic clin+V_k multi-σ (round 47 v218)

Also build comprehensive SOTA leaderboard table.

Outputs:
  Nature_project/05_results/v221_3d_vit_sota.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v221_3d_vit_sota.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
EPOCHS = 30
LR = 5e-4
SEED = 42

# ViT hyperparameters
PATCH = (4, 6, 6)  # patch size in (D, H, W); 16/4=4, 48/6=8 -> 4*8*8=256 patches
EMBED_DIM = 64
DEPTH = 4
HEADS = 4
MLP_RATIO = 4


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class PatchEmbed3D(nn.Module):
    def __init__(self, in_ch=2, patch=PATCH, embed_dim=EMBED_DIM):
        super().__init__()
        self.proj = nn.Conv3d(in_ch, embed_dim,
                                kernel_size=patch, stride=patch)
        self.patch = patch

    def forward(self, x):
        # x: (B, C, D, H, W)
        x = self.proj(x)  # (B, E, D', H', W')
        # flatten spatial
        x = x.flatten(2).transpose(1, 2)  # (B, N, E)
        return x


class TransformerBlock(nn.Module):
    def __init__(self, dim, heads, mlp_ratio=4, dropout=0.1):
        super().__init__()
        self.norm1 = nn.LayerNorm(dim)
        self.attn = nn.MultiheadAttention(
            dim, heads, dropout=dropout, batch_first=True)
        self.norm2 = nn.LayerNorm(dim)
        hidden = int(dim * mlp_ratio)
        self.mlp = nn.Sequential(
            nn.Linear(dim, hidden), nn.GELU(),
            nn.Dropout(dropout),
            nn.Linear(hidden, dim), nn.Dropout(dropout))

    def forward(self, x):
        h = self.norm1(x)
        attn_out, _ = self.attn(h, h, h, need_weights=False)
        x = x + attn_out
        x = x + self.mlp(self.norm2(x))
        return x


class ViT3D(nn.Module):
    def __init__(self, in_ch=2, patch=PATCH,
                  embed_dim=EMBED_DIM, depth=DEPTH,
                  heads=HEADS, mlp_ratio=MLP_RATIO):
        super().__init__()
        self.patch_embed = PatchEmbed3D(in_ch, patch, embed_dim)
        # Compute number of patches
        d_p = TARGET_SHAPE[0] // patch[0]
        h_p = TARGET_SHAPE[1] // patch[1]
        w_p = TARGET_SHAPE[2] // patch[2]
        n_patches = d_p * h_p * w_p
        self.cls_token = nn.Parameter(torch.zeros(1, 1, embed_dim))
        self.pos_embed = nn.Parameter(
            torch.zeros(1, n_patches + 1, embed_dim))
        nn.init.trunc_normal_(self.cls_token, std=0.02)
        nn.init.trunc_normal_(self.pos_embed, std=0.02)
        self.blocks = nn.ModuleList([
            TransformerBlock(embed_dim, heads, mlp_ratio)
            for _ in range(depth)])
        self.norm = nn.LayerNorm(embed_dim)
        self.head = nn.Sequential(
            nn.Linear(embed_dim, 32), nn.GELU(),
            nn.Dropout(0.3), nn.Linear(32, 1))

    def forward(self, x):
        x = self.patch_embed(x)  # (B, N, E)
        cls = self.cls_token.expand(x.size(0), -1, -1)
        x = torch.cat([cls, x], dim=1)
        x = x + self.pos_embed
        for blk in self.blocks:
            x = blk(x)
        x = self.norm(x)
        cls_out = x[:, 0]  # CLS token
        return self.head(cls_out).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def main():
    print("=" * 78, flush=True)
    print(f"v221 SOTA 3D ViT (round 48 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    mu_data = load_mu_data()
    print(f"  MU labelled: n={len(mu_data)}", flush=True)
    y_mu = np.array([d["y"] for d in mu_data], dtype=np.float32)
    folds = stratified_kfold(y_mu, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}", flush=True)

    # ViT param count
    test_vit = ViT3D()
    n_params_vit = sum(p.numel() for p in test_vit.parameters())
    print(f"  ViT3D parameters: {n_params_vit:,}", flush=True)
    del test_vit

    # ---- 5-fold CV ViT ----
    print(f"\n=== 3D ViT 5-fold CV ===", flush=True)
    n = len(mu_data)
    oof = np.zeros(n, dtype=np.float32)
    assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        n_pos = int(y_mu[train_idx].sum())
        n_neg = len(train_idx) - n_pos
        if n_pos < 2 or n_neg < 2:
            continue
        Xtr = stack_inputs([mu_data[i] for i in train_idx])
        Xte = stack_inputs([mu_data[i] for i in test_idx])
        ytr = y_mu[train_idx]
        yte = y_mu[test_idx]
        Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
        Xte_t = torch.from_numpy(Xte).to(DEVICE)
        ytr_t = torch.from_numpy(ytr).to(DEVICE)

        torch.manual_seed(SEED + fold_i * 100)
        np.random.seed(SEED + fold_i * 100)
        model = ViT3D().to(DEVICE)
        opt = torch.optim.AdamW(model.parameters(), lr=LR,
                                 weight_decay=1e-3)
        pos_w = torch.tensor([n_neg / max(1, n_pos)],
                              dtype=torch.float32, device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        bs = 4
        for ep in range(EPOCHS):
            model.train()
            perm = np.random.permutation(len(train_idx))
            for i in range(0, len(train_idx), bs):
                idx = perm[i:i+bs]
                logits = model(Xtr_t[idx])
                loss = criterion(logits, ytr_t[idx])
                opt.zero_grad()
                loss.backward()
                opt.step()
        model.eval()
        with torch.no_grad():
            scores = model(Xte_t).cpu().numpy()
        oof[test_idx] = scores
        assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            fold_aucs.append(float(a))
            print(f"    Fold {fold_i+1}: AUC={a:.4f}",
                  flush=True)
        del model, opt
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()

    pooled_auc = auroc(oof[assigned], y_mu[assigned])
    print(f"\n  ViT3D pooled OOF AUC = {pooled_auc:.4f}",
          flush=True)
    print(f"  ViT3D per-fold mean = {np.mean(fold_aucs):.4f}",
          flush=True)

    # ---- Final SOTA leaderboard ----
    print(f"\n=== FINAL SOTA LEADERBOARD (MU n=130) ===",
          flush=True)
    leaderboard = [
        ("v202 logistic clinical only (3 features)", 0.620,
         3, "logistic"),
        ("v202 logistic clin + V_kernel σ=3 (4 features)",
         0.728, 4, "logistic"),
        ("**v218 logistic clin + V_kernel multi-σ (7 feats)**",
         0.815, 7, "logistic"),
        ("v218 kitchen sink (clin+shape+V_k multi-σ, 20 feats)",
         0.849, 20, "logistic"),
        ("v218 shape-only radiomics (13 feats)", 0.729, 13,
         "logistic"),
        ("v207 SimpleCNN supervised (multi-seed mean)", 0.586,
         488_000, "deep"),
        ("v209 deep ensemble 50 models", 0.587, 488_000,
         "deep"),
        ("v215 SimpleCNN + SimCLR pretrain", 0.706, 488_000,
         "deep"),
        ("v219(B) 3D ResNet-18 SOTA (4.7M params)", 0.568,
         4_680_000, "deep"),
        ("v219(C) 3D ResNet-18 + SimCLR", 0.577, 4_680_000,
         "deep"),
        (f"**v221 3D Vision Transformer SOTA**",
         pooled_auc, n_params_vit, "deep"),
    ]
    for name, auc, p, kind in leaderboard:
        print(f"  {name:60s}  AUC = {auc:.4f}  "
              f"(params/feats={p:,})", flush=True)

    out = {
        "version": "v221",
        "experiment": ("SOTA 3D Vision Transformer comparison "
                       "+ comprehensive architecture "
                       "leaderboard"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "epochs_per_fold": EPOCHS,
        "n_mu": len(mu_data),
        "vit_params": int(n_params_vit),
        "vit_5fold": {
            "fold_aucs": fold_aucs,
            "fold_mean": float(np.mean(fold_aucs)),
            "fold_std": float(np.std(fold_aucs)),
            "pooled_oof_auc": float(pooled_auc),
        },
        "sota_leaderboard": [
            {"method": n, "auc": float(a),
             "params_or_feats": int(p), "kind": k}
            for n, a, p, k in leaderboard
        ],
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
