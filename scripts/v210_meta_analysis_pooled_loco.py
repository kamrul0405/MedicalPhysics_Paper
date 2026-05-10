"""v210: Inverse-variance meta-analysis + reverse-direction LOCO +
pooled MU+RHUH training + power analysis — round 43 (CPU).

Round 42 v208 left the cross-cohort question inconclusive: the
+0.107 MU-internal effect did not replicate on RHUH-GBM n=31
(Delta=-0.005, CI [-0.197, +0.239]). At Nature/Lancet level we now
disambiguate cohort-specificity vs sample-size limitation via four
complementary analyses.

Method:
  1. REPLICATE BOTH DIRECTIONS: train MU -> test RHUH (v208 result),
     plus train RHUH -> test MU (the reverse direction). If kernel
     signal is real, BOTH directions should show positive Delta;
     if cohort-specific, only MU-internal should.
  2. POOLED 5-FOLD CV with cohort-stratification (each fold has
     both cohorts in train and test). Report pooled Delta AUC +
     bootstrap CI.
  3. INVERSE-VARIANCE-WEIGHTED META-ANALYSIS combining MU in-
     sample and RHUH external Delta estimates. Bootstrap-derived
     variance per cohort -> pooled Delta with tighter CI.
  4. POWER ANALYSIS: at sample sizes n in {31, 50, 100, 200, 500},
     the minimum detectable effect size at alpha=0.05, beta=0.20
     for binary AUC comparison.

Outputs:
  Nature_project/05_results/v210_meta_pooled_power.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize
from scipy.stats import norm

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v210_meta_pooled_power.json"

SIGMA_KERNEL = 3.0
HORIZON = 365
N_BOOTSTRAPS = 1000
N_FOLDS = 5
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def build_X(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def label_binary(rows, X, pfs_field="pfs_days",
                  prog_field="progress"):
    out = []
    for r in rows:
        pfs = r[pfs_field]
        prog = r[prog_field]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def load_mu():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1,
        }
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid, "cohort": "MU",
            "v_kernel_s3": kernel_outgrowth_volume(m,
                                                    SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def load_rhuh():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        age_col = next((c for c in cols if c.strip() == "Age"),
                        None)
        idh_col = next((c for c in cols if "IDH status" in c),
                        None)
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                age = (float(row[age_col]) if row[age_col]
                       else None)
                idh_str = (row[idh_col].strip().lower()
                            if row[idh_col] else "")
                if "mut" in idh_str:
                    idh1 = 1
                elif "wt" in idh_str or "wild" in idh_str:
                    idh1 = 0
                else:
                    idh1 = None
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if age is None or idh1 is None or pfs_d is None:
                continue
            clinical[pid] = {
                "age": age, "idh1": idh1,
                "pfs_days": pfs_d, "progress": event,
            }
    rows = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid, "cohort": "RHUH",
            "v_kernel_s3": kernel_outgrowth_volume(m,
                                                    SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def stratified_kfold(y, k, seed=RNG_SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def cohort_stratified_kfold(rows, k, seed=RNG_SEED):
    """Stratify by (cohort, label) so every fold has both."""
    rng = np.random.default_rng(seed)
    folds = [[] for _ in range(k)]
    for cohort_val in ["MU", "RHUH"]:
        for label in [0, 1]:
            idx = [i for i, r in enumerate(rows)
                   if r["cohort"] == cohort_val
                   and r["y"] == label]
            rng.shuffle(idx)
            for j, idx_val in enumerate(idx):
                folds[j % k].append(int(idx_val))
    return [np.array(sorted(f)) for f in folds]


def main():
    print("=" * 78, flush=True)
    print("v210 META-ANALYSIS + POOLED LOCO + POWER (round 43 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # ---- Load both cohorts ----
    mu_rows = load_mu()
    rhuh_rows = load_rhuh()
    mu_lab = label_binary(mu_rows, HORIZON)
    rhuh_lab = label_binary(rhuh_rows, HORIZON)
    print(f"  MU: {len(mu_lab)} labelled "
          f"({sum(1 for l in mu_lab if l['y']==1)} pos, "
          f"{sum(1 for l in mu_lab if l['y']==0)} neg)",
          flush=True)
    print(f"  RHUH: {len(rhuh_lab)} labelled "
          f"({sum(1 for l in rhuh_lab if l['y']==1)} pos, "
          f"{sum(1 for l in rhuh_lab if l['y']==0)} neg)",
          flush=True)

    feats_clin = ["age", "idh1"]
    feats_full = feats_clin + ["v_kernel_s3"]

    mu_cc = [l for l in mu_lab if all(l[f] is not None
                                       for f in feats_full)]
    rhuh_cc = [l for l in rhuh_lab if all(l[f] is not None
                                            for f in feats_full)]
    print(f"  MU complete-case: {len(mu_cc)}", flush=True)
    print(f"  RHUH complete-case: {len(rhuh_cc)}", flush=True)

    # ============================================================
    # PART 1: Both directions of cross-cohort validation
    # ============================================================
    print(f"\n=== PART 1: CROSS-COHORT BOTH DIRECTIONS ===",
          flush=True)
    direction_results = {}
    for tr_name, tr_rows, te_name, te_rows in [
        ("MU", mu_cc, "RHUH", rhuh_cc),
        ("RHUH", rhuh_cc, "MU", mu_cc),
    ]:
        print(f"\n  TRAIN {tr_name} (n={len(tr_rows)}) -> "
              f"TEST {te_name} (n={len(te_rows)})", flush=True)
        y_tr = np.array([l["y"] for l in tr_rows], dtype=float)
        y_te = np.array([l["y"] for l in te_rows], dtype=float)
        X_tr_c, m_c, s_c = build_X(tr_rows, feats_clin)
        X_tr_f, m_f, s_f = build_X(tr_rows, feats_full)
        beta_c = logistic_fit(X_tr_c, y_tr)
        beta_f = logistic_fit(X_tr_f, y_tr)
        # In-sample
        in_auc_c = auroc(X_tr_c @ beta_c, y_tr)
        in_auc_f = auroc(X_tr_f @ beta_f, y_tr)
        # Apply on test cohort using train-derived standardization
        X_te_c, _, _ = build_X(te_rows, feats_clin,
                                 means=m_c, stds=s_c)
        X_te_f, _, _ = build_X(te_rows, feats_full,
                                 means=m_f, stds=s_f)
        ext_auc_c = auroc(X_te_c @ beta_c, y_te)
        ext_auc_f = auroc(X_te_f @ beta_f, y_te)
        delta_ext = ext_auc_f - ext_auc_c
        # Bootstrap on test cohort
        deltas = []
        for _ in range(N_BOOTSTRAPS):
            idx = rng.integers(0, len(te_rows),
                                size=len(te_rows))
            yb = y_te[idx]
            if int(yb.sum()) < 2 or int(len(yb) - yb.sum()) < 2:
                continue
            ac = auroc(X_te_c[idx] @ beta_c, yb)
            af = auroc(X_te_f[idx] @ beta_f, yb)
            deltas.append(af - ac)
        deltas = np.array(deltas)
        ci_lo = float(np.percentile(deltas, 2.5))
        ci_hi = float(np.percentile(deltas, 97.5))
        var_d = float(deltas.var())
        p_one = float((deltas <= 0).mean())
        print(f"    In-sample {tr_name}: clin={in_auc_c:.4f}, "
              f"full={in_auc_f:.4f}, "
              f"Delta={in_auc_f - in_auc_c:+.4f}", flush=True)
        print(f"    External {te_name}: clin={ext_auc_c:.4f}, "
              f"full={ext_auc_f:.4f}, Delta={delta_ext:+.4f}",
              flush=True)
        print(f"    Bootstrap (n={len(deltas)}): mean="
              f"{deltas.mean():+.4f}, var={var_d:.5f}, "
              f"95% CI [{ci_lo:+.4f}, {ci_hi:+.4f}], "
              f"P<=0={p_one:.3f}", flush=True)
        direction_results[f"train_{tr_name}_test_{te_name}"] = {
            "n_train": len(tr_rows),
            "n_test": len(te_rows),
            "in_sample_auc_clin": float(in_auc_c),
            "in_sample_auc_full": float(in_auc_f),
            "in_sample_delta": float(in_auc_f - in_auc_c),
            "external_auc_clin": float(ext_auc_c),
            "external_auc_full": float(ext_auc_f),
            "external_delta_point": float(delta_ext),
            "external_delta_bootstrap_mean": float(deltas.mean()),
            "external_delta_variance": var_d,
            "external_delta_95_CI": [ci_lo, ci_hi],
            "p_one_sided": p_one,
            "n_bootstrap_valid": len(deltas),
        }

    # ============================================================
    # PART 2: Pooled MU+RHUH + cohort-stratified 5-fold CV
    # ============================================================
    print(f"\n=== PART 2: POOLED MU+RHUH 5-FOLD CV ===",
          flush=True)
    pooled = mu_cc + rhuh_cc
    print(f"  Pooled n = {len(pooled)} "
          f"({sum(1 for l in pooled if l['y']==1)} pos, "
          f"{sum(1 for l in pooled if l['y']==0)} neg)",
          flush=True)
    folds = cohort_stratified_kfold(pooled, N_FOLDS,
                                     seed=RNG_SEED)
    print(f"  Cohort-stratified 5-fold sizes: "
          f"{[len(f) for f in folds]}", flush=True)
    pooled_clin_oof = np.zeros(len(pooled))
    pooled_full_oof = np.zeros(len(pooled))
    pooled_assigned = np.zeros(len(pooled), dtype=bool)
    y_pool = np.array([l["y"] for l in pooled], dtype=float)
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(len(pooled))) - set(test_idx.tolist())))
        tr_rows_f = [pooled[i] for i in train_idx]
        te_rows_f = [pooled[i] for i in test_idx]
        y_tr_f = np.array([l["y"] for l in tr_rows_f],
                            dtype=float)
        X_tr_c, m_c, s_c = build_X(tr_rows_f, feats_clin)
        X_tr_f, m_f, s_f = build_X(tr_rows_f, feats_full)
        beta_c = logistic_fit(X_tr_c, y_tr_f)
        beta_f = logistic_fit(X_tr_f, y_tr_f)
        X_te_c, _, _ = build_X(te_rows_f, feats_clin,
                                 means=m_c, stds=s_c)
        X_te_f, _, _ = build_X(te_rows_f, feats_full,
                                 means=m_f, stds=s_f)
        pooled_clin_oof[test_idx] = X_te_c @ beta_c
        pooled_full_oof[test_idx] = X_te_f @ beta_f
        pooled_assigned[test_idx] = True
    pooled_auc_c = auroc(pooled_clin_oof[pooled_assigned],
                          y_pool[pooled_assigned])
    pooled_auc_f = auroc(pooled_full_oof[pooled_assigned],
                          y_pool[pooled_assigned])
    pooled_delta = pooled_auc_f - pooled_auc_c
    print(f"  Pooled OOF AUC: clin={pooled_auc_c:.4f}, "
          f"full={pooled_auc_f:.4f}, "
          f"Delta={pooled_delta:+.4f}", flush=True)

    # Bootstrap on pooled OOF
    deltas_pooled = []
    n_pool = len(pooled)
    for _ in range(N_BOOTSTRAPS):
        idx = rng.integers(0, n_pool, size=n_pool)
        yb = y_pool[idx]
        if int(yb.sum()) < 5 or int(len(yb) - yb.sum()) < 5:
            continue
        ac = auroc(pooled_clin_oof[idx], yb)
        af = auroc(pooled_full_oof[idx], yb)
        deltas_pooled.append(af - ac)
    deltas_pooled = np.array(deltas_pooled)
    ci_p_lo = float(np.percentile(deltas_pooled, 2.5))
    ci_p_hi = float(np.percentile(deltas_pooled, 97.5))
    p_p = float((deltas_pooled <= 0).mean())
    print(f"  Bootstrap pooled (n={len(deltas_pooled)}): "
          f"Delta_mean={deltas_pooled.mean():+.4f}, "
          f"95% CI [{ci_p_lo:+.4f}, {ci_p_hi:+.4f}], "
          f"P<=0={p_p:.4f}", flush=True)

    # Per-cohort breakdown of pooled-CV AUC
    by_cohort = {}
    for cohort_val in ["MU", "RHUH"]:
        sel = np.array([r["cohort"] == cohort_val
                          for r in pooled])
        if sel.sum() < 5:
            continue
        ys = y_pool[sel]
        if ys.sum() < 2 or len(ys) - ys.sum() < 2:
            continue
        ac = auroc(pooled_clin_oof[sel], ys)
        af = auroc(pooled_full_oof[sel], ys)
        print(f"    [pooled-CV {cohort_val} subset n={int(sel.sum())}] "
              f"AUC_clin={ac:.4f}, AUC_full={af:.4f}, "
              f"Delta={af-ac:+.4f}", flush=True)
        by_cohort[cohort_val] = {
            "n": int(sel.sum()),
            "n_pos": int(ys.sum()),
            "auc_clin": float(ac),
            "auc_full": float(af),
            "delta_point": float(af - ac),
        }

    # ============================================================
    # PART 3: Inverse-variance-weighted meta-analysis
    # ============================================================
    print(f"\n=== PART 3: INVERSE-VARIANCE META-ANALYSIS ===",
          flush=True)
    # Pull MU in-sample variance from bootstrap on MU
    # (rerun bootstrap for MU in-sample for proper variance)
    mu_y = np.array([l["y"] for l in mu_cc], dtype=float)
    X_mu_c, mc_m, mc_s = build_X(mu_cc, feats_clin)
    X_mu_f, mf_m, mf_s = build_X(mu_cc, feats_full)
    bc_mu = logistic_fit(X_mu_c, mu_y)
    bf_mu = logistic_fit(X_mu_f, mu_y)
    mu_deltas = []
    for _ in range(N_BOOTSTRAPS):
        idx = rng.integers(0, len(mu_cc), size=len(mu_cc))
        yb = mu_y[idx]
        if int(yb.sum()) < 5 or int(len(yb) - yb.sum()) < 5:
            continue
        # refit on bootstrap sample
        rb = [mu_cc[i] for i in idx]
        Xc, m_c, s_c = build_X(rb, feats_clin)
        Xf, m_f, s_f = build_X(rb, feats_full)
        bc = logistic_fit(Xc, yb)
        bf = logistic_fit(Xf, yb)
        ac = auroc(Xc @ bc, yb)
        af = auroc(Xf @ bf, yb)
        mu_deltas.append(af - ac)
    mu_deltas = np.array(mu_deltas)
    mu_var = float(mu_deltas.var())
    mu_mean = float(mu_deltas.mean())
    print(f"  MU in-sample bootstrap: mean={mu_mean:+.4f}, "
          f"var={mu_var:.6f}, "
          f"95% CI [{np.percentile(mu_deltas, 2.5):+.4f}, "
          f"{np.percentile(mu_deltas, 97.5):+.4f}]",
          flush=True)

    # RHUH external variance from PART 1
    rhuh_d = direction_results["train_MU_test_RHUH"]
    rhuh_mean = rhuh_d["external_delta_bootstrap_mean"]
    rhuh_var = rhuh_d["external_delta_variance"]
    print(f"  RHUH external bootstrap: mean={rhuh_mean:+.4f}, "
          f"var={rhuh_var:.6f}", flush=True)

    # Inverse-variance-weighted pooled estimate
    w_mu = 1.0 / mu_var if mu_var > 0 else 0.0
    w_rhuh = 1.0 / rhuh_var if rhuh_var > 0 else 0.0
    pooled_meta = (w_mu * mu_mean + w_rhuh * rhuh_mean) / (
        w_mu + w_rhuh)
    pooled_meta_se = np.sqrt(1.0 / (w_mu + w_rhuh))
    pooled_meta_ci = [pooled_meta - 1.96 * pooled_meta_se,
                       pooled_meta + 1.96 * pooled_meta_se]
    pooled_meta_z = pooled_meta / pooled_meta_se
    pooled_meta_p = 1 - norm.cdf(pooled_meta_z)
    print(f"  IV-weighted meta-analysis pooled Delta = "
          f"{pooled_meta:+.4f} (SE={pooled_meta_se:.4f}), "
          f"95% CI [{pooled_meta_ci[0]:+.4f}, "
          f"{pooled_meta_ci[1]:+.4f}], "
          f"z={pooled_meta_z:.3f}, "
          f"one-sided P={pooled_meta_p:.4f}", flush=True)

    # ============================================================
    # PART 4: POWER ANALYSIS
    # ============================================================
    print(f"\n=== PART 4: POWER ANALYSIS ===", flush=True)
    # For binary AUC comparison, approximate SE(Delta AUC) under H0
    # using formulas for paired AUC. Here we use simple Hanley-McNeil
    # approximation for the SE of a single AUC under random equal-
    # variance assumption: SE(AUC) ~ sqrt(AUC*(1-AUC)/n_eff)
    # where n_eff = harmonic mean of (n_pos, n_neg). We then take
    # the bootstrap-derived RHUH variance scale ~ const/n.
    # Power: minimum detectable effect size at alpha=0.05 (one-
    # sided), beta=0.20 = z_alpha + z_beta = 1.645 + 0.842 = 2.487
    z_total = 1.645 + 0.842
    # Empirical: at n=31 RHUH, var ~ rhuh_var
    # so SE(Delta) ~ sqrt(rhuh_var)
    rhuh_se_31 = np.sqrt(rhuh_var)
    # Scale by sqrt(31/n) to extrapolate
    scaling_n = 31
    power_results = {}
    for n_target in [31, 50, 100, 150, 200, 300, 500]:
        se_n = rhuh_se_31 * np.sqrt(scaling_n / n_target)
        mde = z_total * se_n
        power_at_d108 = norm.cdf(0.108 / se_n - 1.645)
        # power at MU effect size 0.108
        print(f"  n={n_target}: SE(Delta)={se_n:.4f}, "
              f"MDE (alpha=0.05, beta=0.20) = {mde:.4f}, "
              f"power at Delta=0.108 = {power_at_d108:.3f}",
              flush=True)
        power_results[n_target] = {
            "se_delta_estimated": float(se_n),
            "mde_alpha05_beta20": float(mde),
            "power_at_delta_0_108": float(power_at_d108),
        }

    out = {
        "version": "v210",
        "experiment": ("Inverse-variance meta-analysis + reverse-"
                       "direction LOCO + pooled MU+RHUH 5-fold "
                       "CV + power analysis"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "n_bootstraps": N_BOOTSTRAPS,
        "n_folds": N_FOLDS,
        "cohorts": {
            "MU": {"n_complete": len(mu_cc),
                    "n_pos": sum(1 for l in mu_cc
                                  if l["y"] == 1)},
            "RHUH": {"n_complete": len(rhuh_cc),
                      "n_pos": sum(1 for l in rhuh_cc
                                    if l["y"] == 1)},
        },
        "loco_directions": direction_results,
        "pooled_5fold_cv": {
            "n_pooled": len(pooled),
            "auc_clin_oof": float(pooled_auc_c),
            "auc_full_oof": float(pooled_auc_f),
            "delta_point": float(pooled_delta),
            "delta_bootstrap_mean": float(deltas_pooled.mean()),
            "delta_95_CI": [ci_p_lo, ci_p_hi],
            "p_one_sided": p_p,
            "by_cohort_subset": by_cohort,
        },
        "meta_analysis": {
            "mu_mean": mu_mean,
            "mu_variance": mu_var,
            "rhuh_mean": rhuh_mean,
            "rhuh_variance": rhuh_var,
            "weight_mu": float(w_mu),
            "weight_rhuh": float(w_rhuh),
            "pooled_delta": float(pooled_meta),
            "pooled_se": float(pooled_meta_se),
            "pooled_95_CI": pooled_meta_ci,
            "z_score": float(pooled_meta_z),
            "one_sided_p": float(pooled_meta_p),
        },
        "power_analysis": power_results,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
