"""v215: Self-supervised SimCLR pretraining on multi-cohort baseline
masks → fine-tune binary 365-d PFS head on MU — round 45 (GPU).

Round 44 v213 showed supervised pretraining on MU (with PFS labels)
+ frozen encoder + head fine-tune on RHUH gives AUC=0.804. Open
question: can LABEL-FREE self-supervised pretraining on raw masks
across all 7 cohorts produce a useful encoder for binary PFS?

This is the "use unlabeled data" Nature/Lancet flagship: self-
supervised representation learning leverages all available data
(no PFS needed) for the encoder, and the (small) labelled binary-
PFS data is used only to fit the head.

Method:
  1. Load baseline tumor masks from MU + RHUH + UCSF + LUMIERE +
     UPENN + PROTEAS cohorts (all available cache_3d files). No
     labels needed.
  2. Compute bimodal kernel (sigma=3) for each. Channel input =
     [mask, kernel] (2 ch).
  3. SimCLR-style contrastive pretraining: for each mask sample,
     two augmented views (random rotation, flip, scale, intensity
     noise). NT-Xent loss with temperature 0.5, projection head
     2-layer MLP. Pretrain encoder for 50 epochs.
  4. Freeze encoder weights.
  5. Fine-tune a new head on MU n=130 binary 365-d PFS, 5-fold
     stratified CV.
  6. Compare to v213 (supervised MU pretraining): is label-free
     pretraining as good?

Outputs:
  Nature_project/05_results/v215_simclr_pretrain.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v215_simclr_pretrain.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
PRETRAIN_EPOCHS = 40
FT_EPOCHS = 50
LR_PRE = 1e-3
LR_FT = 1e-3
TEMPERATURE = 0.5
PROJ_DIM = 64
BASE = 24
SEED = 42

COHORT_PREFIXES = ["MU-Glioma-Post", "RHUH-GBM", "UCSF-POSTOP",
                    "LUMIERE", "UPENN-GBM", "PROTEAS-brain-mets",
                    "Yale-Brain-Mets-Longitudinal"]


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class Encoder(nn.Module):
    def __init__(self, in_ch=2, base=BASE):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.feat_dim = base * 4

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.global_pool(e3).flatten(1)


class ProjectionHead(nn.Module):
    def __init__(self, feat_dim=BASE * 4, proj_dim=PROJ_DIM):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, feat_dim),
            nn.GELU(),
            nn.Linear(feat_dim, proj_dim),
        )

    def forward(self, feat):
        z = self.net(feat)
        return F.normalize(z, dim=-1)


class BinaryHead(nn.Module):
    def __init__(self, feat_dim=BASE * 4):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, 32), nn.GELU(),
            nn.Dropout(0.3), nn.Linear(32, 1))

    def forward(self, feat):
        return self.net(feat).squeeze(-1)


def random_3d_augment(x, rng):
    """x: torch tensor (B, C, D, H, W). Apply random flips, axis
    swap, intensity noise."""
    out = x.clone()
    # random flips along each axis with prob 0.5
    for axis in [2, 3, 4]:
        if rng.random() > 0.5:
            out = torch.flip(out, dims=[axis])
    # random intensity noise (multiplicative in [0.9, 1.1])
    scale = float(rng.uniform(0.9, 1.1))
    out = out * scale
    # additive Gaussian noise
    noise_std = 0.02
    out = out + noise_std * torch.randn_like(out)
    out = torch.clamp(out, 0.0, 1.5)
    return out


def nt_xent_loss(z1, z2, temperature=TEMPERATURE):
    """SimCLR NT-Xent loss for two view batches."""
    n = z1.size(0)
    z = torch.cat([z1, z2], dim=0)  # (2n, d)
    sim = torch.matmul(z, z.T) / temperature  # (2n, 2n)
    mask = torch.eye(2 * n, dtype=torch.bool,
                      device=sim.device)
    sim = sim.masked_fill(mask, float("-inf"))
    # For sample i in [0, n), positive is i+n (and vice versa)
    targets = torch.cat([torch.arange(n, 2 * n),
                          torch.arange(0, n)],
                          dim=0).to(sim.device)
    return F.cross_entropy(sim, targets)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def load_all_masks_for_pretrain():
    """Load all baseline masks from all cohorts, no labels."""
    samples = []
    for prefix in COHORT_PREFIXES:
        files = sorted(CACHE.glob(f"{prefix}_*_b.npy"))
        for f in files:
            try:
                m_full = (np.load(f) > 0).astype(np.float32)
                if m_full.sum() == 0:
                    continue
                m = resize_to_target(m_full, TARGET_SHAPE)
                k = heat_bimodal(m, SIGMA_KERNEL)
                samples.append({
                    "pid": f.stem,
                    "cohort": prefix,
                    "mask": m, "kernel": k,
                })
            except Exception:
                continue
    return samples


def load_mu_labelled():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def main():
    print("=" * 78, flush=True)
    print(f"v215 SIMCLR PRETRAIN (round 45 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()
    rng_np = np.random.default_rng(SEED)
    torch.manual_seed(SEED)

    # ---- Step 1: Load all masks (label-free) ----
    print(f"\n=== STEP 1: load multi-cohort masks ===",
          flush=True)
    pretrain_samples = load_all_masks_for_pretrain()
    print(f"  Loaded {len(pretrain_samples)} masks across "
          f"{len(set(s['cohort'] for s in pretrain_samples))} "
          f"cohorts", flush=True)
    cohort_counts = {}
    for s in pretrain_samples:
        cohort_counts[s["cohort"]] = cohort_counts.get(
            s["cohort"], 0) + 1
    for c, n in cohort_counts.items():
        print(f"    {c}: {n}", flush=True)

    # ---- Step 2: SimCLR pretraining ----
    print(f"\n=== STEP 2: SimCLR pretraining "
          f"({PRETRAIN_EPOCHS} epochs) ===", flush=True)
    X_pre = stack_inputs(pretrain_samples)
    Xt_pre = torch.from_numpy(X_pre).to(DEVICE)
    enc = Encoder(in_ch=2, base=BASE).to(DEVICE)
    proj = ProjectionHead(feat_dim=enc.feat_dim,
                            proj_dim=PROJ_DIM).to(DEVICE)
    opt = torch.optim.AdamW(
        list(enc.parameters()) + list(proj.parameters()),
        lr=LR_PRE, weight_decay=1e-4)
    n_pre = len(pretrain_samples)
    bs = 16
    losses = []
    for ep in range(PRETRAIN_EPOCHS):
        enc.train()
        proj.train()
        perm = np.random.permutation(n_pre)
        ep_loss = 0.0
        n_batches = 0
        for i in range(0, n_pre, bs):
            idx = perm[i:i+bs]
            if len(idx) < 2:
                continue
            x_batch = Xt_pre[idx]
            v1 = random_3d_augment(x_batch, rng_np)
            v2 = random_3d_augment(x_batch, rng_np)
            f1 = enc(v1)
            f2 = enc(v2)
            z1 = proj(f1)
            z2 = proj(f2)
            loss = nt_xent_loss(z1, z2)
            opt.zero_grad()
            loss.backward()
            opt.step()
            ep_loss += loss.item()
            n_batches += 1
        avg_loss = ep_loss / max(1, n_batches)
        losses.append(avg_loss)
        if (ep + 1) % 5 == 0 or ep == 0:
            print(f"    ep {ep+1}/{PRETRAIN_EPOCHS}  "
                  f"loss={avg_loss:.4f}", flush=True)

    print(f"  Pretrain done. Final loss = {losses[-1]:.4f}",
          flush=True)

    # ---- Step 3: 5-fold CV on MU using frozen SimCLR encoder ----
    print(f"\n=== STEP 3: 5-fold CV on MU with frozen SimCLR "
          f"encoder ===", flush=True)
    mu_data = load_mu_labelled()
    print(f"  MU labelled: n={len(mu_data)}", flush=True)
    y_mu = np.array([d["y"] for d in mu_data], dtype=np.float32)
    folds = stratified_kfold(y_mu, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}", flush=True)

    X_mu = stack_inputs(mu_data)
    Xt_mu = torch.from_numpy(X_mu).to(DEVICE)
    # Precompute features (encoder frozen)
    for p in enc.parameters():
        p.requires_grad = False
    enc.eval()
    with torch.no_grad():
        feat_mu = enc(Xt_mu)
    print(f"  Encoded {len(mu_data)} MU patients", flush=True)

    oof_scores = np.zeros(len(mu_data), dtype=np.float32)
    oof_assigned = np.zeros(len(mu_data), dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(len(mu_data))) - set(test_idx.tolist())))
        n_tr_pos = int(y_mu[train_idx].sum())
        n_tr_neg = len(train_idx) - n_tr_pos
        if n_tr_pos < 2 or n_tr_neg < 2:
            continue
        feat_tr = feat_mu[train_idx]
        feat_te = feat_mu[test_idx]
        ytr = torch.from_numpy(y_mu[train_idx]).to(DEVICE)
        yte = y_mu[test_idx]

        torch.manual_seed(SEED + fold_i * 100)
        head = BinaryHead(feat_dim=enc.feat_dim).to(DEVICE)
        opt_h = torch.optim.AdamW(head.parameters(), lr=LR_FT,
                                    weight_decay=1e-3)
        pos_w = torch.tensor([n_tr_neg / max(1, n_tr_pos)],
                              dtype=torch.float32, device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        n_tr = len(train_idx)
        bs_ft = min(8, n_tr)
        for ep in range(FT_EPOCHS):
            head.train()
            perm = np.random.permutation(n_tr)
            for i in range(0, n_tr, bs_ft):
                idx = perm[i:i+bs_ft]
                logits = head(feat_tr[idx])
                loss = criterion(logits, ytr[idx])
                opt_h.zero_grad()
                loss.backward()
                opt_h.step()
        head.eval()
        with torch.no_grad():
            scores = head(feat_te).cpu().numpy()
        oof_scores[test_idx] = scores
        oof_assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            fold_aucs.append(float(a))
            print(f"    Fold {fold_i+1}/{N_FOLDS}: AUC={a:.4f}",
                  flush=True)
    pooled_simclr_auc = auroc(oof_scores[oof_assigned],
                                y_mu[oof_assigned])
    print(f"\n  SimCLR pretrained pooled OOF AUC = "
          f"{pooled_simclr_auc:.4f}, "
          f"per-fold mean = {np.mean(fold_aucs):.4f}",
          flush=True)

    # Bootstrap CI
    rng = np.random.default_rng(SEED)
    boot_aucs = []
    n_assigned = int(oof_assigned.sum())
    for _ in range(200):
        idx = rng.integers(0, n_assigned, size=n_assigned)
        valid = np.where(oof_assigned)[0][idx]
        yb = y_mu[valid]
        if int(yb.sum()) < 2 or int(len(yb) - yb.sum()) < 2:
            continue
        boot_aucs.append(auroc(oof_scores[valid], yb))
    boot_aucs = np.array(boot_aucs)
    print(f"  Bootstrap (200): mean={boot_aucs.mean():.4f}, "
          f"95% CI [{np.percentile(boot_aucs, 2.5):.4f}, "
          f"{np.percentile(boot_aucs, 97.5):.4f}]",
          flush=True)

    print(f"\n=== COMPARISON SUMMARY ===", flush=True)
    print(f"  v202 logistic clinical+V_kernel (MU): AUC = 0.728",
          flush=True)
    print(f"  v207 5-seed CNN mask+kernel mean (MU): AUC = 0.586",
          flush=True)
    print(f"  v209 deep ensemble pooled OOF (MU): AUC = 0.587",
          flush=True)
    print(f"  v213 supervised pretrain+FT on RHUH: AUC = 0.804",
          flush=True)
    print(f"  v215 SimCLR pretrain (label-free) + head FT on MU: "
          f"AUC = {pooled_simclr_auc:.4f}", flush=True)

    out = {
        "version": "v215",
        "experiment": ("SimCLR self-supervised pretraining on "
                       "multi-cohort masks + binary 365-d PFS "
                       "head fine-tune on MU"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "pretrain_epochs": PRETRAIN_EPOCHS,
        "ft_epochs": FT_EPOCHS,
        "temperature": TEMPERATURE,
        "n_pretrain_samples": n_pre,
        "cohort_counts": cohort_counts,
        "pretrain_loss_curve": [float(x) for x in losses],
        "pretrain_final_loss": float(losses[-1]),
        "n_mu_labelled": len(mu_data),
        "n_folds": N_FOLDS,
        "fold_aucs": fold_aucs,
        "fold_mean_auc": float(np.mean(fold_aucs)),
        "fold_std_auc": float(np.std(fold_aucs)),
        "pooled_oof_auc": float(pooled_simclr_auc),
        "bootstrap": {
            "mean": float(boot_aucs.mean()),
            "95_CI": [float(np.percentile(boot_aucs, 2.5)),
                       float(np.percentile(boot_aucs, 97.5))],
        },
        "comparison": {
            "v202_logistic_clin_plus_Vk_MU": 0.728,
            "v207_5seed_cnn_mean_MU": 0.586,
            "v209_deep_ensemble_MU": 0.587,
            "v213_supervised_pretrain_FT_RHUH": 0.804,
            "v215_simclr_pretrain_FT_MU": float(
                pooled_simclr_auc),
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
