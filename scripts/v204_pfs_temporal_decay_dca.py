"""v204: PFS-binary-screening temporal decay curve + bootstrap CIs +
decision-curve analysis + calibration — round 40 (CPU).

Motivated by round 39 v202 finding: V_kernel adds +10.8 pp AUC over
clinical features (age + IDH + MGMT) for binary 365-day PFS prediction
(MU-Glioma-Post n=130). Round 39 only tested 3 fixed horizons. To
characterize the kernel's clinical-utility window beyond Nature MI:

  1. Sweep H in {90, 180, 270, 365, 450, 540, 730} days
  2. Bootstrap 1000 resamples per horizon for 95% CIs on Delta AUC
  3. Decision-curve analysis at H=365: net-benefit vs threshold prob
  4. Calibration analysis at H=365: 10-bin Hosmer-Lemeshow

If Delta AUC peaks at 365 days with bootstrap-significant CIs and DCA
shows positive net-benefit across clinically meaningful thresholds,
this is the clinical-utility-quantification flagship finding.

Outputs:
  Nature_project/05_results/v204_pfs_temporal_decay_dca.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_XLSX = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v204_pfs_temporal_decay_dca.json"

SIGMA_KERNEL = 3.0
HORIZONS = [90, 180, 270, 365, 450, 540, 730]
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float32)
    y = np.array(labels, dtype=np.float32)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        nll = -np.sum(y * eta - log_one_plus) + l2 * np.sum(beta * beta)
        return nll

    beta0 = np.zeros(p)
    result = minimize(neg_log_lik, beta0, method="L-BFGS-B")
    return result.x


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X(rows, feats):
    arr = np.array([[r[f] for f in feats] for r in rows], dtype=float)
    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0)
    stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return np.column_stack([np.ones(len(arr_z)), arr_z])


def label_binary(rows, X):
    out = []
    for r in rows:
        pfs = r["pfs_days"]
        prog = r["progress"]
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def fit_eval_pair(rows_lab, feats_clin, feats_clin_vk):
    """Fit clinical-only and clinical+V_kernel; return (auc_clin,
    auc_full, predicted_prob_full, y, scores_full)."""
    if len(rows_lab) < 20:
        return None
    cc = [l for l in rows_lab if all(
        l[f] is not None and not (isinstance(l[f], float)
                                  and np.isnan(l[f]))
        for f in feats_clin_vk)]
    if len(cc) < 20:
        return None
    y = np.array([l["y"] for l in cc], dtype=float)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos < 5 or n_neg < 5:
        return None
    X_clin = build_X(cc, feats_clin)
    X_full = build_X(cc, feats_clin_vk)
    beta_clin = logistic_fit(X_clin, y)
    beta_full = logistic_fit(X_full, y)
    eta_clin = X_clin @ beta_clin
    eta_full = X_full @ beta_full
    return {
        "n": len(cc),
        "n_pos": n_pos,
        "n_neg": n_neg,
        "auc_clin": auroc(eta_clin, y),
        "auc_full": auroc(eta_full, y),
        "p_clin": sigmoid(eta_clin),
        "p_full": sigmoid(eta_full),
        "y": y,
    }


def main():
    print("=" * 78, flush=True)
    print("v204 PFS TEMPORAL-DECAY + BOOTSTRAP-CI + DCA + CALIBRATION "
          "(round 40 CPU)", flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # ---- Load clinical ----
    wb = openpyxl.load_workbook(CLINICAL_XLSX, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress, "pfs_days": pfs_d,
            "idh1": idh1, "mgmt": mgmt,
        }
    print(f"  Loaded clinical for {len(clinical)} MU patients", flush=True)

    # ---- Compute kernel features per patient ----
    print("\nComputing kernel features per patient...", flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        v_k = kernel_outgrowth_volume(m, SIGMA_KERNEL)
        baseline_volume = int(m.sum())
        c = clinical[pid]
        rows.append({
            "pid": pid,
            "v_kernel_s3": v_k,
            "baseline_volume": baseline_volume,
            "age": c["age"],
            "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"],
            "progress": c["progress"],
        })
    print(f"  {len(rows)} MU patients with mask + clinical", flush=True)
    base_filter = [r for r in rows
                   if r["pfs_days"] is not None
                   and r["progress"] is not None
                   and r["age"] is not None
                   and r["idh1"] is not None
                   and r["mgmt"] is not None]
    print(f"  {len(base_filter)} with complete clinical + PFS",
          flush=True)

    feats_clin = ["age", "idh1", "mgmt"]
    feats_clin_vk = feats_clin + ["v_kernel_s3"]

    # ---- Per-horizon: point estimate + bootstrap CI on Delta AUC ----
    horizon_results = {}
    for X in HORIZONS:
        print(f"\n=== Horizon: progressed within {X} days ===",
              flush=True)
        labelled = label_binary(base_filter, X)
        n_pos = sum(1 for l in labelled if l["y"] == 1)
        n_neg = len(labelled) - n_pos
        print(f"  n_labelled = {len(labelled)} ({n_pos} pos, {n_neg} neg)",
              flush=True)
        if n_pos < 5 or n_neg < 5:
            print("  Insufficient class balance, skipping", flush=True)
            continue
        # Point estimate (full sample)
        full = fit_eval_pair(labelled, feats_clin, feats_clin_vk)
        if full is None:
            continue
        auc_clin = full["auc_clin"]
        auc_full = full["auc_full"]
        delta = auc_full - auc_clin
        print(f"  Point: clin={auc_clin:.4f}  clin+Vk={auc_full:.4f}  "
              f"Delta={delta:+.4f}", flush=True)

        # Bootstrap on Delta AUC
        deltas = []
        clin_aucs = []
        full_aucs = []
        n = len(labelled)
        for b in range(N_BOOTSTRAPS):
            idx = rng.integers(0, n, size=n)
            boot_rows = [labelled[i] for i in idx]
            res = fit_eval_pair(boot_rows, feats_clin, feats_clin_vk)
            if res is None:
                continue
            deltas.append(res["auc_full"] - res["auc_clin"])
            clin_aucs.append(res["auc_clin"])
            full_aucs.append(res["auc_full"])
        deltas = np.array(deltas)
        clin_aucs = np.array(clin_aucs)
        full_aucs = np.array(full_aucs)
        ci_low = float(np.percentile(deltas, 2.5))
        ci_high = float(np.percentile(deltas, 97.5))
        delta_mean = float(deltas.mean())
        delta_med = float(np.median(deltas))
        # one-sided p_value: P(delta <= 0)
        p_one_sided = float((deltas <= 0).mean())
        print(f"  Bootstrap (n={len(deltas)}): "
              f"Delta_mean={delta_mean:+.4f}, Delta_median={delta_med:+.4f}, "
              f"95% CI [{ci_low:+.4f}, {ci_high:+.4f}], "
              f"P(Delta<=0)={p_one_sided:.4f}", flush=True)
        horizon_results[X] = {
            "n_labelled": len(labelled),
            "n_pos": n_pos,
            "n_neg": n_neg,
            "n_complete": full["n"],
            "auc_clin_point": float(auc_clin),
            "auc_full_point": float(auc_full),
            "delta_AUC_point": float(delta),
            "delta_AUC_bootstrap_mean": delta_mean,
            "delta_AUC_bootstrap_median": delta_med,
            "delta_AUC_95_CI": [ci_low, ci_high],
            "p_one_sided_delta_le_0": p_one_sided,
            "auc_clin_bootstrap_mean": float(clin_aucs.mean()),
            "auc_full_bootstrap_mean": float(full_aucs.mean()),
            "n_bootstrap_valid": len(deltas),
        }

    # ---- Decision-curve analysis at H=365 ----
    print("\n=== Decision-curve analysis at H=365 days ===",
          flush=True)
    labelled_365 = label_binary(base_filter, 365)
    res_365 = fit_eval_pair(labelled_365, feats_clin, feats_clin_vk)
    p_full = res_365["p_full"]
    p_clin = res_365["p_clin"]
    y = res_365["y"]
    n = len(y)
    prevalence = float(y.mean())
    print(f"  n={n} patients, prevalence={prevalence:.3f}", flush=True)

    thresholds = np.arange(0.05, 0.96, 0.05)
    dca_clin = []
    dca_full = []
    dca_treat_all = []
    for pt in thresholds:
        # Net benefit: NB = TP/N - FP/N * (pt / (1 - pt))
        tp_full = int(((p_full >= pt) & (y == 1)).sum())
        fp_full = int(((p_full >= pt) & (y == 0)).sum())
        nb_full = tp_full / n - fp_full / n * (pt / (1 - pt))
        tp_clin = int(((p_clin >= pt) & (y == 1)).sum())
        fp_clin = int(((p_clin >= pt) & (y == 0)).sum())
        nb_clin = tp_clin / n - fp_clin / n * (pt / (1 - pt))
        nb_all = prevalence - (1 - prevalence) * (pt / (1 - pt))
        dca_clin.append(nb_clin)
        dca_full.append(nb_full)
        dca_treat_all.append(nb_all)
        if pt in [0.1, 0.25, 0.5, 0.75]:
            print(f"  pt={pt:.2f}: NB_clin={nb_clin:+.4f}, "
                  f"NB_full={nb_full:+.4f}, "
                  f"Delta_NB={nb_full - nb_clin:+.4f}", flush=True)
    dca_clin = np.array(dca_clin)
    dca_full = np.array(dca_full)
    dca_diff = dca_full - dca_clin
    print(f"  Mean Delta NB across thresholds: {dca_diff.mean():+.5f}",
          flush=True)
    print(f"  N thresholds where full > clinical: "
          f"{int((dca_diff > 0).sum())} / {len(dca_diff)}",
          flush=True)

    # ---- Calibration analysis at H=365 ----
    print("\n=== Calibration analysis at H=365 days ===", flush=True)
    n_bins = 10
    order = np.argsort(p_full)
    p_sorted = p_full[order]
    y_sorted = y[order]
    bin_size = max(1, n // n_bins)
    cal_bins = []
    for b in range(n_bins):
        lo = b * bin_size
        hi = (b + 1) * bin_size if b < n_bins - 1 else n
        if hi - lo < 1:
            continue
        p_bin = p_sorted[lo:hi]
        y_bin = y_sorted[lo:hi]
        cal_bins.append({
            "bin": b + 1,
            "n": int(len(p_bin)),
            "mean_predicted": float(p_bin.mean()),
            "observed_pos_rate": float(y_bin.mean()),
        })
    # Hosmer-Lemeshow chi-square
    hl_chi2 = 0.0
    for cb in cal_bins:
        obs_pos = cb["observed_pos_rate"] * cb["n"]
        exp_pos = cb["mean_predicted"] * cb["n"]
        if exp_pos > 0 and (cb["n"] - exp_pos) > 0:
            hl_chi2 += ((obs_pos - exp_pos) ** 2 / exp_pos
                        + ((cb["n"] - obs_pos) - (cb["n"] - exp_pos)) ** 2
                        / (cb["n"] - exp_pos))
    print(f"  Hosmer-Lemeshow chi-square (df={len(cal_bins) - 2}): "
          f"{hl_chi2:.3f}", flush=True)
    for cb in cal_bins:
        print(f"    Bin {cb['bin']}: n={cb['n']}, "
              f"mean_pred={cb['mean_predicted']:.3f}, "
              f"obs_rate={cb['observed_pos_rate']:.3f}",
              flush=True)

    out = {
        "version": "v204",
        "experiment": ("PFS-binary-screening temporal decay + bootstrap "
                       "CIs + DCA + calibration"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "n_bootstrap": N_BOOTSTRAPS,
        "horizons_days": HORIZONS,
        "horizon_results": {str(k): v for k, v in horizon_results.items()},
        "dca_at_365": {
            "thresholds": thresholds.tolist(),
            "nb_clinical": dca_clin.tolist(),
            "nb_clin_plus_Vkernel": dca_full.tolist(),
            "nb_treat_all": dca_treat_all,
            "mean_delta_nb": float(dca_diff.mean()),
            "n_thresholds_full_better": int((dca_diff > 0).sum()),
            "prevalence_365": prevalence,
            "n_patients_365": n,
        },
        "calibration_at_365": {
            "hl_chi2": float(hl_chi2),
            "df": len(cal_bins) - 2,
            "bins": cal_bins,
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
