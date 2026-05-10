"""v222: IPCW time-dependent AUC + multi-horizon DCA with multi-σ
V_kernel — round 49 (CPU).

Round 40 v204 multi-horizon AUC (single-σ) excluded censored
patients before each horizon — losing information. Beyond-NMI
correction: use Inverse Probability of Censoring Weighting (IPCW)
to properly include censored patients via Kaplan-Meier-based
censoring weights.

Method: at each horizon H, IPCW weights = 1/G(min(T_i, H)) where
G is the Kaplan-Meier estimator of the censoring distribution.
Compute IPCW-weighted AUC for clinical-only vs clin+V_k(σ=3) vs
clin+V_k multi-σ. Bootstrap CIs.

Outputs:
  Nature_project/05_results/v222_ipcw_multi_horizon.json
"""
from __future__ import annotations

import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v222_ipcw_multi_horizon.json"

HORIZONS = [180, 270, 365, 450, 540, 730]
N_BOOTSTRAPS = 500
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def km_censoring(time, event):
    """Kaplan-Meier estimator of censoring distribution G(t) =
    P(C > t). Returns sorted unique times and step values."""
    time = np.asarray(time)
    event = np.asarray(event)
    # Censoring indicator (1 if censored)
    cens = 1 - event
    # KM for censoring: at each unique time, # at risk = sum(T >= t),
    # # censored = sum(T==t & cens==1)
    sort_idx = np.argsort(time)
    t_sorted = time[sort_idx]
    c_sorted = cens[sort_idx]
    n = len(time)
    G = 1.0
    grid_t = []
    grid_G = []
    grid_t.append(0.0)
    grid_G.append(1.0)
    i = 0
    while i < n:
        j = i
        # find all events at same time t_sorted[i]
        while j < n and t_sorted[j] == t_sorted[i]:
            j += 1
        n_risk = n - i
        d = int(c_sorted[i:j].sum())
        if d > 0 and n_risk > 0:
            G *= (1 - d / n_risk)
        grid_t.append(float(t_sorted[i]))
        grid_G.append(float(G))
        i = j
    return np.array(grid_t), np.array(grid_G)


def G_at(t, grid_t, grid_G):
    """Step KM evaluation; returns G(t)."""
    idx = np.searchsorted(grid_t, t, side="right") - 1
    idx = np.clip(idx, 0, len(grid_G) - 1)
    return grid_G[idx]


def ipcw_auc(scores, time, event, horizon, grid_t, grid_G):
    """IPCW-weighted AUC at horizon H, comparing 'cases' (T<=H,
    event=1) to 'controls' (T>H), with case weights 1/G(T_i^-)
    and control weights 1/G(H)."""
    time = np.asarray(time, dtype=float)
    event = np.asarray(event, dtype=int)
    scores = np.asarray(scores, dtype=float)
    cases = (time <= horizon) & (event == 1)
    controls = time > horizon
    n_c = int(cases.sum())
    n_co = int(controls.sum())
    if n_c == 0 or n_co == 0:
        return float("nan"), n_c, n_co
    # Weights: cases: 1/G(T_i^-) approximated as 1/G(T_i - eps)
    # Controls: 1/G(H)
    w_cases = np.array([1.0 / max(1e-3, G_at(t, grid_t, grid_G))
                          for t in time[cases]])
    w_co = 1.0 / max(1e-3, G_at(horizon, grid_t, grid_G))
    s_cases = scores[cases]
    s_co = scores[controls]
    # IPCW concordance: sum_i sum_j w_i * w_j * I(s_cases_i > s_co_j)
    num = 0.0
    den = 0.0
    for i in range(n_c):
        wi = w_cases[i]
        for j in range(n_co):
            num += wi * w_co * ((s_cases[i] > s_co[j]) +
                                  0.5 * (s_cases[i] == s_co[j]))
            den += wi * w_co
    if den == 0:
        return float("nan"), n_c, n_co
    return float(num / den), n_c, n_co


def auroc_simple(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X(rows, feats):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0)
    stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return np.column_stack([np.ones(len(arr_z)), arr_z])


def label_binary(rows, X):
    """Round 39-48 binary labelling: progressed within X days,
    excluding censored before X."""
    out = []
    for r in rows:
        pfs = r["pfs_days"]
        prog = r["progress"]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def main():
    print("=" * 78, flush=True)
    print("v222 IPCW MULTI-HORIZON MULTI-σ (round 49 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # Load MU
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }

    # Compute features
    print("\nLoading masks + computing multi-σ kernels...",
          flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None or c["pfs_days"] is None
                or c["progress"] is None):
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        rows.append({
            "pid": pid,
            "v_kernel_s2": float(kernel_outgrowth_volume(m, 2.0)),
            "v_kernel_s3": float(kernel_outgrowth_volume(m, 3.0)),
            "v_kernel_s4": float(kernel_outgrowth_volume(m, 4.0)),
            "v_kernel_s5": float(kernel_outgrowth_volume(m, 5.0)),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"],
            "progress": c["progress"],
        })
    n = len(rows)
    print(f"  {n} MU patients", flush=True)

    feats_clin = ["age", "idh1", "mgmt"]
    feats_vk_s3 = feats_clin + ["v_kernel_s3"]
    feats_vk_multi = feats_clin + [
        "v_kernel_s2", "v_kernel_s3", "v_kernel_s4",
        "v_kernel_s5"]

    # KM censoring distribution
    time_arr = np.array([r["pfs_days"] for r in rows],
                         dtype=float)
    event_arr = np.array([r["progress"] for r in rows],
                          dtype=float)
    n_events = int(event_arr.sum())
    n_cens = n - n_events
    print(f"  Total events={n_events}, censored={n_cens}",
          flush=True)
    grid_t, grid_G = km_censoring(time_arr, event_arr)
    print(f"  KM censoring grid: {len(grid_t)} unique times, "
          f"final G={grid_G[-1]:.4f}", flush=True)

    # ---- IPCW AUC at each horizon ----
    print(f"\n=== IPCW multi-horizon AUC ===", flush=True)
    horizon_results = {}
    for H in HORIZONS:
        print(f"\n  Horizon H={H} d:", flush=True)
        # Cases & controls
        cases_mask = (time_arr <= H) & (event_arr == 1)
        controls_mask = time_arr > H
        n_c = int(cases_mask.sum())
        n_co = int(controls_mask.sum())
        n_excluded = n - n_c - n_co  # censored before H
        print(f"    cases={n_c}, controls={n_co}, "
              f"excluded (censored before H)={n_excluded}",
              flush=True)
        if n_c < 5 or n_co < 5:
            print(f"    insufficient class balance, skipping",
                  flush=True)
            continue

        # Build feature matrices on full sample (no exclusion)
        X_clin = build_X(rows, feats_clin)
        X_vks3 = build_X(rows, feats_vk_s3)
        X_multi = build_X(rows, feats_vk_multi)

        # For training, use binary label exclusion as round 39
        # (avoid leakage from censored)
        bin_lab = label_binary(rows, H)
        if len(bin_lab) < 20:
            continue
        y_bin = np.array([l["y"] for l in bin_lab], dtype=float)
        # Build training X for those rows
        idx_lab = [i for i, r in enumerate(rows)
                    if any(l["pid"] == r["pid"]
                           for l in bin_lab)]
        # Cleaner: refit logistic on labelled subset
        rows_lab = bin_lab
        X_clin_l = build_X(rows_lab, feats_clin)
        X_vks3_l = build_X(rows_lab, feats_vk_s3)
        X_multi_l = build_X(rows_lab, feats_vk_multi)
        y_l = np.array([r["y"] for r in rows_lab], dtype=float)
        b_clin = logistic_fit(X_clin_l, y_l)
        b_vks3 = logistic_fit(X_vks3_l, y_l)
        b_multi = logistic_fit(X_multi_l, y_l)

        # Score ALL patients (including censored)
        s_clin = X_clin @ b_clin
        s_vks3 = X_vks3 @ b_vks3
        s_multi = X_multi @ b_multi

        # IPCW AUC
        auc_clin_ipcw, _, _ = ipcw_auc(s_clin, time_arr,
                                          event_arr, H,
                                          grid_t, grid_G)
        auc_vks3_ipcw, _, _ = ipcw_auc(s_vks3, time_arr,
                                          event_arr, H,
                                          grid_t, grid_G)
        auc_multi_ipcw, _, _ = ipcw_auc(s_multi, time_arr,
                                           event_arr, H,
                                           grid_t, grid_G)
        # Naive AUC (excluding censored)
        rows_lab_arr = np.array(rows_lab)
        s_clin_naive = build_X(rows_lab, feats_clin) @ b_clin
        s_vks3_naive = build_X(rows_lab, feats_vk_s3) @ b_vks3
        s_multi_naive = (build_X(rows_lab, feats_vk_multi)
                          @ b_multi)
        auc_clin_naive = auroc_simple(s_clin_naive, y_l)
        auc_vks3_naive = auroc_simple(s_vks3_naive, y_l)
        auc_multi_naive = auroc_simple(s_multi_naive, y_l)

        delta_vks3_ipcw = auc_vks3_ipcw - auc_clin_ipcw
        delta_multi_ipcw = auc_multi_ipcw - auc_clin_ipcw
        delta_multi_naive = auc_multi_naive - auc_clin_naive
        print(f"    NAIVE (excl censored): clin={auc_clin_naive:.4f}, "
              f"V_k σ=3={auc_vks3_naive:.4f}, multi-σ="
              f"{auc_multi_naive:.4f}, "
              f"Δ multi-σ={delta_multi_naive:+.4f}",
              flush=True)
        print(f"    IPCW (uses censored): clin={auc_clin_ipcw:.4f}, "
              f"V_k σ=3={auc_vks3_ipcw:.4f}, multi-σ="
              f"{auc_multi_ipcw:.4f}, "
              f"Δ multi-σ={delta_multi_ipcw:+.4f}",
              flush=True)

        # Bootstrap on IPCW AUC for multi-σ
        bs_deltas = []
        n_full = n
        for _ in range(N_BOOTSTRAPS):
            idx = rng.integers(0, n_full, size=n_full)
            t_b = time_arr[idx]
            e_b = event_arr[idx]
            sc_b = s_clin[idx]
            sm_b = s_multi[idx]
            try:
                gt_b, gG_b = km_censoring(t_b, e_b)
                a_c, _, _ = ipcw_auc(sc_b, t_b, e_b, H,
                                        gt_b, gG_b)
                a_m, _, _ = ipcw_auc(sm_b, t_b, e_b, H,
                                        gt_b, gG_b)
                if not (np.isnan(a_c) or np.isnan(a_m)):
                    bs_deltas.append(a_m - a_c)
            except Exception:
                continue
        bs_deltas = np.array(bs_deltas)
        if len(bs_deltas) > 0:
            ci_lo = float(np.percentile(bs_deltas, 2.5))
            ci_hi = float(np.percentile(bs_deltas, 97.5))
            p_one = float((bs_deltas <= 0).mean())
            print(f"    Bootstrap IPCW Δ multi-σ "
                  f"(n={len(bs_deltas)}): mean="
                  f"{bs_deltas.mean():+.4f}, 95% CI ["
                  f"{ci_lo:+.4f}, {ci_hi:+.4f}], "
                  f"P<=0={p_one:.4f}", flush=True)
        else:
            ci_lo = ci_hi = p_one = float("nan")

        horizon_results[H] = {
            "n_cases": n_c,
            "n_controls": n_co,
            "n_excluded_naive": n_excluded,
            "naive": {
                "auc_clin": float(auc_clin_naive),
                "auc_Vk_s3": float(auc_vks3_naive),
                "auc_multi_sigma": float(auc_multi_naive),
                "delta_multi_sigma": float(delta_multi_naive),
            },
            "ipcw": {
                "auc_clin": float(auc_clin_ipcw),
                "auc_Vk_s3": float(auc_vks3_ipcw),
                "auc_multi_sigma": float(auc_multi_ipcw),
                "delta_Vk_s3": float(delta_vks3_ipcw),
                "delta_multi_sigma": float(delta_multi_ipcw),
            },
            "bootstrap_ipcw_multi_sigma": {
                "n_valid": len(bs_deltas),
                "mean": (float(bs_deltas.mean())
                          if len(bs_deltas) > 0
                          else float("nan")),
                "95_CI": [ci_lo, ci_hi],
                "p_one_sided": p_one,
            },
        }

    out = {
        "version": "v222",
        "experiment": ("IPCW time-dependent AUC at multi-"
                       "horizons with multi-σ V_kernel; "
                       "compares to naive (excl-censored) AUC"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "n_total": n,
        "n_events": n_events,
        "n_censored": n_cens,
        "horizons": HORIZONS,
        "horizon_results": {str(k): v for k, v in
                             horizon_results.items()},
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
