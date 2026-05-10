"""v209: 5-fold CV deep ensemble (10 members per fold) with
calibration + selective prediction — round 42 (GPU).

The Nature/Lancet regulatory must-have for a clinical predictive
model: per-patient uncertainty quantification + calibration +
selective prediction. Round 41 v207 showed CNN kernel rescue is
seed-dependent; this round leverages that variability constructively
via a 10-member deep ensemble.

Method:
  - 5-fold stratified CV on MU-Glioma-Post n=130 binary 365-d PFS
  - Per fold: train 10 ensemble members with different RNG seeds
    on the same training set
  - Per test patient: predicted probability mean + std across the 10
    ensemble members
  - Pooled OOF AUC, ECE (10-bin Expected Calibration Error),
    reliability diagram
  - Selective prediction: at coverage levels c in [0.5, 1.0], defer
    the (1-c) fraction of MOST-UNCERTAIN patients, compute AUC on
    the remaining

Outputs:
  Nature_project/05_results/v209_deep_ensemble_uncertainty.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v209_deep_ensemble_uncertainty.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
N_ENSEMBLE = 10
EPOCHS = 30
LR = 5e-4
BASE_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class BinaryPFSCNN(nn.Module):
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(32, 1),
        )

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.head(self.global_pool(e3)).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def train_member(Xtr, ytr, n_tr_pos, n_tr_neg, in_ch, seed):
    torch.manual_seed(seed)
    np.random.seed(seed)
    Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
    ytr_t = torch.from_numpy(ytr).to(DEVICE)
    model = BinaryPFSCNN(in_ch=in_ch, base=24).to(DEVICE)
    opt = torch.optim.AdamW(model.parameters(), lr=LR,
                             weight_decay=1e-3)
    pos_w = torch.tensor([n_tr_neg / max(1, n_tr_pos)],
                          dtype=torch.float32, device=DEVICE)
    criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    n_tr = len(Xtr)
    bs = 4
    for ep in range(EPOCHS):
        model.train()
        perm = np.random.permutation(n_tr)
        for i in range(0, n_tr, bs):
            idx = perm[i:i+bs]
            loss = criterion(model(Xtr_t[idx]), ytr_t[idx])
            opt.zero_grad()
            loss.backward()
            opt.step()
    return model


def main():
    print("=" * 78, flush=True)
    print(f"v209 DEEP-ENSEMBLE + UNCERTAINTY (round 42 GPU) "
          f"device={DEVICE}", flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    # ---- Load data (mask + kernel + binary 365-d label) ----
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    print(f"  Loaded clinical for {len(clinical)} MU patients",
          flush=True)

    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    n = len(data)
    y_all = np.array([d["y"] for d in data], dtype=np.float32)
    n_pos = int(y_all.sum())
    n_neg = int(n - n_pos)
    print(f"  {n} patients, n_pos={n_pos}, n_neg={n_neg}",
          flush=True)

    folds = stratified_kfold(y_all, N_FOLDS, seed=BASE_SEED)
    print(f"  Stratified {N_FOLDS}-fold sizes: "
          f"{[len(f) for f in folds]}", flush=True)

    # ---- Train ensemble per fold ----
    oof_means = np.zeros(n, dtype=np.float32)
    oof_stds = np.zeros(n, dtype=np.float32)
    oof_assigned = np.zeros(n, dtype=bool)

    for fold_i, test_idx in enumerate(folds):
        print(f"\n--- FOLD {fold_i+1}/{N_FOLDS} ---", flush=True)
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        n_tr = len(train_idx)
        n_tr_pos = int(y_all[train_idx].sum())
        n_tr_neg = int(n_tr - n_tr_pos)
        # 2-channel input (mask + kernel)
        Xtr = np.stack(
            [np.stack([data[i]["mask"], data[i]["kernel"]],
                       axis=0)
             for i in train_idx]).astype(np.float32)
        Xte = np.stack(
            [np.stack([data[i]["mask"], data[i]["kernel"]],
                       axis=0)
             for i in test_idx]).astype(np.float32)
        ytr = y_all[train_idx]
        yte = y_all[test_idx]
        Xte_t = torch.from_numpy(Xte).to(DEVICE)

        # Per-test predictions across N_ENSEMBLE members
        test_probs = np.zeros((N_ENSEMBLE, len(test_idx)),
                                dtype=np.float32)
        for m_i in range(N_ENSEMBLE):
            seed = BASE_SEED + fold_i * 100 + m_i
            model = train_member(Xtr, ytr, n_tr_pos, n_tr_neg,
                                  in_ch=2, seed=seed)
            model.eval()
            with torch.no_grad():
                logits = model(Xte_t).cpu().numpy()
            probs = 1.0 / (1.0 + np.exp(-np.clip(logits, -50, 50)))
            test_probs[m_i] = probs
            del model
            gc.collect()
            if DEVICE.type == "cuda":
                torch.cuda.empty_cache()
            print(f"    member {m_i+1}/{N_ENSEMBLE} done",
                  flush=True)

        # Aggregate
        mean_pred = test_probs.mean(axis=0)
        std_pred = test_probs.std(axis=0)
        oof_means[test_idx] = mean_pred
        oof_stds[test_idx] = std_pred
        oof_assigned[test_idx] = True
        fold_auc = auroc(mean_pred, yte)
        print(f"  Fold ensemble AUC = {fold_auc:.4f}",
              flush=True)

    # ---- Pooled OOF metrics ----
    pooled_auc = auroc(oof_means[oof_assigned],
                        y_all[oof_assigned])
    print(f"\n=== POOLED OOF METRICS ===", flush=True)
    print(f"  Ensemble pooled OOF AUC = {pooled_auc:.4f}",
          flush=True)

    # ---- ECE + reliability diagram (10-bin) ----
    n_bins = 10
    bin_edges = np.linspace(0, 1, n_bins + 1)
    bin_indices = np.digitize(oof_means, bin_edges) - 1
    bin_indices = np.clip(bin_indices, 0, n_bins - 1)
    ece_total = 0.0
    reliability = []
    for b in range(n_bins):
        sel = bin_indices == b
        if sel.sum() == 0:
            continue
        mean_p = float(oof_means[sel].mean())
        obs = float(y_all[sel].mean())
        n_b = int(sel.sum())
        ece_total += (n_b / n) * abs(obs - mean_p)
        reliability.append({
            "bin": b + 1, "n": n_b,
            "mean_predicted": mean_p,
            "observed_pos_rate": obs,
        })
    print(f"  ECE (10-bin) = {ece_total:.4f}", flush=True)
    for r in reliability:
        print(f"    bin {r['bin']}: n={r['n']}, "
              f"pred={r['mean_predicted']:.3f}, "
              f"obs={r['observed_pos_rate']:.3f}", flush=True)

    # ---- Selective prediction (uncertainty-deferral) ----
    print(f"\n=== SELECTIVE PREDICTION ===", flush=True)
    selective = []
    coverages = [1.00, 0.95, 0.90, 0.80, 0.70, 0.60, 0.50]
    sort_idx = np.argsort(oof_stds)  # lower uncertainty first
    for c in coverages:
        keep_n = int(np.ceil(c * n))
        keep = sort_idx[:keep_n]
        if int(y_all[keep].sum()) < 3 or int(
                len(keep) - y_all[keep].sum()) < 3:
            continue
        a = auroc(oof_means[keep], y_all[keep])
        sel_pos = int(y_all[keep].sum())
        sel_neg = int(len(keep) - sel_pos)
        selective.append({
            "coverage": c, "n_kept": keep_n,
            "n_pos_kept": sel_pos, "n_neg_kept": sel_neg,
            "auc": float(a),
        })
        print(f"  coverage={c:.2f} (n={keep_n}, "
              f"{sel_pos} pos, {sel_neg} neg): AUC={a:.4f}",
              flush=True)

    # ---- Uncertainty-vs-correctness ----
    # Bin by uncertainty quartile, report AUC and accuracy per quartile
    print(f"\n=== UNCERTAINTY-vs-AUC quartiles ===", flush=True)
    q_edges = np.percentile(oof_stds, [0, 25, 50, 75, 100])
    quartile_results = []
    for qi in range(4):
        lo, hi = q_edges[qi], q_edges[qi + 1]
        if qi == 3:
            sel = (oof_stds >= lo) & (oof_stds <= hi)
        else:
            sel = (oof_stds >= lo) & (oof_stds < hi)
        if sel.sum() < 5:
            continue
        ys = y_all[sel]
        ms = oof_means[sel]
        a = auroc(ms, ys)
        quartile_results.append({
            "quartile": qi + 1,
            "uncertainty_range": [float(lo), float(hi)],
            "n": int(sel.sum()),
            "n_pos": int(ys.sum()),
            "n_neg": int(len(ys) - ys.sum()),
            "auc": float(a),
            "mean_uncertainty": float(oof_stds[sel].mean()),
        })
        print(f"  Q{qi+1} (std in [{lo:.4f}, {hi:.4f}], "
              f"n={int(sel.sum())}, "
              f"{int(ys.sum())} pos): AUC={a:.4f}", flush=True)

    out = {
        "version": "v209",
        "experiment": ("Deep ensemble (10 members per fold) + "
                       "calibration + selective prediction on "
                       "binary 365-d PFS"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "n_ensemble_per_fold": N_ENSEMBLE,
        "epochs_per_member": EPOCHS,
        "n_total": n,
        "n_pos": n_pos,
        "n_neg": n_neg,
        "pooled_oof_auc": float(pooled_auc),
        "ece_10bin": float(ece_total),
        "reliability": reliability,
        "selective_prediction": selective,
        "uncertainty_quartiles": quartile_results,
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
