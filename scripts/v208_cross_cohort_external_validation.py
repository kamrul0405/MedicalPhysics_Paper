"""v208: Cross-cohort external validation — train on MU-Glioma-Post,
test on RHUH-GBM — round 42 (CPU).

The single biggest Nature/Lancet vulnerability after round 41: every
empirical-grounding piece (L1-L7) was on the same MU-Glioma-Post
cohort (n=130). The flagship move is to **train the multivariate
logistic on MU, then evaluate on a fully held-out external cohort
(RHUH-GBM)** with the same binary 365-day PFS task.

RHUH-GBM has IDH + Age + PFS days + censor flag (no MGMT). Therefore
the cross-cohort model uses a 3-feature multivariate logistic:
age + IDH + V_kernel(sigma=3). Both cohorts: binary "progressed
within 365 days" outcome.

Method:
  1. Load both cohorts; build identical features
  2. For both cohorts: fit logistic clinical-only (age + IDH) and
     clinical + V_kernel (age + IDH + V_kernel) using MU as the
     training set
  3. Apply the MU-trained model to RHUH; compute cross-cohort AUC
     for both variants; compute Delta cross-cohort AUC
  4. Bootstrap (1000 resamples on the test cohort RHUH) to get
     95% CI on cross-cohort Delta AUC
  5. Report direction-of-effect: positive Delta_cross_cohort
     would replicate the round-39/40 finding externally

Outputs:
  Nature_project/05_results/v208_cross_cohort_external.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v208_cross_cohort_external.json"

SIGMA_KERNEL = 3.0
HORIZON = 365
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def build_X_with_stats(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def label_binary(rows, X, pfs_field, prog_field):
    out = []
    for r in rows:
        pfs = r[pfs_field]
        prog = r[prog_field]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def load_mu():
    """Load MU-Glioma-Post features and binary 365-d PFS labels."""
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress, "pfs_days": pfs_d,
            "idh1": idh1,
        }

    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid,
            "v_kernel_s3": kernel_outgrowth_volume(m, SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def load_rhuh():
    """Load RHUH-GBM features and binary 365-d PFS labels."""
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        # Find columns
        age_col = next((c for c in cols if c.strip() == "Age"),
                        None)
        idh_col = next((c for c in cols if "IDH status" in c),
                        None)
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        print(f"  RHUH columns: age='{age_col}', idh='{idh_col}'",
              flush=True)
        print(f"  RHUH columns: pfs='{pfs_col}', "
              f"cens='{cens_col}'", flush=True)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                age = (float(row[age_col]) if row[age_col]
                       else None)
                idh_str = (row[idh_col].strip().lower()
                            if row[idh_col] else "")
                if "mut" in idh_str:
                    idh1 = 1
                elif "wt" in idh_str or "wild" in idh_str:
                    idh1 = 0
                else:
                    idh1 = None
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                # Right Censored = "yes" -> NOT a progression event
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if age is None or idh1 is None or pfs_d is None:
                continue
            clinical[pid] = {
                "age": age, "idh1": idh1, "pfs_days": pfs_d,
                "progress": event,
            }
    print(f"  Loaded RHUH clinical for {len(clinical)} patients "
          f"(non-null IDH/Age/PFS/event)", flush=True)

    rows = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid,
            "v_kernel_s3": kernel_outgrowth_volume(m, SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def main():
    print("=" * 78, flush=True)
    print("v208 CROSS-COHORT EXTERNAL VALIDATION (round 42 CPU)",
          flush=True)
    print("  Train on MU-Glioma-Post, test on RHUH-GBM",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # ---- Load both cohorts ----
    mu_rows = load_mu()
    print(f"  Loaded {len(mu_rows)} MU patients with "
          f"masks + clinical", flush=True)
    rhuh_rows = load_rhuh()
    print(f"  Loaded {len(rhuh_rows)} RHUH patients with "
          f"masks + clinical", flush=True)

    mu_lab = label_binary(mu_rows, HORIZON, "pfs_days",
                            "progress")
    rhuh_lab = label_binary(rhuh_rows, HORIZON, "pfs_days",
                              "progress")
    mu_pos = sum(1 for l in mu_lab if l["y"] == 1)
    mu_neg = len(mu_lab) - mu_pos
    rhuh_pos = sum(1 for l in rhuh_lab if l["y"] == 1)
    rhuh_neg = len(rhuh_lab) - rhuh_pos
    print(f"\n  MU labelled: n={len(mu_lab)} ({mu_pos} pos, "
          f"{mu_neg} neg)", flush=True)
    print(f"  RHUH labelled: n={len(rhuh_lab)} ({rhuh_pos} pos, "
          f"{rhuh_neg} neg)", flush=True)

    if rhuh_pos < 3 or rhuh_neg < 3:
        print("  ERROR: insufficient class balance on RHUH",
              flush=True)
        return

    feats_clin = ["age", "idh1"]
    feats_full = feats_clin + ["v_kernel_s3"]

    # ---- Filter to complete cases on each cohort ----
    mu_cc = [l for l in mu_lab if all(l[f] is not None
                                       for f in feats_full)]
    rhuh_cc = [l for l in rhuh_lab if all(l[f] is not None
                                            for f in feats_full)]
    print(f"\n  MU complete-case: {len(mu_cc)}", flush=True)
    print(f"  RHUH complete-case: {len(rhuh_cc)}", flush=True)

    # ---- Train on MU, with MU-derived standardization ----
    y_mu = np.array([l["y"] for l in mu_cc], dtype=float)
    X_mu_clin, mu_means_c, mu_stds_c = build_X_with_stats(
        mu_cc, feats_clin)
    X_mu_full, mu_means_f, mu_stds_f = build_X_with_stats(
        mu_cc, feats_full)
    beta_clin = logistic_fit(X_mu_clin, y_mu)
    beta_full = logistic_fit(X_mu_full, y_mu)
    print(f"\n  MU-trained beta_clin = {beta_clin}", flush=True)
    print(f"  MU-trained beta_full = {beta_full}", flush=True)

    # In-sample MU AUC (sanity check)
    mu_auc_clin = auroc(X_mu_clin @ beta_clin, y_mu)
    mu_auc_full = auroc(X_mu_full @ beta_full, y_mu)
    print(f"\n  MU in-sample: AUC_clin={mu_auc_clin:.4f}  "
          f"AUC_full={mu_auc_full:.4f}  "
          f"Delta={mu_auc_full - mu_auc_clin:+.4f}", flush=True)

    # ---- Apply to RHUH (using MU's standardization) ----
    y_rhuh = np.array([l["y"] for l in rhuh_cc], dtype=float)
    X_rhuh_clin, _, _ = build_X_with_stats(
        rhuh_cc, feats_clin, means=mu_means_c, stds=mu_stds_c)
    X_rhuh_full, _, _ = build_X_with_stats(
        rhuh_cc, feats_full, means=mu_means_f, stds=mu_stds_f)
    rhuh_auc_clin = auroc(X_rhuh_clin @ beta_clin, y_rhuh)
    rhuh_auc_full = auroc(X_rhuh_full @ beta_full, y_rhuh)
    delta_external = rhuh_auc_full - rhuh_auc_clin
    print(f"\n  *** CROSS-COHORT EXTERNAL VALIDATION ***",
          flush=True)
    print(f"  RHUH (held-out): AUC_clin={rhuh_auc_clin:.4f}  "
          f"AUC_full={rhuh_auc_full:.4f}  "
          f"Delta_external={delta_external:+.4f}", flush=True)

    # ---- Bootstrap on RHUH (test cohort) ----
    print(f"\n  Bootstrap on RHUH n={len(rhuh_cc)} "
          f"({N_BOOTSTRAPS} resamples)...", flush=True)
    delta_boots = []
    auc_clin_boots = []
    auc_full_boots = []
    n_rhuh = len(rhuh_cc)
    for b in range(N_BOOTSTRAPS):
        idx = rng.integers(0, n_rhuh, size=n_rhuh)
        y_b = y_rhuh[idx]
        if int(y_b.sum()) < 2 or int(len(y_b) - y_b.sum()) < 2:
            continue
        Xc_b = X_rhuh_clin[idx]
        Xf_b = X_rhuh_full[idx]
        a_clin = auroc(Xc_b @ beta_clin, y_b)
        a_full = auroc(Xf_b @ beta_full, y_b)
        delta_boots.append(a_full - a_clin)
        auc_clin_boots.append(a_clin)
        auc_full_boots.append(a_full)
    delta_boots = np.array(delta_boots)
    auc_clin_boots = np.array(auc_clin_boots)
    auc_full_boots = np.array(auc_full_boots)
    ci_lo = float(np.percentile(delta_boots, 2.5))
    ci_hi = float(np.percentile(delta_boots, 97.5))
    p_one = float((delta_boots <= 0).mean())
    print(f"  Bootstrap (n_valid={len(delta_boots)}):", flush=True)
    print(f"    Delta_external mean={delta_boots.mean():+.4f}, "
          f"median={np.median(delta_boots):+.4f}", flush=True)
    print(f"    95% CI [{ci_lo:+.4f}, {ci_hi:+.4f}], "
          f"P(Delta<=0)={p_one:.4f}", flush=True)
    print(f"    AUC_clin: mean={auc_clin_boots.mean():.4f}  "
          f"95% CI [{np.percentile(auc_clin_boots, 2.5):.4f}, "
          f"{np.percentile(auc_clin_boots, 97.5):.4f}]",
          flush=True)
    print(f"    AUC_full: mean={auc_full_boots.mean():.4f}  "
          f"95% CI [{np.percentile(auc_full_boots, 2.5):.4f}, "
          f"{np.percentile(auc_full_boots, 97.5):.4f}]",
          flush=True)

    out = {
        "version": "v208",
        "experiment": ("Cross-cohort external validation: train "
                       "on MU-Glioma-Post, test on RHUH-GBM"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "sigma_kernel": SIGMA_KERNEL,
        "features_clin": feats_clin,
        "features_full": feats_full,
        "training": {
            "cohort": "MU-Glioma-Post",
            "n_complete": len(mu_cc),
            "n_pos": int(y_mu.sum()),
            "n_neg": int(len(y_mu) - y_mu.sum()),
            "auc_clin_in_sample": float(mu_auc_clin),
            "auc_full_in_sample": float(mu_auc_full),
            "delta_in_sample": float(mu_auc_full - mu_auc_clin),
            "beta_clin": beta_clin.tolist(),
            "beta_full": beta_full.tolist(),
        },
        "external_test": {
            "cohort": "RHUH-GBM",
            "n_complete": len(rhuh_cc),
            "n_pos": int(y_rhuh.sum()),
            "n_neg": int(len(y_rhuh) - y_rhuh.sum()),
            "auc_clin_external": float(rhuh_auc_clin),
            "auc_full_external": float(rhuh_auc_full),
            "delta_external_point": float(delta_external),
            "delta_external_bootstrap_mean": float(
                delta_boots.mean()),
            "delta_external_bootstrap_median": float(
                np.median(delta_boots)),
            "delta_external_95_CI": [ci_lo, ci_hi],
            "p_one_sided": p_one,
            "auc_clin_external_95_CI": [
                float(np.percentile(auc_clin_boots, 2.5)),
                float(np.percentile(auc_clin_boots, 97.5))],
            "auc_full_external_95_CI": [
                float(np.percentile(auc_full_boots, 2.5)),
                float(np.percentile(auc_full_boots, 97.5))],
            "n_bootstrap_valid": len(delta_boots),
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
