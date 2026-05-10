"""v220: Multi-σ V_kernel comprehensive validation —
cross-cohort + meta-analysis + continuous PFS Cox — round 48 (CPU).

Round 47 v218 established multi-σ V_kernel breakthrough on MU n=130
(AUC=0.815, NRI=+0.805, IDI=+0.112). Open questions:

  1. Does multi-σ V_kernel CROSS-COHORT replicate? (round 43 v210
     tested single-σ; got Δ=-0.005 on RHUH due to power)
  2. Does IV-weighted meta-analysis with multi-σ give a stronger
     P-value than round 43's 0.036?
  3. Does multi-σ improve CONTINUOUS PFS Cox over round 45 v214's
     single-σ V_kernel + clinical Cox C=0.616 (LRT P=0.007)?
  4. What is the FINAL Nature/Lancet evidence summary?

Method:
  A. Train MU multi-σ logistic (clin+V_k(σ=2,3,4,5)) on full MU,
     evaluate on RHUH external (binary 365-d PFS). Bootstrap CI.
  B. Inverse-variance meta-analysis: pool MU bootstrap + RHUH
     bootstrap distributions to get pooled Delta AUC.
  C. Cox PH on continuous MU PFS with all multi-σ V_kernel
     features added to clinical. Compare to v214 Cox.
  D. Transfer learning: apply MU-trained multi-σ logistic to
     RHUH directly (no fine-tuning).

Outputs:
  Nature_project/05_results/v220_multi_sigma_comprehensive.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize
from scipy.stats import chi2, norm

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v220_multi_sigma_comprehensive.json"

HORIZON = 365
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X_with_stats(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def label_binary(rows, X, pfs_field="pfs_days",
                  prog_field="progress"):
    out = []
    for r in rows:
        pfs = r[pfs_field]
        prog = r[prog_field]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def cox_partial_loglik(beta, X, time, event, l2=1e-3):
    eta = X @ beta
    order = np.argsort(-time)
    eta_o = eta[order]
    eta_max_o = eta_o.max()
    es_o = np.exp(eta_o - eta_max_o)
    cum_es = np.cumsum(es_o)
    log_cum = eta_max_o + np.log(cum_es)
    log_cum_orig = np.empty(len(time))
    log_cum_orig[order] = log_cum
    valid = event == 1
    log_pl = (eta[valid] - log_cum_orig[valid]).sum()
    return -log_pl + l2 * np.sum(beta * beta)


def cox_fit(X, time, event):
    p = X.shape[1]
    res = minimize(cox_partial_loglik, np.zeros(p),
                    args=(X, time, event), method="L-BFGS-B")
    return res.x, -res.fun


def harrells_c(time, event, risk):
    time = np.asarray(time)
    event = np.asarray(event)
    risk = np.asarray(risk)
    n = len(time)
    n_pairs = 0
    n_concordant = 0
    n_ties = 0
    for i in range(n):
        if event[i] != 1:
            continue
        for j in range(n):
            if i == j:
                continue
            if time[i] >= time[j]:
                continue
            n_pairs += 1
            if risk[i] > risk[j]:
                n_concordant += 1
            elif risk[i] == risk[j]:
                n_ties += 1
    if n_pairs == 0:
        return float("nan")
    return (n_concordant + 0.5 * n_ties) / n_pairs


def load_mu():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid, "cohort": "MU",
            "v_kernel_s2": float(kernel_outgrowth_volume(m, 2.0)),
            "v_kernel_s3": float(kernel_outgrowth_volume(m, 3.0)),
            "v_kernel_s4": float(kernel_outgrowth_volume(m, 4.0)),
            "v_kernel_s5": float(kernel_outgrowth_volume(m, 5.0)),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def load_rhuh():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        age_col = next((c for c in cols if c.strip() == "Age"),
                        None)
        idh_col = next((c for c in cols if "IDH status" in c),
                        None)
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                age = (float(row[age_col]) if row[age_col]
                       else None)
                idh_str = (row[idh_col].strip().lower()
                            if row[idh_col] else "")
                if "mut" in idh_str:
                    idh1 = 1
                elif "wt" in idh_str or "wild" in idh_str:
                    idh1 = 0
                else:
                    idh1 = None
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if age is None or idh1 is None or pfs_d is None:
                continue
            clinical[pid] = {
                "age": age, "idh1": idh1,
                "pfs_days": pfs_d, "progress": event,
            }
    rows = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid, "cohort": "RHUH",
            "v_kernel_s2": float(kernel_outgrowth_volume(m, 2.0)),
            "v_kernel_s3": float(kernel_outgrowth_volume(m, 3.0)),
            "v_kernel_s4": float(kernel_outgrowth_volume(m, 4.0)),
            "v_kernel_s5": float(kernel_outgrowth_volume(m, 5.0)),
            "age": c["age"], "idh1": c["idh1"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def main():
    print("=" * 78, flush=True)
    print("v220 MULTI-σ COMPREHENSIVE VALIDATION (round 48 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    mu_rows = load_mu()
    rhuh_rows = load_rhuh()
    mu_lab = label_binary(mu_rows, HORIZON)
    rhuh_lab = label_binary(rhuh_rows, HORIZON)
    print(f"  MU labelled: n={len(mu_lab)} "
          f"({sum(1 for l in mu_lab if l['y']==1)} pos)",
          flush=True)
    print(f"  RHUH labelled: n={len(rhuh_lab)} "
          f"({sum(1 for l in rhuh_lab if l['y']==1)} pos)",
          flush=True)

    # Note: RHUH lacks MGMT, so for cross-cohort we use 3-feature
    # clinical (age + IDH1) + 4 multi-σ V_kernel = 6 features
    feats_clin_3 = ["age", "idh1", "mgmt"]  # MU 4-feature
    feats_clin_2 = ["age", "idh1"]  # cross-cohort 2-feature
    feats_vk_multi = ["v_kernel_s2", "v_kernel_s3",
                       "v_kernel_s4", "v_kernel_s5"]
    feats_full_mu = feats_clin_3 + feats_vk_multi
    feats_full_xc = feats_clin_2 + feats_vk_multi

    # Filter to complete cases
    mu_cc = [l for l in mu_lab if all(l[f] is not None
                                       for f in feats_full_mu)]
    rhuh_cc = [l for l in rhuh_lab if all(l[f] is not None
                                            for f in feats_full_xc)]
    print(f"  MU complete-case (4-feat clin + multi-σ): "
          f"{len(mu_cc)}", flush=True)
    print(f"  RHUH complete-case (3-feat clin + multi-σ): "
          f"{len(rhuh_cc)}", flush=True)

    # ============================================================
    # PART A: MU in-sample multi-σ + cross-cohort RHUH external
    # (using 3-feature clinical: no MGMT)
    # ============================================================
    print(f"\n=== PART A: cross-cohort multi-σ external "
          f"validation ===", flush=True)
    feats_clin_xc = feats_clin_2  # age + idh1
    feats_full_xc = feats_clin_xc + feats_vk_multi  # 6 features

    # Train multi-σ logistic on MU
    y_mu = np.array([l["y"] for l in mu_cc], dtype=float)
    X_mu_clin, mc, sc = build_X_with_stats(mu_cc, feats_clin_xc)
    X_mu_full, mf, sf = build_X_with_stats(mu_cc, feats_full_xc)
    bc_mu = logistic_fit(X_mu_clin, y_mu)
    bf_mu = logistic_fit(X_mu_full, y_mu)
    auc_mu_clin = auroc(X_mu_clin @ bc_mu, y_mu)
    auc_mu_full = auroc(X_mu_full @ bf_mu, y_mu)
    delta_mu = auc_mu_full - auc_mu_clin
    print(f"  MU in-sample (2-feat clin + 4-feat multi-σ):",
          flush=True)
    print(f"    AUC_clin={auc_mu_clin:.4f}, "
          f"AUC_full={auc_mu_full:.4f}, "
          f"Delta={delta_mu:+.4f}", flush=True)

    # Apply to RHUH
    y_rhuh = np.array([l["y"] for l in rhuh_cc], dtype=float)
    X_rhuh_clin, _, _ = build_X_with_stats(
        rhuh_cc, feats_clin_xc, means=mc, stds=sc)
    X_rhuh_full, _, _ = build_X_with_stats(
        rhuh_cc, feats_full_xc, means=mf, stds=sf)
    auc_rh_clin = auroc(X_rhuh_clin @ bc_mu, y_rhuh)
    auc_rh_full = auroc(X_rhuh_full @ bf_mu, y_rhuh)
    delta_rh = auc_rh_full - auc_rh_clin
    print(f"  RHUH external (multi-σ MU-trained logistic):",
          flush=True)
    print(f"    AUC_clin={auc_rh_clin:.4f}, "
          f"AUC_full={auc_rh_full:.4f}, "
          f"Delta_external={delta_rh:+.4f}", flush=True)

    # Bootstrap on RHUH
    print(f"  Bootstrapping {N_BOOTSTRAPS} on RHUH...",
          flush=True)
    rh_deltas = []
    n_rh = len(rhuh_cc)
    for _ in range(N_BOOTSTRAPS):
        idx = rng.integers(0, n_rh, size=n_rh)
        yb = y_rhuh[idx]
        if int(yb.sum()) < 2 or int(len(yb) - yb.sum()) < 2:
            continue
        ac = auroc(X_rhuh_clin[idx] @ bc_mu, yb)
        af = auroc(X_rhuh_full[idx] @ bf_mu, yb)
        rh_deltas.append(af - ac)
    rh_deltas = np.array(rh_deltas)
    rh_var = float(rh_deltas.var())
    rh_mean = float(rh_deltas.mean())
    rh_ci = [float(np.percentile(rh_deltas, 2.5)),
              float(np.percentile(rh_deltas, 97.5))]
    print(f"    Bootstrap (n={len(rh_deltas)}): "
          f"mean={rh_mean:+.4f}, var={rh_var:.5f}, "
          f"95% CI [{rh_ci[0]:+.4f}, {rh_ci[1]:+.4f}], "
          f"P<=0={float((rh_deltas <= 0).mean()):.4f}",
          flush=True)

    # MU bootstrap (in-sample) for meta-analysis
    print(f"  Bootstrapping MU in-sample multi-σ...",
          flush=True)
    mu_deltas = []
    n_mu = len(mu_cc)
    for _ in range(N_BOOTSTRAPS):
        idx = rng.integers(0, n_mu, size=n_mu)
        yb = y_mu[idx]
        if int(yb.sum()) < 5 or int(len(yb) - yb.sum()) < 5:
            continue
        rb = [mu_cc[i] for i in idx]
        Xc, _, _ = build_X_with_stats(rb, feats_clin_xc)
        Xf, _, _ = build_X_with_stats(rb, feats_full_xc)
        bc = logistic_fit(Xc, yb)
        bf = logistic_fit(Xf, yb)
        ac = auroc(Xc @ bc, yb)
        af = auroc(Xf @ bf, yb)
        mu_deltas.append(af - ac)
    mu_deltas = np.array(mu_deltas)
    mu_var = float(mu_deltas.var())
    mu_mean = float(mu_deltas.mean())
    mu_ci = [float(np.percentile(mu_deltas, 2.5)),
              float(np.percentile(mu_deltas, 97.5))]
    print(f"    MU bootstrap: mean={mu_mean:+.4f}, "
          f"var={mu_var:.5f}, 95% CI "
          f"[{mu_ci[0]:+.4f}, {mu_ci[1]:+.4f}]", flush=True)

    # ============================================================
    # PART B: Inverse-variance meta-analysis with multi-σ
    # ============================================================
    print(f"\n=== PART B: IV-weighted meta-analysis (multi-σ) ===",
          flush=True)
    w_mu = 1.0 / mu_var if mu_var > 0 else 0.0
    w_rh = 1.0 / rh_var if rh_var > 0 else 0.0
    pooled = (w_mu * mu_mean + w_rh * rh_mean) / (w_mu + w_rh)
    pooled_se = np.sqrt(1.0 / (w_mu + w_rh))
    pooled_ci = [pooled - 1.96 * pooled_se,
                  pooled + 1.96 * pooled_se]
    z = pooled / pooled_se
    p_meta = 1 - norm.cdf(z)
    print(f"  IV-weighted pooled multi-σ Delta = {pooled:+.4f}",
          flush=True)
    print(f"    SE = {pooled_se:.4f}", flush=True)
    print(f"    95% CI [{pooled_ci[0]:+.4f}, "
          f"{pooled_ci[1]:+.4f}]", flush=True)
    print(f"    z = {z:.3f}, one-sided P = {p_meta:.4f}",
          flush=True)
    print(f"  vs round 43 (single-σ): pooled=+0.083, P=0.036",
          flush=True)

    # ============================================================
    # PART C: Continuous PFS Cox with multi-σ
    # ============================================================
    print(f"\n=== PART C: Continuous PFS Cox with multi-σ ===",
          flush=True)
    # Use full MU (including patients excluded from binary)
    rows_mu_cox = [r for r in mu_rows
                    if all(r[f] is not None
                           for f in feats_full_mu)
                    and r["pfs_days"] is not None
                    and r["progress"] is not None]
    print(f"  Cox sample: n={len(rows_mu_cox)}", flush=True)
    time_arr = np.array([r["pfs_days"] for r in rows_mu_cox],
                         dtype=float)
    event_arr = np.array([r["progress"] for r in rows_mu_cox],
                          dtype=float)
    n_events = int(event_arr.sum())
    print(f"  Events: {n_events}, censored: "
          f"{len(rows_mu_cox) - n_events}", flush=True)

    def build_cox_X(rows, feats):
        arr = np.array([[r[f] for f in feats] for r in rows],
                        dtype=float)
        means = arr.mean(axis=0)
        stds = arr.std(axis=0)
        stds[stds == 0] = 1
        return (arr - means) / stds

    # Cox: clinical only
    X_clin = build_cox_X(rows_mu_cox, feats_clin_3)
    b_clin, pll_clin = cox_fit(X_clin, time_arr, event_arr)
    c_clin = harrells_c(time_arr, event_arr, X_clin @ b_clin)

    # Cox: clinical + V_kernel(σ=3)
    X_vks3 = build_cox_X(rows_mu_cox, feats_clin_3 +
                           ["v_kernel_s3"])
    b_vks3, pll_vks3 = cox_fit(X_vks3, time_arr, event_arr)
    c_vks3 = harrells_c(time_arr, event_arr, X_vks3 @ b_vks3)

    # Cox: clinical + multi-σ
    X_multi = build_cox_X(rows_mu_cox, feats_full_mu)
    b_multi, pll_multi = cox_fit(X_multi, time_arr, event_arr)
    c_multi = harrells_c(time_arr, event_arr, X_multi @ b_multi)

    # LRTs
    lr_vks3 = 2 * (pll_vks3 - pll_clin)
    df_vks3 = 1
    p_lr_vks3 = (1 - chi2.cdf(lr_vks3, df_vks3)
                  if lr_vks3 > 0 else 1.0)
    lr_multi = 2 * (pll_multi - pll_clin)
    df_multi = 4
    p_lr_multi = (1 - chi2.cdf(lr_multi, df_multi)
                   if lr_multi > 0 else 1.0)
    lr_multi_vs_vks3 = 2 * (pll_multi - pll_vks3)
    df_multi_vs_vks3 = 3
    p_lr_multi_vs_vks3 = (1 - chi2.cdf(lr_multi_vs_vks3,
                                         df_multi_vs_vks3)
                           if lr_multi_vs_vks3 > 0 else 1.0)
    print(f"  clin only: C={c_clin:.4f}, pll={pll_clin:.2f}",
          flush=True)
    print(f"  clin + V_k(σ=3): C={c_vks3:.4f}, pll="
          f"{pll_vks3:.2f}, LR={lr_vks3:.2f} (df=1, P="
          f"{p_lr_vks3:.4f})", flush=True)
    print(f"  clin + V_k multi-σ: C={c_multi:.4f}, pll="
          f"{pll_multi:.2f}, LR={lr_multi:.2f} (df=4, P="
          f"{p_lr_multi:.4f})", flush=True)
    print(f"  multi-σ vs σ=3 incremental: LR="
          f"{lr_multi_vs_vks3:.2f} (df=3, P="
          f"{p_lr_multi_vs_vks3:.4f})", flush=True)

    out = {
        "version": "v220",
        "experiment": ("Multi-σ V_kernel comprehensive "
                       "validation: cross-cohort + meta-"
                       "analysis + continuous PFS Cox"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "n_bootstraps": N_BOOTSTRAPS,
        "cohorts": {
            "MU": {"n_cc": len(mu_cc),
                    "n_pos": int(y_mu.sum())},
            "RHUH": {"n_cc": len(rhuh_cc),
                      "n_pos": int(y_rhuh.sum())},
        },
        "cross_cohort": {
            "features_clin": feats_clin_xc,
            "features_full": feats_full_xc,
            "MU_in_sample": {
                "auc_clin": float(auc_mu_clin),
                "auc_full": float(auc_mu_full),
                "delta": float(delta_mu),
                "bootstrap_mean": mu_mean,
                "bootstrap_var": mu_var,
                "bootstrap_95_CI": mu_ci,
            },
            "RHUH_external": {
                "auc_clin": float(auc_rh_clin),
                "auc_full": float(auc_rh_full),
                "delta_point": float(delta_rh),
                "bootstrap_mean": rh_mean,
                "bootstrap_var": rh_var,
                "bootstrap_95_CI": rh_ci,
                "p_one_sided": float(
                    (rh_deltas <= 0).mean()),
            },
        },
        "meta_analysis_multi_sigma": {
            "weight_MU": float(w_mu),
            "weight_RHUH": float(w_rh),
            "pooled_delta": float(pooled),
            "pooled_se": float(pooled_se),
            "pooled_95_CI": pooled_ci,
            "z_score": float(z),
            "one_sided_p": float(p_meta),
            "vs_round_43_single_sigma": {
                "pooled_delta": 0.083,
                "p_one_sided": 0.036,
            },
        },
        "cox_pfs_multi_sigma": {
            "n_total": len(rows_mu_cox),
            "n_events": n_events,
            "clin_only": {
                "C": float(c_clin),
                "partial_loglik": float(pll_clin),
            },
            "clin_plus_Vk_s3": {
                "C": float(c_vks3),
                "partial_loglik": float(pll_vks3),
                "LR_vs_clin": float(lr_vks3),
                "df": int(df_vks3),
                "P_one_sided": float(p_lr_vks3),
            },
            "clin_plus_Vk_multi_sigma": {
                "C": float(c_multi),
                "partial_loglik": float(pll_multi),
                "LR_vs_clin": float(lr_multi),
                "df": int(df_multi),
                "P_one_sided": float(p_lr_multi),
            },
            "multi_sigma_incremental_vs_single_sigma": {
                "LR": float(lr_multi_vs_vks3),
                "df": int(df_multi_vs_vks3),
                "P_one_sided": float(p_lr_multi_vs_vks3),
            },
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
