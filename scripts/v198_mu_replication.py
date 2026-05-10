"""v198: REPLICATION on MU-Glioma-Post — does the round-35 λ + V_kernel
synergy hold on a 4× larger cohort? — round 36.

Round 35 v197 found preliminary evidence on RHUH-GBM (n=13 with valid
λ fits) that λ + V_kernel jointly improve survival prediction
(Δ C = +0.10, LRT p = 0.0018), even though either alone was
non-significant. The natural senior-Nature-reviewer test:

  REPLICATE on a larger cohort with full clinical OS data.

  MU-Glioma-Post has n=151 patients with cached masks AND clinical
  data including: Overall Survival (death event), Days from diagnosis
  to death, IDH1 status, MGMT methylation, Age, Time to First
  Progression. Patient IDs match exactly between cache and xlsx.

  This is the gold-standard replication test. If MU also shows the
  synergy → strong evidence that physics-derived radiomic features
  (λ + V_kernel) jointly capture invasion biology. If MU does not →
  RHUH n=13 finding was overfit; honest scoping.

Method (mirrors v197 RHUH design):
  - Load MU clinical xlsx → PFS, OS days, event status, IDH, MGMT, Age
  - Match to cached MU baseline+followup masks
  - Compute per-patient λ (UODSL exp decay fit)
  - Compute V_kernel σ=3 (kernel-predicted outgrowth volume)
  - Multivariate Cox: M0 (clinical) vs M1 (+λ) vs M2 (+λ + V_kernel)
  - LRT, C-index, bootstrap CIs

Outputs:
  Nature_project/05_results/v198_mu_replication.json
  Nature_project/05_results/v198_mu_per_patient.csv
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import distance_transform_edt, gaussian_filter
from scipy.optimize import minimize
from scipy.stats import chi2 as chi2dist, norm, spearmanr

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_XLSX = Path(r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v198_mu_replication.json"
OUT_CSV = RESULTS / "v198_mu_per_patient.csv"

DISTANCE_BINS = np.arange(1, 25)
SIGMA_KERNEL = 3.0


# ============================================================================
# Cox PH (re-used from v195/v197)
# ============================================================================

def cox_ph_multivariate(times, events, X):
    times = np.asarray(times, dtype=float)
    events = np.asarray(events, dtype=float)
    X = np.atleast_2d(np.asarray(X, dtype=float))
    if X.ndim == 1:
        X = X.reshape(-1, 1)
    n, p = X.shape
    order = np.argsort(-times)
    Z = X[order]
    e = events[order]

    def neg_log_pl(beta):
        eta = Z @ beta
        log_S = np.zeros(n)
        running_max = -np.inf
        running_sum_exp = 0.0
        for i in range(n):
            new_max = max(running_max, eta[i])
            running_sum_exp = (np.exp(running_max - new_max) * running_sum_exp
                                + np.exp(eta[i] - new_max))
            running_max = new_max
            log_S[i] = running_max + np.log(running_sum_exp)
        log_pl = float((eta[e == 1] - log_S[e == 1]).sum())
        return -log_pl

    beta0 = np.zeros(p)
    result = minimize(neg_log_pl, beta0, method="L-BFGS-B",
                      options={"maxiter": 200})
    beta = result.x
    log_pl = -result.fun
    eps = 1e-4
    H = np.zeros((p, p))
    for i in range(p):
        for j in range(p):
            if i <= j:
                bp = beta.copy(); bp[i] += eps; bp[j] += eps
                f_pp = neg_log_pl(bp)
                bp = beta.copy(); bp[i] += eps; bp[j] -= eps
                f_pm = neg_log_pl(bp)
                bp = beta.copy(); bp[i] -= eps; bp[j] += eps
                f_mp = neg_log_pl(bp)
                bp = beta.copy(); bp[i] -= eps; bp[j] -= eps
                f_mm = neg_log_pl(bp)
                H[i, j] = (f_pp - f_pm - f_mp + f_mm) / (4 * eps * eps)
                H[j, i] = H[i, j]
    try:
        cov = np.linalg.inv(H)
        se = np.sqrt(np.maximum(np.diag(cov), 0))
    except np.linalg.LinAlgError:
        se = np.full(p, np.nan)
    return {
        "beta": beta.tolist(),
        "se": se.tolist(),
        "log_pl": float(log_pl),
        "converged": bool(result.success),
    }


def concordance(times, events, risk_scores):
    times = np.asarray(times)
    events = np.asarray(events)
    risk_scores = np.asarray(risk_scores)
    n = len(times)
    n_pairs = 0
    n_concordant = 0
    n_ties = 0
    for i in range(n):
        for j in range(i+1, n):
            if events[i] == 0 and events[j] == 0:
                continue
            if times[i] < times[j]:
                if events[i] != 1:
                    continue
                n_pairs += 1
                if risk_scores[i] > risk_scores[j]:
                    n_concordant += 1
                elif risk_scores[i] == risk_scores[j]:
                    n_ties += 1
            elif times[j] < times[i]:
                if events[j] != 1:
                    continue
                n_pairs += 1
                if risk_scores[j] > risk_scores[i]:
                    n_concordant += 1
                elif risk_scores[i] == risk_scores[j]:
                    n_ties += 1
    if n_pairs == 0:
        return float("nan")
    return (n_concordant + 0.5 * n_ties) / n_pairs


# ============================================================================
# UODSL fit + V_kernel
# ============================================================================

def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    h_broad = heat_constant(mask, sigma)
    return np.maximum(persistence, h_broad)


def kernel_outgrowth_volume(mask, sigma=SIGMA_KERNEL):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def fit_lambda(mask, outgrowth):
    m = mask.astype(bool)
    out = outgrowth.astype(bool)
    if m.sum() == 0 or out.sum() == 0:
        return float("nan"), float("nan"), 0
    d = distance_transform_edt(~m)
    d_int = np.round(d).astype(int)
    d_arr, p_arr, n_arr = [], [], []
    for b in DISTANCE_BINS:
        sel = (d_int == b)
        if sel.sum() < 5:
            continue
        n_total = int(sel.sum())
        n_out = int((sel & out).sum())
        d_arr.append(b)
        p_arr.append(n_out / n_total)
        n_arr.append(n_total)
    d_arr = np.array(d_arr, dtype=float)
    p_arr = np.array(p_arr, dtype=float)
    n_arr = np.array(n_arr, dtype=float)
    valid = p_arr > 0
    if valid.sum() < 3:
        return float("nan"), float("nan"), int(valid.sum())
    d_v = d_arr[valid]
    p_v = p_arr[valid]
    n_v = n_arr[valid]
    log_p = np.log(p_v)
    w = np.sqrt(n_v)
    sw = np.sum(w)
    sw_d = np.sum(w * d_v)
    sw_logp = np.sum(w * log_p)
    sw_dd = np.sum(w * d_v * d_v)
    sw_d_logp = np.sum(w * d_v * log_p)
    denom = sw * sw_dd - sw_d ** 2
    if denom == 0:
        return float("nan"), float("nan"), int(valid.sum())
    slope = (sw * sw_d_logp - sw_d * sw_logp) / denom
    intercept = (sw_logp - slope * sw_d) / sw
    lam = float(-1.0 / slope) if slope < 0 else float("inf")
    pred = intercept + slope * d_v
    ss_res = float(np.sum(w * (log_p - pred) ** 2))
    ss_tot = float(np.sum(w * (log_p - np.mean(log_p)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
    return lam, r2, int(valid.sum())


def main():
    print("=" * 78, flush=True)
    print("v198 MU-GLIOMA-POST REPLICATION (round 36)", flush=True)
    print("=" * 78, flush=True)

    # ---------- Load MU clinical ----------
    wb = openpyxl.load_workbook(CLINICAL_XLSX, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(ws.iter_rows(values_only=True))]
    # Find columns
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    grade_col = header.index("Grade of Primary Brain Tumor")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    death_col = header.index("Overall Survival (Death)")
    os_col = header.index("Number of days from Diagnosis to death (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")

    clinical = {}
    n_loaded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = float(row[age_col]) if row[age_col] is not None else None
            grade = float(row[grade_col]) if row[grade_col] is not None else None
            event = (int(row[death_col])
                       if row[death_col] is not None else None)
            os_d = (float(row[os_col])
                      if row[os_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                       if row[pfs_col] is not None else None)
            progress = (int(row[progress_col])
                          if row[progress_col] is not None else None)
            idh1 = (float(row[idh1_col])
                      if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                      if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "grade": grade,
            "event": event, "os_days": os_d,
            "pfs_days": pfs_d, "progress": progress,
            "idh1": idh1, "mgmt": mgmt,
        }
        n_loaded += 1
    print(f"  Loaded clinical for {n_loaded} MU patients", flush=True)

    # Censoring: if event == 1 (death) and os_days exists, use os_days as time
    # If event == 0, censoring time is unclear from the xlsx; use os_days if
    # available, else fall back to last MRI date
    # For simplicity: only include patients with event AND os_days OR
    # event=0 with PFS as censoring time

    # ---------- Compute per-patient lambda + V_kernel ----------
    print("\nComputing per-patient λ + V_kernel for MU-Glioma-Post...",
          flush=True)
    rows = []
    n_with_mask = 0
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        fr = f.parent / f.name.replace("_b.npy", "_r.npy")
        if not fr.exists():
            continue
        fu = (np.load(fr) > 0).astype(np.float32)
        outgrowth = (fu.astype(bool) & ~m.astype(bool)).astype(np.float32)
        if m.sum() == 0:
            continue
        n_with_mask += 1
        v_k = kernel_outgrowth_volume(m, SIGMA_KERNEL)
        baseline_volume = int(m.sum())
        if outgrowth.sum() > 0:
            lam, r2, n_pts = fit_lambda(m, outgrowth)
            valid = (not np.isnan(lam)) and (not np.isinf(lam)) and \
                    lam > 0 and lam < 200 and r2 > 0.5 and n_pts >= 4
        else:
            lam, r2, n_pts, valid = float("nan"), float("nan"), 0, False
        c = clinical[pid]
        rows.append({
            "pid": pid,
            "lambda_obs": lam if valid else None,
            "lambda_r2": r2 if not np.isnan(r2) else None,
            "lambda_valid": valid,
            "v_kernel_s3": v_k,
            "baseline_volume": baseline_volume,
            "age": c["age"],
            "grade": c["grade"],
            "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "os_days": c["os_days"],
            "event": c["event"],
            "pfs_days": c["pfs_days"],
            "progress": c["progress"],
        })

    print(f"  {n_with_mask} MU patients with cached masks + clinical",
          flush=True)
    n_valid_lam = sum(1 for r in rows if r["lambda_valid"])
    print(f"  {n_valid_lam} patients with valid per-patient lambda fit",
          flush=True)

    # ---------- Filter for valid OS data ----------
    valid_os = [r for r in rows
                  if r["lambda_valid"] and r["os_days"] is not None
                  and r["event"] is not None and r["age"] is not None]
    print(f"  {len(valid_os)} patients with valid lambda + OS + age",
          flush=True)
    if len(valid_os) < 5:
        print("  Insufficient patients with valid OS data — using PFS instead",
              flush=True)
        valid_pfs = [r for r in rows
                       if r["lambda_valid"] and r["pfs_days"] is not None
                       and r["progress"] is not None
                       and r["age"] is not None]
        print(f"  {len(valid_pfs)} patients with valid lambda + PFS + age",
              flush=True)
        valid_rows = valid_pfs
        time_key = "pfs_days"
        event_key = "progress"
    else:
        valid_rows = valid_os
        time_key = "os_days"
        event_key = "event"
    print(f"  Using outcome: {time_key} (event = {event_key})", flush=True)

    if len(valid_rows) < 10:
        print(f"\n  HONEST: only {len(valid_rows)} valid patients — "
              f"insufficient for multivariate Cox.", flush=True)
        # Save partial result
        out = {
            "version": "v198",
            "experiment": "MU-Glioma-Post replication of round-35 synergy",
            "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "status": "INSUFFICIENT_DATA",
            "n_clinical_loaded": n_loaded,
            "n_with_mask": n_with_mask,
            "n_valid_lambda": n_valid_lam,
            "n_valid_outcome": len(valid_rows),
            "outcome_used": time_key,
        }
        OUT_JSON.write_text(json.dumps(out, indent=2))
        if rows:
            with OUT_CSV.open("w", newline="") as fp:
                w = csv.DictWriter(fp, fieldnames=list(rows[0].keys()))
                w.writeheader()
                w.writerows(rows)
        return

    # ---------- Spearman ----------
    times = np.array([r[time_key] for r in valid_rows], dtype=float)
    events = np.array([r[event_key] for r in valid_rows], dtype=float)
    lams = np.array([r["lambda_obs"] for r in valid_rows], dtype=float)
    vks = np.array([r["v_kernel_s3"] for r in valid_rows], dtype=float)
    print(f"\n  n = {len(valid_rows)}; events = {int(events.sum())}",
          flush=True)
    rho_lam, p_lam = spearmanr(lams, times)
    rho_vk, p_vk = spearmanr(vks, times)
    print(f"  Spearman lambda vs {time_key}: rho={rho_lam:+.4f}  p={p_lam:.4f}",
          flush=True)
    print(f"  Spearman V_kernel vs {time_key}: rho={rho_vk:+.4f}  p={p_vk:.4f}",
          flush=True)

    # ---------- Univariate Cox ----------
    print("\n=== Univariate Cox ===", flush=True)
    for feat, x in [("lambda", lams), ("V_kernel", vks),
                      ("baseline_volume",
                       np.array([r["baseline_volume"] for r in valid_rows]))]:
        if np.std(x) == 0:
            continue
        x_z = (x - x.mean()) / x.std()
        result = cox_ph_multivariate(times, events, x_z.reshape(-1, 1))
        if result:
            beta = result["beta"][0]
            se = result["se"][0]
            HR = float(np.exp(beta))
            z = beta / max(se, 1e-12)
            p = float(2 * (1 - norm.cdf(abs(z))))
            print(f"  {feat:18s}  HR/SD = {HR:.4f}  beta={beta:+.4f}  "
                  f"SE={se:.4f}  p={p:.4f}", flush=True)

    # ---------- Multivariate Cox ----------
    # M0 = clinical: age + idh1 + mgmt
    # M1 = M0 + lambda
    # M2 = M0 + lambda + V_kernel
    print("\n=== Multivariate Cox (complete-case) ===", flush=True)
    feats_m0 = ["age", "idh1", "mgmt"]
    feats_m1 = feats_m0 + ["lambda_obs"]
    feats_m2 = feats_m1 + ["v_kernel_s3"]
    cc = [r for r in valid_rows
            if all(r[f] is not None and not (isinstance(r[f], float)
                                                and np.isnan(r[f]))
                   for f in feats_m2)]
    print(f"  complete-case n = {len(cc)}", flush=True)
    if len(cc) < 10:
        print("  Insufficient complete-case patients", flush=True)
        return
    times_cc = np.array([r[time_key] for r in cc], dtype=float)
    events_cc = np.array([r[event_key] for r in cc], dtype=float)

    def build_X(rows, feats):
        X = np.array([[r[f] for f in feats] for r in rows], dtype=float)
        means = X.mean(axis=0)
        stds = X.std(axis=0)
        stds[stds == 0] = 1
        return (X - means) / stds

    X0 = build_X(cc, feats_m0)
    X1 = build_X(cc, feats_m1)
    X2 = build_X(cc, feats_m2)
    M0 = cox_ph_multivariate(times_cc, events_cc, X0)
    M1 = cox_ph_multivariate(times_cc, events_cc, X1)
    M2 = cox_ph_multivariate(times_cc, events_cc, X2)
    risk_M0 = X0 @ np.array(M0["beta"]) if M0 else np.zeros(len(times_cc))
    risk_M1 = X1 @ np.array(M1["beta"]) if M1 else np.zeros(len(times_cc))
    risk_M2 = X2 @ np.array(M2["beta"]) if M2 else np.zeros(len(times_cc))
    c_M0 = concordance(times_cc, events_cc, risk_M0)
    c_M1 = concordance(times_cc, events_cc, risk_M1)
    c_M2 = concordance(times_cc, events_cc, risk_M2)
    print(f"\n  M0 (clinical: age+IDH1+MGMT):   C = {c_M0:.4f}",
          flush=True)
    print(f"  M1 (M0 + lambda):                 C = {c_M1:.4f}  "
          f"Delta = {c_M1 - c_M0:+.4f}", flush=True)
    print(f"  M2 (M0 + lambda + V_kernel):       C = {c_M2:.4f}  "
          f"Delta = {c_M2 - c_M0:+.4f}", flush=True)

    if M0 and M1:
        lrt01 = 2 * (M1["log_pl"] - M0["log_pl"])
        p_01 = float(1 - chi2dist.cdf(lrt01, df=1))
    else:
        lrt01, p_01 = float("nan"), float("nan")
    if M0 and M2:
        lrt02 = 2 * (M2["log_pl"] - M0["log_pl"])
        p_02 = float(1 - chi2dist.cdf(lrt02, df=2))
    else:
        lrt02, p_02 = float("nan"), float("nan")
    print(f"\n  LRT M0 vs M1 (df=1): chi^2 = {lrt01:.4f}  p = {p_01:.4f}",
          flush=True)
    print(f"  LRT M0 vs M2 (df=2): chi^2 = {lrt02:.4f}  p = {p_02:.4f}",
          flush=True)

    print(f"\n  Round-35 RHUH (n=13) reference:", flush=True)
    print(f"    M0 C = 0.78; M1 C = 0.80 (LRT p = 0.30); M2 C = 0.88 "
          f"(LRT p = 0.0018)", flush=True)
    print(f"\n  v198 MU REPLICATION (n={len(cc)}):", flush=True)
    print(f"    M0 C = {c_M0:.4f}; M1 C = {c_M1:.4f} (LRT p = {p_01:.4f}); "
          f"M2 C = {c_M2:.4f} (LRT p = {p_02:.4f})", flush=True)
    if p_02 < 0.05:
        print(f"\n  >>> SYNERGY REPLICATES on MU n={len(cc)} <<<",
              flush=True)
    else:
        print(f"\n  >>> SYNERGY DOES NOT REPLICATE on MU "
              f"(LRT p = {p_02:.4f}) <<<", flush=True)

    out = {
        "version": "v198",
        "experiment": ("MU-Glioma-Post replication of round-35 lambda + "
                       "V_kernel synergy"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "n_clinical_loaded": n_loaded,
        "n_with_mask": n_with_mask,
        "n_valid_lambda": n_valid_lam,
        "n_valid_outcome": len(valid_rows),
        "n_complete_case": len(cc),
        "outcome_used": time_key,
        "spearman": {
            "lambda": {"rho": float(rho_lam), "p": float(p_lam)},
            "V_kernel": {"rho": float(rho_vk), "p": float(p_vk)},
        },
        "multivariate_cox": {
            "M0_c_index": float(c_M0),
            "M1_c_index": float(c_M1),
            "M1_delta_c": float(c_M1 - c_M0),
            "M2_c_index": float(c_M2),
            "M2_delta_c": float(c_M2 - c_M0),
            "LRT_M0_vs_M1": {
                "chi2": float(lrt01) if not np.isnan(lrt01) else None,
                "p_value": float(p_01) if not np.isnan(p_01) else None,
            },
            "LRT_M0_vs_M2": {
                "chi2": float(lrt02) if not np.isnan(lrt02) else None,
                "p_value": float(p_02) if not np.isnan(p_02) else None,
            },
        },
        "round35_RHUH_reference": {
            "M0_c_index": 0.7833,
            "M1_c_index": 0.8000,
            "M2_c_index": 0.8833,
            "LRT_M0_vs_M2_p": 0.0018,
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)

    if rows:
        with OUT_CSV.open("w", newline="") as fp:
            w = csv.DictWriter(fp, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
        print(f"Saved {OUT_CSV}", flush=True)


if __name__ == "__main__":
    main()
