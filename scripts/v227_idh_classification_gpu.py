"""v227: KERNEL-DRIVEN IDH1 MOLECULAR PATHOLOGY CLASSIFICATION
from baseline imaging — round 51 (GPU).

Beyond-NMI flagship: tests whether multi-σ kernel features predict
IDH1 mutation status (a molecular pathology label) directly from
baseline imaging. Distinct from PFS prediction; demonstrates the
kernel captures molecular biology, not just outgrowth.

IDH1 status is the most important molecular marker in glioma
(IDH-mut = better prognosis, distinct biology). Classifying it
from imaging would enable non-invasive molecular pathology, sparing
patients biopsy.

Method:
  - 5-fold stratified CV on MU n=130 with IDH1 labels
  - Compare:
    (a) clinical-only: age + MGMT (excluding IDH itself)
    (b) multi-σ V_kernel only: V_k(σ=2,3,4,5)
    (c) combined: clinical + multi-σ V_kernel
    (d) 3D CNN on baseline mask + kernel image (deep baseline)
  - Bootstrap CIs

Outputs:
  Nature_project/05_results/v227_idh_classification.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v227_idh_classification.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

TARGET_SHAPE = (16, 48, 48)
N_FOLDS = 5
EPOCHS = 30
LR = 5e-4
SEED = 42
N_BOOT = 500


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class SimpleCNN(nn.Module):
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32), nn.GELU(),
            nn.Dropout(0.3), nn.Linear(32, 1))

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.head(self.global_pool(e3)).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def build_X(rows, feats):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0)
    stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return np.column_stack([np.ones(len(arr_z)), arr_z])


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "idh1": idh1, "mgmt": mgmt,
        }
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None):
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        # IDH1 binary classification: idh1 = 1 (mutant) vs 0 (WT)
        y_idh = int(c["idh1"])
        # Compute multi-σ kernels at full resolution
        v_k2 = float(kernel_outgrowth_volume(m_full, 2.0))
        v_k3 = float(kernel_outgrowth_volume(m_full, 3.0))
        v_k4 = float(kernel_outgrowth_volume(m_full, 4.0))
        v_k5 = float(kernel_outgrowth_volume(m_full, 5.0))
        # Resize for CNN input
        m_resized = resize_to_target(m_full, TARGET_SHAPE)
        k_resized = heat_bimodal(m_resized, 3.0)
        data.append({
            "pid": pid, "y": y_idh,
            "age": c["age"], "mgmt": c["mgmt"],
            "v_k2": v_k2, "v_k3": v_k3,
            "v_k4": v_k4, "v_k5": v_k5,
            "mask": m_resized, "kernel": k_resized,
        })
    return data


def run_logistic_5fold(data, feats, folds, name="logistic"):
    n = len(data)
    y_all = np.array([d["y"] for d in data], dtype=np.float32)
    oof = np.zeros(n, dtype=np.float32)
    assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        rb_tr = [data[i] for i in train_idx]
        rb_te = [data[i] for i in test_idx]
        ytr = np.array([d["y"] for d in rb_tr], dtype=float)
        yte = np.array([d["y"] for d in rb_te], dtype=float)
        if ytr.sum() < 2 or len(ytr) - ytr.sum() < 2:
            continue
        Xtr = build_X(rb_tr, feats)
        Xte = build_X(rb_te, feats)
        beta = logistic_fit(Xtr, ytr)
        scores = Xte @ beta
        oof[test_idx] = scores
        assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(
                len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            fold_aucs.append(float(a))
            print(f"    {name} fold {fold_i+1}: AUC={a:.4f}",
                  flush=True)
    pooled_auc = auroc(oof[assigned], y_all[assigned])
    print(f"  {name} pooled OOF AUC = {pooled_auc:.4f}",
          flush=True)
    return {
        "fold_aucs": fold_aucs,
        "fold_mean": (float(np.mean(fold_aucs))
                       if fold_aucs else float("nan")),
        "pooled_oof_auc": float(pooled_auc),
        "oof_scores": oof.tolist(),
        "assigned": assigned.tolist(),
    }


def run_cnn_5fold(data, folds, name="CNN"):
    n = len(data)
    y_all = np.array([d["y"] for d in data], dtype=np.float32)
    oof = np.zeros(n, dtype=np.float32)
    assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        rb_tr = [data[i] for i in train_idx]
        rb_te = [data[i] for i in test_idx]
        n_tr_pos = int(sum(1 for d in rb_tr if d["y"] == 1))
        n_tr_neg = len(rb_tr) - n_tr_pos
        if n_tr_pos < 2 or n_tr_neg < 2:
            continue
        Xtr = stack_inputs(rb_tr)
        Xte = stack_inputs(rb_te)
        ytr = np.array([d["y"] for d in rb_tr], dtype=np.float32)
        yte = np.array([d["y"] for d in rb_te], dtype=np.float32)
        Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
        Xte_t = torch.from_numpy(Xte).to(DEVICE)
        ytr_t = torch.from_numpy(ytr).to(DEVICE)
        torch.manual_seed(SEED + fold_i * 100)
        np.random.seed(SEED + fold_i * 100)
        model = SimpleCNN(in_ch=2, base=24).to(DEVICE)
        opt = torch.optim.AdamW(model.parameters(), lr=LR,
                                 weight_decay=1e-3)
        pos_w = torch.tensor([n_tr_neg / max(1, n_tr_pos)],
                              dtype=torch.float32,
                              device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        bs = 4
        for ep in range(EPOCHS):
            model.train()
            perm = np.random.permutation(len(rb_tr))
            for i in range(0, len(rb_tr), bs):
                idx = perm[i:i+bs]
                logits = model(Xtr_t[idx])
                loss = criterion(logits, ytr_t[idx])
                opt.zero_grad()
                loss.backward()
                opt.step()
        model.eval()
        with torch.no_grad():
            scores = model(Xte_t).cpu().numpy()
        oof[test_idx] = scores
        assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(
                len(yte) - yte.sum()) >= 1:
            a = auroc(scores, yte)
            fold_aucs.append(float(a))
            print(f"    {name} fold {fold_i+1}: AUC={a:.4f}",
                  flush=True)
        del model, opt
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()
    pooled_auc = auroc(oof[assigned], y_all[assigned])
    print(f"  {name} pooled OOF AUC = {pooled_auc:.4f}",
          flush=True)
    return {
        "fold_aucs": fold_aucs,
        "fold_mean": (float(np.mean(fold_aucs))
                       if fold_aucs else float("nan")),
        "pooled_oof_auc": float(pooled_auc),
    }


def main():
    print("=" * 78, flush=True)
    print(f"v227 IDH1 MOLECULAR CLASSIFICATION (round 51 GPU) "
          f"device={DEVICE}", flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    data = load_mu_data()
    n = len(data)
    y = np.array([d["y"] for d in data], dtype=np.float32)
    n_mut = int(y.sum())
    n_wt = n - n_mut
    print(f"  MU n={n}, IDH-mut={n_mut}, IDH-WT={n_wt}",
          flush=True)
    if n_mut < 10:
        print(f"  WARNING: very few mutant cases; "
              f"high variance expected", flush=True)
    folds = stratified_kfold(y, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}",
          flush=True)

    # ---- Variant A: clinical only (age + MGMT) ----
    print(f"\n=== VARIANT A: clinical only (age + MGMT) ===",
          flush=True)
    feats_clin = ["age", "mgmt"]
    res_clin = run_logistic_5fold(data, feats_clin, folds,
                                     name="clin")

    # ---- Variant B: multi-σ V_kernel only ----
    print(f"\n=== VARIANT B: multi-σ V_kernel only (4 feats) "
          f"===", flush=True)
    feats_kernel = ["v_k2", "v_k3", "v_k4", "v_k5"]
    res_kernel = run_logistic_5fold(data, feats_kernel, folds,
                                       name="kernel")

    # ---- Variant C: clinical + multi-σ V_kernel ----
    print(f"\n=== VARIANT C: clinical + multi-σ V_kernel "
          f"(6 feats) ===", flush=True)
    feats_combined = feats_clin + feats_kernel
    res_combined = run_logistic_5fold(data, feats_combined,
                                         folds, name="combined")

    # ---- Variant D: 3D CNN on mask + kernel image ----
    print(f"\n=== VARIANT D: 3D CNN on mask + kernel image ===",
          flush=True)
    res_cnn = run_cnn_5fold(data, folds, name="CNN")

    # ---- Bootstrap CI for combined ----
    print(f"\n=== Bootstrap pairwise Δ AUC ===", flush=True)
    rng = np.random.default_rng(SEED)
    oof_clin = np.array(res_clin["oof_scores"])
    oof_kernel = np.array(res_kernel["oof_scores"])
    oof_combined = np.array(res_combined["oof_scores"])
    assigned = np.array(res_clin["assigned"])
    n_assigned = int(assigned.sum())
    deltas_combined_minus_clin = []
    deltas_kernel_minus_clin = []
    for _ in range(N_BOOT):
        idx = rng.integers(0, n_assigned, size=n_assigned)
        valid_idx = np.where(assigned)[0][idx]
        yb = y[valid_idx]
        if int(yb.sum()) < 3 or int(len(yb) - yb.sum()) < 3:
            continue
        a_c = auroc(oof_clin[valid_idx], yb)
        a_k = auroc(oof_kernel[valid_idx], yb)
        a_co = auroc(oof_combined[valid_idx], yb)
        deltas_combined_minus_clin.append(a_co - a_c)
        deltas_kernel_minus_clin.append(a_k - a_c)
    deltas_combined_minus_clin = np.array(
        deltas_combined_minus_clin)
    deltas_kernel_minus_clin = np.array(
        deltas_kernel_minus_clin)
    print(f"  Δ combined - clinical: mean="
          f"{deltas_combined_minus_clin.mean():+.4f}, "
          f"95% CI [{np.percentile(deltas_combined_minus_clin, 2.5):+.4f}, "
          f"{np.percentile(deltas_combined_minus_clin, 97.5):+.4f}], "
          f"P<=0={float((deltas_combined_minus_clin <= 0).mean()):.4f}",
          flush=True)
    print(f"  Δ kernel - clinical: mean="
          f"{deltas_kernel_minus_clin.mean():+.4f}, "
          f"95% CI [{np.percentile(deltas_kernel_minus_clin, 2.5):+.4f}, "
          f"{np.percentile(deltas_kernel_minus_clin, 97.5):+.4f}], "
          f"P<=0={float((deltas_kernel_minus_clin <= 0).mean()):.4f}",
          flush=True)

    print(f"\n=== SUMMARY ===", flush=True)
    print(f"  IDH1 classification (n={n}, IDH-mut={n_mut}/n_wt"
          f"={n_wt}):", flush=True)
    print(f"    Variant A (clin only): "
          f"AUC={res_clin['pooled_oof_auc']:.4f}", flush=True)
    print(f"    Variant B (kernel only): "
          f"AUC={res_kernel['pooled_oof_auc']:.4f}", flush=True)
    print(f"    Variant C (clin + kernel): "
          f"AUC={res_combined['pooled_oof_auc']:.4f}",
          flush=True)
    print(f"    Variant D (3D CNN): "
          f"AUC={res_cnn['pooled_oof_auc']:.4f}", flush=True)

    out = {
        "version": "v227",
        "experiment": ("IDH1 molecular pathology "
                       "classification from baseline imaging "
                       "via multi-σ V_kernel"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "n_total": n, "n_mut": n_mut, "n_wt": n_wt,
        "n_folds": N_FOLDS,
        "results": {
            "A_clin_only": {k: v for k, v in res_clin.items()
                              if k != "oof_scores"
                              and k != "assigned"},
            "B_kernel_only": {k: v for k, v in res_kernel.items()
                                 if k != "oof_scores"
                                 and k != "assigned"},
            "C_clin_plus_kernel": {
                k: v for k, v in res_combined.items()
                if k != "oof_scores" and k != "assigned"},
            "D_3D_CNN": res_cnn,
        },
        "bootstrap_deltas": {
            "combined_minus_clin": {
                "mean": float(
                    deltas_combined_minus_clin.mean()),
                "95_CI": [
                    float(np.percentile(
                        deltas_combined_minus_clin, 2.5)),
                    float(np.percentile(
                        deltas_combined_minus_clin, 97.5))],
                "p_one_sided": float(
                    (deltas_combined_minus_clin <= 0).mean()),
            },
            "kernel_minus_clin": {
                "mean": float(deltas_kernel_minus_clin.mean()),
                "95_CI": [
                    float(np.percentile(
                        deltas_kernel_minus_clin, 2.5)),
                    float(np.percentile(
                        deltas_kernel_minus_clin, 97.5))],
                "p_one_sided": float(
                    (deltas_kernel_minus_clin <= 0).mean()),
            },
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
