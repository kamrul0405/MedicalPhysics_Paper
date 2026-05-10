"""v212: NRI + IDI + Brier-score reclassification statistics —
round 44 (CPU).

The Lancet / JAMA / NEJM-standard tests for "does adding V_kernel
improve clinical classification?" beyond the AUC lift in round 39
v202 (Delta AUC=+0.108) and the meta-analysis pooled Delta in round
43 v210. Three complementary reclassification metrics:

  1. CONTINUOUS NRI (Net Reclassification Index): for each patient,
     compute the change in predicted probability between clinical-
     only and clinical+V_kernel. NRI_pos: fraction of progressors
     whose probability moved up minus moved down. NRI_neg:
     fraction of non-progressors whose probability moved down minus
     up. Total NRI = NRI_pos + NRI_neg.

  2. CATEGORICAL NRI: using clinically meaningful risk thresholds
     [0.25, 0.50, 0.75], count patients moving to higher vs lower
     risk categories.

  3. IDI (Integrated Discrimination Improvement): difference in
     mean predicted probability between events and non-events,
     compared between the two models.

  4. BRIER SCORE: BS = mean((p_pred - y)^2). Decomposition into
     reliability (calibration), resolution (variance of conditional
     expectations), and uncertainty (irreducible variance).

  5. BRIER SKILL SCORE: BSS = 1 - BS_model / BS_reference, where
     reference = predict prevalence for everyone.

Outputs:
  Nature_project/05_results/v212_nri_idi_brier.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize
from scipy.stats import norm

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v212_nri_idi_brier.json"

SIGMA_KERNEL = 3.0
HORIZON = 365
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def build_X(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def label_binary(rows, X, pfs_field="pfs_days",
                  prog_field="progress"):
    out = []
    for r in rows:
        pfs = r[pfs_field]
        prog = r[prog_field]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def continuous_nri(p_old, p_new, y):
    """Continuous NRI: fraction of events with p_new > p_old minus
    p_new < p_old, plus same for non-events with sign flipped."""
    y = np.asarray(y).astype(int)
    p_old = np.asarray(p_old)
    p_new = np.asarray(p_new)
    pos = y == 1
    neg = y == 0
    if pos.sum() == 0 or neg.sum() == 0:
        return float("nan"), float("nan"), float("nan")
    nri_pos = ((p_new[pos] > p_old[pos]).mean()
                - (p_new[pos] < p_old[pos]).mean())
    nri_neg = ((p_new[neg] < p_old[neg]).mean()
                - (p_new[neg] > p_old[neg]).mean())
    return float(nri_pos + nri_neg), float(nri_pos), float(nri_neg)


def categorical_nri(p_old, p_new, y, thresholds):
    """Categorical NRI using risk-category thresholds. Categories
    are defined by intervals between thresholds (e.g., [0.25, 0.5,
    0.75] -> 4 categories: <0.25, 0.25-0.5, 0.5-0.75, >=0.75)."""
    y = np.asarray(y).astype(int)
    p_old = np.asarray(p_old)
    p_new = np.asarray(p_new)
    cat_old = np.digitize(p_old, thresholds)
    cat_new = np.digitize(p_new, thresholds)
    pos = y == 1
    neg = y == 0
    if pos.sum() == 0 or neg.sum() == 0:
        return float("nan"), float("nan"), float("nan")
    up_pos = (cat_new[pos] > cat_old[pos]).mean()
    dn_pos = (cat_new[pos] < cat_old[pos]).mean()
    up_neg = (cat_new[neg] > cat_old[neg]).mean()
    dn_neg = (cat_new[neg] < cat_old[neg]).mean()
    nri_pos = up_pos - dn_pos
    nri_neg = dn_neg - up_neg
    return float(nri_pos + nri_neg), float(nri_pos), float(nri_neg)


def idi(p_old, p_new, y):
    """Integrated Discrimination Improvement.
    IDI = (mean(p_new[y=1]) - mean(p_old[y=1])) -
          (mean(p_new[y=0]) - mean(p_old[y=0]))."""
    y = np.asarray(y).astype(int)
    p_old = np.asarray(p_old)
    p_new = np.asarray(p_new)
    pos = y == 1
    neg = y == 0
    if pos.sum() == 0 or neg.sum() == 0:
        return float("nan")
    delta_pos = float(p_new[pos].mean() - p_old[pos].mean())
    delta_neg = float(p_new[neg].mean() - p_old[neg].mean())
    return delta_pos - delta_neg


def brier_score(p, y):
    p = np.asarray(p)
    y = np.asarray(y).astype(int)
    return float(((p - y) ** 2).mean())


def brier_decomposition(p, y, n_bins=10):
    """Brier = Reliability - Resolution + Uncertainty."""
    p = np.asarray(p)
    y = np.asarray(y).astype(int)
    bin_edges = np.linspace(0, 1, n_bins + 1)
    cidx = np.digitize(p, bin_edges) - 1
    cidx = np.clip(cidx, 0, n_bins - 1)
    n = len(y)
    o_total = float(y.mean())
    rel = 0.0
    res = 0.0
    for b in range(n_bins):
        sel = cidx == b
        if sel.sum() == 0:
            continue
        nb = int(sel.sum())
        pb = float(p[sel].mean())
        ob = float(y[sel].mean())
        rel += (nb / n) * (pb - ob) ** 2
        res += (nb / n) * (ob - o_total) ** 2
    unc = o_total * (1 - o_total)
    bs = float(rel - res + unc)
    return {
        "brier": bs,
        "reliability": float(rel),
        "resolution": float(res),
        "uncertainty": float(unc),
    }


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def load_mu():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid, "cohort": "MU",
            "v_kernel_s3": kernel_outgrowth_volume(m,
                                                    SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def load_rhuh():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        age_col = next((c for c in cols if c.strip() == "Age"),
                        None)
        idh_col = next((c for c in cols if "IDH status" in c),
                        None)
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                age = (float(row[age_col]) if row[age_col]
                       else None)
                idh_str = (row[idh_col].strip().lower()
                            if row[idh_col] else "")
                if "mut" in idh_str:
                    idh1 = 1
                elif "wt" in idh_str or "wild" in idh_str:
                    idh1 = 0
                else:
                    idh1 = None
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if age is None or idh1 is None or pfs_d is None:
                continue
            clinical[pid] = {
                "age": age, "idh1": idh1,
                "pfs_days": pfs_d, "progress": event,
            }
    rows = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        rows.append({
            "pid": pid, "cohort": "RHUH",
            "v_kernel_s3": kernel_outgrowth_volume(m,
                                                    SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    return rows


def reclassification_analysis(rows_lab, feats_clin, feats_full,
                                rng):
    """Compute NRI/IDI/Brier with bootstrap CIs."""
    cc = [l for l in rows_lab if all(
        l[f] is not None and not (isinstance(l[f], float)
                                  and np.isnan(l[f]))
        for f in feats_full)]
    n = len(cc)
    if n < 20:
        return None
    y = np.array([l["y"] for l in cc], dtype=float)
    n_pos = int(y.sum())
    n_neg = n - n_pos
    if n_pos < 5 or n_neg < 5:
        return None
    Xc, m_c, s_c = build_X(cc, feats_clin)
    Xf, m_f, s_f = build_X(cc, feats_full)
    bc = logistic_fit(Xc, y)
    bf = logistic_fit(Xf, y)
    p_old = sigmoid(Xc @ bc)
    p_new = sigmoid(Xf @ bf)

    # Point estimates
    nri_c, nri_pos_c, nri_neg_c = continuous_nri(p_old, p_new, y)
    nri_cat, nri_pos_cat, nri_neg_cat = categorical_nri(
        p_old, p_new, y, thresholds=[0.25, 0.5, 0.75])
    idi_v = idi(p_old, p_new, y)
    bs_c = brier_score(p_old, y)
    bs_f = brier_score(p_new, y)
    prev = float(y.mean())
    bs_ref = prev * (1 - prev)  # always-prevalence Brier
    bss_c = 1 - bs_c / bs_ref if bs_ref > 0 else float("nan")
    bss_f = 1 - bs_f / bs_ref if bs_ref > 0 else float("nan")
    decomp_c = brier_decomposition(p_old, y)
    decomp_f = brier_decomposition(p_new, y)
    auc_c = auroc(p_old, y)
    auc_f = auroc(p_new, y)

    # Bootstrap CIs
    nri_c_bs = []
    nri_cat_bs = []
    idi_bs = []
    delta_bss_bs = []
    delta_brier_bs = []
    for _ in range(N_BOOTSTRAPS):
        idx = rng.integers(0, n, size=n)
        yb = y[idx]
        if int(yb.sum()) < 3 or int(len(yb) - yb.sum()) < 3:
            continue
        po_b = p_old[idx]
        pn_b = p_new[idx]
        # NRI
        n_c, _, _ = continuous_nri(po_b, pn_b, yb)
        n_cat, _, _ = categorical_nri(po_b, pn_b, yb,
                                        thresholds=[0.25, 0.5, 0.75])
        ii = idi(po_b, pn_b, yb)
        bsc_b = brier_score(po_b, yb)
        bsf_b = brier_score(pn_b, yb)
        prev_b = float(yb.mean())
        bsref_b = prev_b * (1 - prev_b)
        bss_d = (1 - bsf_b / bsref_b) - (1 - bsc_b / bsref_b) \
                  if bsref_b > 0 else float("nan")
        if not np.isnan(n_c):
            nri_c_bs.append(n_c)
        if not np.isnan(n_cat):
            nri_cat_bs.append(n_cat)
        if not np.isnan(ii):
            idi_bs.append(ii)
        if not np.isnan(bss_d):
            delta_bss_bs.append(bss_d)
        delta_brier_bs.append(bsf_b - bsc_b)
    nri_c_bs = np.array(nri_c_bs)
    nri_cat_bs = np.array(nri_cat_bs)
    idi_bs = np.array(idi_bs)
    delta_bss_bs = np.array(delta_bss_bs)
    delta_brier_bs = np.array(delta_brier_bs)

    return {
        "n": n, "n_pos": n_pos, "n_neg": n_neg,
        "auc_clin": auc_c, "auc_full": auc_f,
        "delta_AUC": auc_f - auc_c,
        "continuous_NRI": {
            "point": nri_c, "nri_pos": nri_pos_c,
            "nri_neg": nri_neg_c,
            "bootstrap_mean": float(nri_c_bs.mean()),
            "95_CI": [float(np.percentile(nri_c_bs, 2.5)),
                       float(np.percentile(nri_c_bs, 97.5))],
            "p_one_sided": float((nri_c_bs <= 0).mean()),
        },
        "categorical_NRI": {
            "thresholds": [0.25, 0.5, 0.75],
            "point": nri_cat, "nri_pos": nri_pos_cat,
            "nri_neg": nri_neg_cat,
            "bootstrap_mean": float(nri_cat_bs.mean()),
            "95_CI": [float(np.percentile(nri_cat_bs, 2.5)),
                       float(np.percentile(nri_cat_bs, 97.5))],
            "p_one_sided": float((nri_cat_bs <= 0).mean()),
        },
        "IDI": {
            "point": idi_v,
            "bootstrap_mean": float(idi_bs.mean()),
            "95_CI": [float(np.percentile(idi_bs, 2.5)),
                       float(np.percentile(idi_bs, 97.5))],
            "p_one_sided": float((idi_bs <= 0).mean()),
        },
        "Brier": {
            "clin": bs_c, "full": bs_f,
            "delta_brier": bs_f - bs_c,
            "delta_brier_mean": float(delta_brier_bs.mean()),
            "delta_brier_95_CI": [
                float(np.percentile(delta_brier_bs, 2.5)),
                float(np.percentile(delta_brier_bs, 97.5))],
            "ref_uncertainty": bs_ref,
        },
        "BSS": {
            "clin": bss_c, "full": bss_f,
            "delta_bss": bss_f - bss_c,
            "delta_bss_bootstrap_mean": float(delta_bss_bs.mean()),
            "delta_bss_95_CI": [
                float(np.percentile(delta_bss_bs, 2.5)),
                float(np.percentile(delta_bss_bs, 97.5))],
            "p_one_sided": float((delta_bss_bs <= 0).mean()),
        },
        "Brier_decomp_clin": decomp_c,
        "Brier_decomp_full": decomp_f,
    }


def main():
    print("=" * 78, flush=True)
    print("v212 NRI + IDI + BRIER (round 44 CPU)", flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    mu_rows = load_mu()
    rhuh_rows = load_rhuh()
    mu_lab = label_binary(mu_rows, HORIZON)
    rhuh_lab = label_binary(rhuh_rows, HORIZON)
    print(f"  MU labelled: n={len(mu_lab)} "
          f"({sum(1 for l in mu_lab if l['y']==1)} pos)",
          flush=True)
    print(f"  RHUH labelled: n={len(rhuh_lab)} "
          f"({sum(1 for l in rhuh_lab if l['y']==1)} pos)",
          flush=True)

    # MU 4-feature analysis (round 39 v202 baseline)
    feats_clin_mu = ["age", "idh1", "mgmt"]
    feats_full_mu = feats_clin_mu + ["v_kernel_s3"]

    print(f"\n=== MU 4-feature reclassification ===", flush=True)
    mu_result = reclassification_analysis(
        mu_lab, feats_clin_mu, feats_full_mu, rng)
    if mu_result is None:
        raise RuntimeError("MU analysis failed")
    nri = mu_result["continuous_NRI"]
    nri_cat = mu_result["categorical_NRI"]
    idi_d = mu_result["IDI"]
    bs = mu_result["Brier"]
    bss = mu_result["BSS"]
    print(f"  AUC_clin={mu_result['auc_clin']:.4f}, "
          f"AUC_full={mu_result['auc_full']:.4f}, "
          f"Delta={mu_result['delta_AUC']:+.4f}", flush=True)
    print(f"  Continuous NRI = {nri['point']:+.4f}  "
          f"(NRI_pos={nri['nri_pos']:+.4f}, "
          f"NRI_neg={nri['nri_neg']:+.4f}); "
          f"95% CI [{nri['95_CI'][0]:+.4f}, "
          f"{nri['95_CI'][1]:+.4f}], P<=0={nri['p_one_sided']:.4f}",
          flush=True)
    print(f"  Categorical NRI [0.25, 0.5, 0.75] = "
          f"{nri_cat['point']:+.4f}; "
          f"95% CI [{nri_cat['95_CI'][0]:+.4f}, "
          f"{nri_cat['95_CI'][1]:+.4f}], "
          f"P<=0={nri_cat['p_one_sided']:.4f}",
          flush=True)
    print(f"  IDI = {idi_d['point']:+.4f}; 95% CI "
          f"[{idi_d['95_CI'][0]:+.4f}, {idi_d['95_CI'][1]:+.4f}], "
          f"P<=0={idi_d['p_one_sided']:.4f}", flush=True)
    print(f"  Brier: clin={bs['clin']:.4f}, full={bs['full']:.4f}, "
          f"Delta={bs['delta_brier']:+.4f}, 95% CI "
          f"[{bs['delta_brier_95_CI'][0]:+.4f}, "
          f"{bs['delta_brier_95_CI'][1]:+.4f}]", flush=True)
    print(f"  BSS: clin={bss['clin']:.4f}, full={bss['full']:.4f}, "
          f"Delta_BSS={bss['delta_bss']:+.4f}, 95% CI "
          f"[{bss['delta_bss_95_CI'][0]:+.4f}, "
          f"{bss['delta_bss_95_CI'][1]:+.4f}], "
          f"P<=0={bss['p_one_sided']:.4f}", flush=True)
    print(f"  Brier decomp (clin): rel={mu_result['Brier_decomp_clin']['reliability']:.4f}, "
          f"res={mu_result['Brier_decomp_clin']['resolution']:.4f}, "
          f"unc={mu_result['Brier_decomp_clin']['uncertainty']:.4f}",
          flush=True)
    print(f"  Brier decomp (full): rel={mu_result['Brier_decomp_full']['reliability']:.4f}, "
          f"res={mu_result['Brier_decomp_full']['resolution']:.4f}, "
          f"unc={mu_result['Brier_decomp_full']['uncertainty']:.4f}",
          flush=True)

    # Pooled MU+RHUH 3-feature analysis
    feats_clin_pool = ["age", "idh1"]
    feats_full_pool = feats_clin_pool + ["v_kernel_s3"]
    pooled = mu_lab + rhuh_lab
    print(f"\n=== Pooled MU+RHUH 3-feature reclassification ===",
          flush=True)
    pooled_result = reclassification_analysis(
        pooled, feats_clin_pool, feats_full_pool, rng)
    if pooled_result is not None:
        nri_p = pooled_result["continuous_NRI"]
        idi_p = pooled_result["IDI"]
        bss_p = pooled_result["BSS"]
        print(f"  AUC_clin={pooled_result['auc_clin']:.4f}, "
              f"AUC_full={pooled_result['auc_full']:.4f}, "
              f"Delta={pooled_result['delta_AUC']:+.4f}",
              flush=True)
        print(f"  Continuous NRI = {nri_p['point']:+.4f}, 95% CI "
              f"[{nri_p['95_CI'][0]:+.4f}, "
              f"{nri_p['95_CI'][1]:+.4f}], "
              f"P<=0={nri_p['p_one_sided']:.4f}", flush=True)
        print(f"  IDI = {idi_p['point']:+.4f}, 95% CI "
              f"[{idi_p['95_CI'][0]:+.4f}, "
              f"{idi_p['95_CI'][1]:+.4f}], "
              f"P<=0={idi_p['p_one_sided']:.4f}", flush=True)
        print(f"  Delta BSS = {bss_p['delta_bss']:+.4f}, 95% CI "
              f"[{bss_p['delta_bss_95_CI'][0]:+.4f}, "
              f"{bss_p['delta_bss_95_CI'][1]:+.4f}]", flush=True)

    # MU 3-feature variant for direct comparability with pooled
    print(f"\n=== MU 3-feature reclassification "
          f"(comparable to pooled) ===", flush=True)
    mu_3f_result = reclassification_analysis(
        mu_lab, feats_clin_pool, feats_full_pool, rng)
    if mu_3f_result is not None:
        nri_3f = mu_3f_result["continuous_NRI"]
        idi_3f = mu_3f_result["IDI"]
        bss_3f = mu_3f_result["BSS"]
        print(f"  AUC_clin={mu_3f_result['auc_clin']:.4f}, "
              f"AUC_full={mu_3f_result['auc_full']:.4f}, "
              f"Delta={mu_3f_result['delta_AUC']:+.4f}",
              flush=True)
        print(f"  Continuous NRI = {nri_3f['point']:+.4f}, 95% CI "
              f"[{nri_3f['95_CI'][0]:+.4f}, "
              f"{nri_3f['95_CI'][1]:+.4f}], "
              f"P<=0={nri_3f['p_one_sided']:.4f}", flush=True)

    out = {
        "version": "v212",
        "experiment": ("NRI + IDI + Brier-score reclassification "
                       "statistics for binary 365-d PFS"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "n_bootstraps": N_BOOTSTRAPS,
        "MU_4feature": mu_result,
        "MU_3feature": mu_3f_result,
        "pooled_3feature": pooled_result,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
