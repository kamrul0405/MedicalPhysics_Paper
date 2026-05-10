"""v216: V_kernel robustness to mask perturbations — round 46 (CPU).

Clinical-deployment-grade robustness analysis. Three perturbation
types simulate realistic segmentation noise:

  1. MORPHOLOGICAL (erosion/dilation): captures segmentation drift
     — a real clinician's mask can differ from the gold standard
     by a few voxels of boundary expansion or contraction.
  2. VOXEL-FLIP NOISE: captures inter-rater disagreement on
     boundary voxels (~5-10% disagreement is typical).
  3. PARTIAL-VOLUME BLUR: captures effects of MRI resolution
     limits (Gaussian blur on the binary mask).

For each perturbation type and magnitude:
  - Apply perturbation to MU baseline masks
  - Recompute V_kernel (sigma=3) on perturbed mask
  - Refit logistic clin+V_kernel on perturbed feature
  - Report Delta AUC and Delta NRI vs unperturbed (which gives
    AUC=0.728, NRI=+0.43 from rounds 39 v202, 44 v212)

If the kernel signal survives moderate perturbations -> deployment-
ready. If it collapses with small perturbations -> not robust.

Outputs:
  Nature_project/05_results/v216_robustness_perturbations.json
"""
from __future__ import annotations

import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import (binary_dilation, binary_erosion,
                              gaussian_filter)
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v216_robustness_perturbations.json"

SIGMA_KERNEL = 3.0
HORIZON = 365
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X(rows, feats):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0)
    stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return np.column_stack([np.ones(len(arr_z)), arr_z])


def continuous_nri(p_old, p_new, y):
    y = np.asarray(y).astype(int)
    p_old = np.asarray(p_old)
    p_new = np.asarray(p_new)
    pos = y == 1
    neg = y == 0
    if pos.sum() == 0 or neg.sum() == 0:
        return float("nan")
    nri_pos = ((p_new[pos] > p_old[pos]).mean()
                - (p_new[pos] < p_old[pos]).mean())
    nri_neg = ((p_new[neg] < p_old[neg]).mean()
                - (p_new[neg] > p_old[neg]).mean())
    return float(nri_pos + nri_neg)


def perturb_morphological(mask, k):
    """k > 0: dilation; k < 0: erosion; k = 0: identity."""
    if k == 0:
        return mask.copy()
    bool_m = mask.astype(bool)
    if k > 0:
        for _ in range(k):
            bool_m = binary_dilation(bool_m)
    else:
        for _ in range(-k):
            bool_m = binary_erosion(bool_m)
    return bool_m.astype(np.float32)


def perturb_voxelflip(mask, p, rng):
    """Random voxel flip with probability p (only flip BOUNDARY-
    adjacent voxels to be realistic)."""
    if p == 0:
        return mask.copy()
    bool_m = mask.astype(bool)
    # Generate random flips only in a 'boundary band' (within 2
    # voxels of the boundary) — more realistic than flipping
    # arbitrary voxels far from the tumour.
    dilated = binary_dilation(bool_m, iterations=2)
    eroded = binary_erosion(bool_m, iterations=2) if bool_m.sum(
        ) > 50 else bool_m
    boundary_band = dilated & ~eroded
    flip_mask = rng.random(mask.shape) < p
    flip_mask = flip_mask & boundary_band
    perturbed = bool_m.copy()
    perturbed[flip_mask] = ~perturbed[flip_mask]
    return perturbed.astype(np.float32)


def perturb_partial_volume(mask, sigma_pv):
    """Apply Gaussian blur and threshold at 0.5 — mimics partial-
    volume effects."""
    if sigma_pv == 0:
        return mask.copy()
    blurred = gaussian_filter(mask.astype(np.float32),
                                sigma=sigma_pv)
    return (blurred >= 0.5).astype(np.float32)


def main():
    print("=" * 78, flush=True)
    print("v216 ROBUSTNESS TO MASK PERTURBATIONS (round 46 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # Load MU clinical
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }

    # Load masks
    print("\nLoading masks + binary 365-d labels...", flush=True)
    samples = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None or c["progress"] is None
                or c["pfs_days"] is None):
            continue
        mask = (np.load(f) > 0).astype(np.float32)
        if mask.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = int(c["progress"])
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        samples.append({
            "pid": pid, "mask": mask, "y": int(y),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
        })
    n_samples = len(samples)
    n_pos = sum(1 for s in samples if s["y"] == 1)
    n_neg = n_samples - n_pos
    print(f"  {n_samples} MU patients (pos={n_pos}, "
          f"neg={n_neg})", flush=True)

    feats_clin = ["age", "idh1", "mgmt"]
    feats_full = feats_clin + ["v_kernel_s3"]

    def evaluate_with_kernel_values(rows_with_vk):
        """Given rows with v_kernel_s3 set, fit logistic and
        return AUC, NRI, p_old, p_new."""
        X_c = build_X(rows_with_vk, feats_clin)
        X_f = build_X(rows_with_vk, feats_full)
        y = np.array([r["y"] for r in rows_with_vk],
                      dtype=float)
        bc = logistic_fit(X_c, y)
        bf = logistic_fit(X_f, y)
        p_old = sigmoid(X_c @ bc)
        p_new = sigmoid(X_f @ bf)
        auc_c = auroc(p_old, y)
        auc_f = auroc(p_new, y)
        nri = continuous_nri(p_old, p_new, y)
        return {
            "auc_clin": float(auc_c),
            "auc_full": float(auc_f),
            "delta_AUC": float(auc_f - auc_c),
            "NRI": float(nri),
        }

    # Baseline (unperturbed)
    print("\n=== BASELINE (unperturbed) ===", flush=True)
    for s in samples:
        s["v_kernel_s3"] = kernel_outgrowth_volume(
            s["mask"], SIGMA_KERNEL)
    baseline = evaluate_with_kernel_values(samples)
    print(f"  AUC_clin={baseline['auc_clin']:.4f}, "
          f"AUC_full={baseline['auc_full']:.4f}, "
          f"Delta AUC={baseline['delta_AUC']:+.4f}, "
          f"NRI={baseline['NRI']:+.4f}", flush=True)

    # ---- Sweep 1: morphological erosion/dilation ----
    print("\n=== PERTURBATION 1: morphological "
          "(erosion/dilation) ===", flush=True)
    morph_results = {}
    for k in [-3, -2, -1, 0, 1, 2, 3]:
        for s in samples:
            mp = perturb_morphological(s["mask"], k)
            s["v_kernel_s3"] = kernel_outgrowth_volume(
                mp, SIGMA_KERNEL)
        r = evaluate_with_kernel_values(samples)
        print(f"  k={k:+d} voxels: AUC_full={r['auc_full']:.4f}, "
              f"Delta AUC={r['delta_AUC']:+.4f}, "
              f"NRI={r['NRI']:+.4f}", flush=True)
        morph_results[str(k)] = r

    # ---- Sweep 2: voxel-flip noise ----
    print("\n=== PERTURBATION 2: voxel-flip noise (boundary "
          "band) ===", flush=True)
    flip_results = {}
    for p in [0.0, 0.05, 0.10, 0.20, 0.30, 0.50]:
        for s in samples:
            mp = perturb_voxelflip(s["mask"], p, rng)
            s["v_kernel_s3"] = kernel_outgrowth_volume(
                mp, SIGMA_KERNEL)
        r = evaluate_with_kernel_values(samples)
        print(f"  p={p:.2f}: AUC_full={r['auc_full']:.4f}, "
              f"Delta AUC={r['delta_AUC']:+.4f}, "
              f"NRI={r['NRI']:+.4f}", flush=True)
        flip_results[f"{p:.2f}"] = r

    # ---- Sweep 3: partial-volume blur ----
    print("\n=== PERTURBATION 3: partial-volume blur ===",
          flush=True)
    pv_results = {}
    for sg in [0.0, 0.5, 1.0, 1.5, 2.0, 3.0]:
        for s in samples:
            mp = perturb_partial_volume(s["mask"], sg)
            s["v_kernel_s3"] = kernel_outgrowth_volume(
                mp, SIGMA_KERNEL)
        r = evaluate_with_kernel_values(samples)
        print(f"  sigma_pv={sg:.1f}: "
              f"AUC_full={r['auc_full']:.4f}, "
              f"Delta AUC={r['delta_AUC']:+.4f}, "
              f"NRI={r['NRI']:+.4f}", flush=True)
        pv_results[f"{sg:.1f}"] = r

    out = {
        "version": "v216",
        "experiment": ("V_kernel robustness to 3 mask "
                       "perturbation types: morphological, "
                       "voxel-flip noise, partial-volume blur"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "n_samples": n_samples,
        "n_pos": n_pos,
        "n_neg": n_neg,
        "baseline": baseline,
        "perturbation_morphological": morph_results,
        "perturbation_voxelflip": flip_results,
        "perturbation_partial_volume": pv_results,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
