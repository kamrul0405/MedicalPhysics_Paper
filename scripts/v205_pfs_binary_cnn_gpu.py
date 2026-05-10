"""v205: End-to-end 3D CNN binary 365-day-PFS classifier with
mask-only vs mask+kernel ablation — round 40 (GPU).

Motivated by round 39 v202 (V_kernel adds +10.8 pp AUC for binary
365-day PFS via simple logistic) and v203 (multi-task DL fails for
continuous Cox survival). Open question: can a deep 3D CNN trained
DIRECTLY on the binary PFS task (rather than continuous Cox) match
or beat the simple logistic? And — critically — does the deep model
need the kernel input, or can it discover the kernel-equivalent
features from the mask alone?

Method: 5-fold stratified CV on MU-Glioma-Post (n with PFS+mask).
3D CNN encoder + global pool + MLP -> binary logit. Two variants:
  A) mask-only input (1 channel)
  B) mask + bimodal kernel sigma=3 (2 channels)
BCE loss with positive-class weighting.

Compare to v202 logistic baselines (clinical-only AUC=0.620;
clinical+V_kernel AUC=0.728).

Outputs:
  Nature_project/05_results/v205_pfs_binary_cnn.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v205_pfs_binary_cnn.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
EPOCHS = 40
LR = 5e-4
SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors, order=0
                    ).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class BinaryPFSCNN(nn.Module):
    """3D CNN encoder + global pool + MLP -> 1 logit."""
    def __init__(self, in_ch=1, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(32, 1),
        )

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.head(self.global_pool(e3)).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def main():
    print("=" * 78, flush=True)
    print(f"v205 BINARY-PFS 3D CNN (round 40 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    # ---- Load clinical PFS ----
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    age_col = header.index("Age at diagnosis")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "progress": prog, "pfs_days": pfs_d,
            "age": age, "idh1": idh1, "mgmt": mgmt,
        }
    print(f"  Loaded clinical for {len(clinical)} MU patients",
          flush=True)

    # ---- Load masks + build kernel + binary 365-d label ----
    print("\nLoading masks + computing kernels + binary 365-day "
          "labels...", flush=True)
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["progress"] is None or c["pfs_days"] is None
                or c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None):
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        # Binary 365-day label
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue  # censored before HORIZON
        # Resize + kernel
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({
            "pid": pid, "mask": m, "kernel": k, "y": int(y),
            "age": c["age"], "idh1": c["idh1"], "mgmt": c["mgmt"],
        })
    n = len(data)
    y_all = np.array([d["y"] for d in data], dtype=np.float32)
    n_pos, n_neg = int(y_all.sum()), int(n - y_all.sum())
    print(f"  {n} patients, n_pos={n_pos}, n_neg={n_neg}", flush=True)

    folds = stratified_kfold(y_all, N_FOLDS, seed=SEED)
    print(f"  Stratified {N_FOLDS}-fold sizes: "
          f"{[len(f) for f in folds]}", flush=True)

    # ---- Train each variant ----
    variant_results = {}
    for variant_name, in_channels in [("mask_only", 1),
                                       ("mask_plus_kernel", 2)]:
        print(f"\n=== VARIANT: {variant_name} (in_ch={in_channels}) "
              f"===", flush=True)
        oof_scores = np.zeros(n, dtype=np.float32)
        oof_assigned = np.zeros(n, dtype=bool)
        fold_aucs = []
        for fold_i, test_idx in enumerate(folds):
            train_idx = np.array(
                sorted(set(range(n)) - set(test_idx.tolist())))
            n_train_pos = int(y_all[train_idx].sum())
            n_train_neg = int(len(train_idx) - n_train_pos)
            if in_channels == 1:
                Xtr = np.stack(
                    [data[i]["mask"][None, :, :, :]
                     for i in train_idx]).astype(np.float32)
                Xte = np.stack(
                    [data[i]["mask"][None, :, :, :]
                     for i in test_idx]).astype(np.float32)
            else:
                Xtr = np.stack(
                    [np.stack([data[i]["mask"],
                               data[i]["kernel"]], axis=0)
                     for i in train_idx]).astype(np.float32)
                Xte = np.stack(
                    [np.stack([data[i]["mask"],
                               data[i]["kernel"]], axis=0)
                     for i in test_idx]).astype(np.float32)
            ytr = y_all[train_idx]
            yte = y_all[test_idx]
            Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
            Xte_t = torch.from_numpy(Xte).to(DEVICE)
            ytr_t = torch.from_numpy(ytr).to(DEVICE)

            torch.manual_seed(SEED + fold_i)
            np.random.seed(SEED + fold_i)
            model = BinaryPFSCNN(in_ch=in_channels, base=24).to(DEVICE)
            opt = torch.optim.AdamW(model.parameters(), lr=LR,
                                     weight_decay=1e-3)
            pos_weight = torch.tensor([n_train_neg / max(1, n_train_pos)],
                                       dtype=torch.float32, device=DEVICE)
            criterion = nn.BCEWithLogitsLoss(pos_weight=pos_weight)

            n_tr = len(train_idx)
            bs = 4
            for ep in range(EPOCHS):
                model.train()
                perm = np.random.permutation(n_tr)
                ep_loss = 0.0
                for i in range(0, n_tr, bs):
                    idx = perm[i:i+bs]
                    xb = Xtr_t[idx]
                    yb = ytr_t[idx]
                    logits = model(xb)
                    loss = criterion(logits, yb)
                    opt.zero_grad()
                    loss.backward()
                    opt.step()
                    ep_loss += loss.item() * len(idx)
                if (ep + 1) % 10 == 0:
                    model.eval()
                    with torch.no_grad():
                        scr_te = model(Xte_t).cpu().numpy()
                    auc_te = auroc(scr_te, yte)
                    print(f"    fold {fold_i+1}/{N_FOLDS} ep {ep+1}/"
                          f"{EPOCHS}  loss={ep_loss/n_tr:.4f}  "
                          f"test_AUC={auc_te:.4f}", flush=True)

            model.eval()
            with torch.no_grad():
                scr_te = model(Xte_t).cpu().numpy()
            oof_scores[test_idx] = scr_te
            oof_assigned[test_idx] = True
            auc_te = auroc(scr_te, yte)
            fold_aucs.append(float(auc_te))
            print(f"  Fold {fold_i+1}/{N_FOLDS} final test AUC = "
                  f"{auc_te:.4f}  "
                  f"(n_train={len(train_idx)}, n_test={len(test_idx)})",
                  flush=True)

            del model, opt, Xtr_t, Xte_t, ytr_t
            gc.collect()
            if DEVICE.type == "cuda":
                torch.cuda.empty_cache()

        # Pooled OOF AUC
        pooled_auc = auroc(oof_scores[oof_assigned], y_all[oof_assigned])
        print(f"\n  {variant_name} POOLED OOF AUC = {pooled_auc:.4f}  "
              f"(per-fold mean = {np.mean(fold_aucs):.4f})", flush=True)
        variant_results[variant_name] = {
            "in_channels": in_channels,
            "fold_aucs": fold_aucs,
            "fold_auc_mean": float(np.mean(fold_aucs)),
            "fold_auc_std": float(np.std(fold_aucs)),
            "pooled_oof_auc": float(pooled_auc),
            "n_folds": N_FOLDS,
        }

    # ---- Summary table ----
    print("\n" + "=" * 78, flush=True)
    print("SUMMARY — v205 binary-PFS 3D CNN vs v202 logistic baselines",
          flush=True)
    print("=" * 78, flush=True)
    print(f"  v202 logistic clinical-only             AUC = 0.6199",
          flush=True)
    print(f"  v202 logistic clinical + V_kernel       AUC = 0.7283",
          flush=True)
    print(f"  v205 3D CNN mask-only          OOF AUC = "
          f"{variant_results['mask_only']['pooled_oof_auc']:.4f}",
          flush=True)
    print(f"  v205 3D CNN mask + kernel(3)   OOF AUC = "
          f"{variant_results['mask_plus_kernel']['pooled_oof_auc']:.4f}",
          flush=True)

    out = {
        "version": "v205",
        "experiment": ("End-to-end 3D CNN binary 365-day-PFS classifier "
                       "with mask-only vs mask+kernel ablation"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "epochs_per_fold": EPOCHS,
        "n_total": n,
        "n_pos": n_pos,
        "n_neg": n_neg,
        "variants": variant_results,
        "v202_baselines": {
            "logistic_clinical_only_AUC": 0.6199,
            "logistic_clinical_plus_Vkernel_AUC": 0.7283,
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
