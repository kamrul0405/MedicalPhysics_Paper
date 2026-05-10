"""v214: Unification of binary-PFS screen with continuous Cox PH —
does the binary-classifier probability score rescue Cox prediction?
— round 45 (CPU).

The metric-mismatch story: rounds 32-38 produced 5 honest negatives
on continuous Cox survival prediction (kernel HR p=0.92, p=0.53,
p=0.25, C=0.45, C=0.46). Round 39 v202 reframed PFS as binary
classification at 365 d -> +0.108 AUC lift, leading to rounds 40-44
of confirmation, meta-analysis, transfer learning.

Open question for unification: if we compute the LOGISTIC-DERIVED
365-day risk score for each patient, and use that single number as
a Cox PH covariate, does it rescue the continuous Cox model? This
would unify the binary-success story with continuous time-to-event
prediction.

Method on MU-Glioma-Post n=130 with continuous PFS:
  1. Fit logistic clin+V_kernel on binary 365-d PFS as before -> per-
     patient predicted probability p_hat (the screening risk score).
  2. Cox PH on continuous PFS_days (with right-censoring) using p_hat
     as a single covariate. Compare against:
     (a) Cox clin-only (age + IDH + MGMT)
     (b) Cox with raw V_kernel
     (c) Cox with raw V_kernel + clinical
     (d) Cox with p_hat + clinical
  3. Compare Harrell's C-index, log-likelihood, BIC.
  4. Restricted cubic spline on V_kernel (3 knots) — does the
     non-linear shape matter?
  5. Likelihood-ratio test for adding p_hat to clinical.
  6. Bootstrap CI on Delta C-index (clin+p_hat) - clin.

Outputs:
  Nature_project/05_results/v214_binary_score_cox.json
"""
from __future__ import annotations

import csv
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize
from scipy.stats import chi2

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v214_binary_score_cox.json"

SIGMA_KERNEL = 3.0
HORIZON = 365
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def build_X(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def cox_partial_loglik(beta, X, time, event, l2=1e-3):
    """Negative partial log-likelihood with L2 regularization."""
    n = len(time)
    eta = X @ beta
    eta_max = eta.max()
    exp_eta_shift = np.exp(eta - eta_max)
    # For each event-i, sum over risk set: log sum_{j in R_i}(exp(eta_j))
    order = np.argsort(-time)  # descending so cumsum gives R_i
    eta_o = eta[order]
    eta_max_o = eta_o.max()
    es_o = np.exp(eta_o - eta_max_o)
    cum_es = np.cumsum(es_o)
    log_cum = eta_max_o + np.log(cum_es)
    # Map back to original order
    log_cum_orig = np.empty(n)
    log_cum_orig[order] = log_cum
    valid = event == 1
    log_pl = (eta[valid] - log_cum_orig[valid]).sum()
    return -log_pl + l2 * np.sum(beta * beta)


def cox_fit(X_no_intercept, time, event):
    """Fit Cox PH (no intercept). X is (n, p); time, event are
    (n,)."""
    p = X_no_intercept.shape[1]
    res = minimize(cox_partial_loglik, np.zeros(p),
                    args=(X_no_intercept, time, event),
                    method="L-BFGS-B")
    return res.x, -res.fun  # beta and partial log-likelihood
                              # (positive value)


def harrells_c(time, event, risk_score):
    """Harrell's C-index."""
    time = np.asarray(time)
    event = np.asarray(event)
    risk = np.asarray(risk_score)
    n = len(time)
    n_pairs = 0
    n_concordant = 0
    n_ties = 0
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            if event[i] != 1:
                continue
            if time[j] <= time[i] and event[j] != 1:
                continue
            if time[i] >= time[j]:
                continue
            n_pairs += 1
            if risk[i] > risk[j]:
                n_concordant += 1
            elif risk[i] == risk[j]:
                n_ties += 1
    if n_pairs == 0:
        return float("nan")
    return (n_concordant + 0.5 * n_ties) / n_pairs


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def restricted_cubic_spline_basis(x, knots):
    """3-knot RCS basis (returns 1 column for 3 knots; just non-
    linear term beyond linear)."""
    x = np.asarray(x, dtype=float)
    k = sorted(knots)
    k1, k2, k3 = k[0], k[1], k[2]
    def t(u):
        return np.maximum(0, u) ** 3
    denom = (k3 - k1)
    return ((t(x - k1) - t(x - k2) * (k3 - k1) / (k3 - k2)
              + t(x - k3) * (k2 - k1) / (k3 - k2)) / denom ** 2)


def main():
    print("=" * 78, flush=True)
    print("v214 BINARY-SCORE-IN-COX UNIFICATION (round 45 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # ---- Load MU ----
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        c = clinical[pid]
        if (c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None or c["pfs_days"] is None
                or c["progress"] is None):
            continue
        rows.append({
            "pid": pid,
            "v_kernel_s3": kernel_outgrowth_volume(m,
                                                    SIGMA_KERNEL),
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"], "progress": c["progress"],
        })
    print(f"  Loaded {len(rows)} MU patients with full clinical "
          f"+ PFS", flush=True)

    # Continuous PFS time and event (regardless of horizon)
    time_arr = np.array([r["pfs_days"] for r in rows],
                         dtype=float)
    event_arr = np.array([r["progress"] for r in rows],
                          dtype=float)
    n_events = int(event_arr.sum())
    print(f"  Total events = {n_events}, censored = "
          f"{len(rows) - n_events}", flush=True)
    print(f"  Median PFS = {np.median(time_arr):.0f} d, max = "
          f"{time_arr.max():.0f} d", flush=True)

    # ---- Step 1: Fit binary classifier (clin + V_kernel) at H=365 ----
    print(f"\n=== STEP 1: Binary 365-d PFS logistic ===",
          flush=True)
    bin_lab = []
    for i, r in enumerate(rows):
        pfs = r["pfs_days"]
        prog = int(r["progress"])
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        bin_lab.append({**r, "y": y, "i": i})
    n_bin = len(bin_lab)
    n_bin_pos = sum(1 for l in bin_lab if l["y"] == 1)
    print(f"  Binary labelled n = {n_bin} (pos={n_bin_pos}, "
          f"neg={n_bin - n_bin_pos})", flush=True)

    feats_clin = ["age", "idh1", "mgmt"]
    feats_full = feats_clin + ["v_kernel_s3"]
    y_bin = np.array([l["y"] for l in bin_lab], dtype=float)
    X_bin_full, m_full, s_full = build_X(bin_lab, feats_full)
    beta_full = logistic_fit(X_bin_full, y_bin)
    print(f"  Logistic clin+V_kernel beta = {beta_full}",
          flush=True)
    auc_bin = auroc(X_bin_full @ beta_full, y_bin)
    print(f"  Binary classifier AUC (in-sample) = "
          f"{auc_bin:.4f}", flush=True)

    # ---- Step 2: compute predicted probability p_hat for ALL
    # MU patients (not just labelled) using the trained logistic ----
    X_all_full, _, _ = build_X(rows, feats_full,
                                 means=m_full, stds=s_full)
    p_hat_all = sigmoid(X_all_full @ beta_full)
    # Add to rows
    for r, p in zip(rows, p_hat_all):
        r["p_hat_365"] = float(p)
    print(f"  p_hat range: [{p_hat_all.min():.3f}, "
          f"{p_hat_all.max():.3f}], mean = "
          f"{p_hat_all.mean():.3f}", flush=True)

    # ---- Step 3: fit Cox PH models ----
    print(f"\n=== STEP 3: Cox PH models on continuous PFS ===",
          flush=True)
    # Standardize all features for Cox
    def build_cox_X(rows, feats):
        arr = np.array([[r[f] for f in feats] for r in rows],
                        dtype=float)
        means = arr.mean(axis=0)
        stds = arr.std(axis=0)
        stds[stds == 0] = 1
        return (arr - means) / stds

    cox_results = {}
    n_total = len(rows)
    for name, feats in [
        ("clin_only", ["age", "idh1", "mgmt"]),
        ("Vkernel_only", ["v_kernel_s3"]),
        ("Vkernel_plus_clin",
         ["age", "idh1", "mgmt", "v_kernel_s3"]),
        ("phat_only", ["p_hat_365"]),
        ("phat_plus_clin",
         ["age", "idh1", "mgmt", "p_hat_365"]),
    ]:
        X_cox = build_cox_X(rows, feats)
        beta, pll = cox_fit(X_cox, time_arr, event_arr)
        risk = X_cox @ beta
        c_idx = harrells_c(time_arr, event_arr, risk)
        # BIC = -2*pll + p*log(n_events)
        bic = -2 * pll + len(feats) * np.log(max(1, n_events))
        print(f"  {name:25s}  C={c_idx:.4f}  pll={pll:+.4f}  "
              f"BIC={bic:.2f}  n_feats={len(feats)}", flush=True)
        cox_results[name] = {
            "features": feats,
            "n_features": len(feats),
            "beta": beta.tolist(),
            "partial_loglik": float(pll),
            "BIC": float(bic),
            "C_index": float(c_idx),
        }

    # Likelihood-ratio test: clin vs clin + V_kernel
    pll_clin = cox_results["clin_only"]["partial_loglik"]
    pll_vk = cox_results["Vkernel_plus_clin"]["partial_loglik"]
    lr_vk = 2 * (pll_vk - pll_clin)
    df_vk = (cox_results["Vkernel_plus_clin"]["n_features"]
             - cox_results["clin_only"]["n_features"])
    p_vk = 1 - chi2.cdf(lr_vk, df_vk) if lr_vk > 0 else 1.0
    print(f"\n  LRT clin vs clin+V_kernel: LR={lr_vk:.4f}, "
          f"df={df_vk}, P={p_vk:.4f}", flush=True)

    pll_phat = cox_results["phat_plus_clin"]["partial_loglik"]
    lr_phat = 2 * (pll_phat - pll_clin)
    df_phat = (cox_results["phat_plus_clin"]["n_features"]
                - cox_results["clin_only"]["n_features"])
    p_phat = (1 - chi2.cdf(lr_phat, df_phat)
              if lr_phat > 0 else 1.0)
    print(f"  LRT clin vs clin+p_hat: LR={lr_phat:.4f}, "
          f"df={df_phat}, P={p_phat:.4f}", flush=True)

    # ---- Step 4: Restricted cubic spline on V_kernel ----
    print(f"\n=== STEP 4: RCS Cox (3-knot non-linear V_kernel) ===",
          flush=True)
    vk_arr = np.array([r["v_kernel_s3"] for r in rows],
                       dtype=float)
    knots = list(np.percentile(vk_arr, [25, 50, 75]))
    print(f"  3-knot RCS at V_kernel percentiles: 25%={knots[0]:.0f}, "
          f"50%={knots[1]:.0f}, 75%={knots[2]:.0f}", flush=True)
    vk_rcs = restricted_cubic_spline_basis(vk_arr, knots)
    feats_rcs_vals = []
    for r, v_lin, v_rcs in zip(rows, vk_arr, vk_rcs):
        r["vk_lin"] = float(v_lin)
        r["vk_rcs"] = float(v_rcs)
    X_rcs = build_cox_X(
        rows, ["age", "idh1", "mgmt", "vk_lin", "vk_rcs"])
    beta_rcs, pll_rcs = cox_fit(X_rcs, time_arr, event_arr)
    risk_rcs = X_rcs @ beta_rcs
    c_rcs = harrells_c(time_arr, event_arr, risk_rcs)
    print(f"  Cox clin + RCS V_kernel: C={c_rcs:.4f}, "
          f"pll={pll_rcs:.4f}", flush=True)
    lr_rcs = 2 * (pll_rcs - pll_vk)
    p_rcs = (1 - chi2.cdf(lr_rcs, 1)
             if lr_rcs > 0 else 1.0)
    print(f"  LRT linear vs RCS V_kernel: LR={lr_rcs:.4f}, "
          f"df=1, P={p_rcs:.4f}", flush=True)

    # ---- Step 5: bootstrap CI on Delta C-index (clin+phat) -
    # (clin) ----
    print(f"\n=== STEP 5: bootstrap CI on Delta C-index ===",
          flush=True)
    delta_c_phat_clin = []
    delta_c_vk_clin = []
    delta_c_phat_vk_clin = []
    for _ in range(N_BOOTSTRAPS):
        idx = rng.integers(0, n_total, size=n_total)
        rb = [rows[i] for i in idx]
        tb = time_arr[idx]
        eb = event_arr[idx]
        if int(eb.sum()) < 5:
            continue
        # Refit each model
        X_c = build_cox_X(rb, ["age", "idh1", "mgmt"])
        b_c, _ = cox_fit(X_c, tb, eb)
        c_c = harrells_c(tb, eb, X_c @ b_c)
        X_v = build_cox_X(rb,
                            ["age", "idh1", "mgmt", "v_kernel_s3"])
        b_v, _ = cox_fit(X_v, tb, eb)
        c_v = harrells_c(tb, eb, X_v @ b_v)
        X_p = build_cox_X(rb,
                            ["age", "idh1", "mgmt", "p_hat_365"])
        b_p, _ = cox_fit(X_p, tb, eb)
        c_p = harrells_c(tb, eb, X_p @ b_p)
        if not (np.isnan(c_c) or np.isnan(c_v) or np.isnan(c_p)):
            delta_c_phat_clin.append(c_p - c_c)
            delta_c_vk_clin.append(c_v - c_c)
            delta_c_phat_vk_clin.append(c_p - c_v)
    delta_c_phat_clin = np.array(delta_c_phat_clin)
    delta_c_vk_clin = np.array(delta_c_vk_clin)
    delta_c_phat_vk_clin = np.array(delta_c_phat_vk_clin)
    print(f"  Delta C (phat+clin) - clin: mean="
          f"{delta_c_phat_clin.mean():+.4f}, "
          f"95% CI [{np.percentile(delta_c_phat_clin, 2.5):+.4f}, "
          f"{np.percentile(delta_c_phat_clin, 97.5):+.4f}], "
          f"P<=0={float((delta_c_phat_clin <= 0).mean()):.4f}",
          flush=True)
    print(f"  Delta C (Vk+clin) - clin: mean="
          f"{delta_c_vk_clin.mean():+.4f}, "
          f"95% CI [{np.percentile(delta_c_vk_clin, 2.5):+.4f}, "
          f"{np.percentile(delta_c_vk_clin, 97.5):+.4f}]",
          flush=True)
    print(f"  Delta C (phat+clin) - (Vk+clin): mean="
          f"{delta_c_phat_vk_clin.mean():+.4f}, "
          f"95% CI [{np.percentile(delta_c_phat_vk_clin, 2.5):+.4f}, "
          f"{np.percentile(delta_c_phat_vk_clin, 97.5):+.4f}], "
          f"P<=0={float((delta_c_phat_vk_clin <= 0).mean()):.4f}",
          flush=True)

    out = {
        "version": "v214",
        "experiment": ("Binary-classifier-derived risk score in "
                       "continuous Cox PH — unification with rounds "
                       "32-38 negatives"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_for_binary": HORIZON,
        "n_total": n_total,
        "n_events": n_events,
        "n_censored": n_total - n_events,
        "binary_classifier": {
            "n_labelled": n_bin,
            "n_pos": n_bin_pos,
            "n_neg": n_bin - n_bin_pos,
            "auc_in_sample": float(auc_bin),
            "p_hat_min": float(p_hat_all.min()),
            "p_hat_max": float(p_hat_all.max()),
            "p_hat_mean": float(p_hat_all.mean()),
        },
        "cox_models": cox_results,
        "LRT_clin_vs_VkernelPlusClin": {
            "LR": float(lr_vk),
            "df": int(df_vk),
            "P": float(p_vk),
        },
        "LRT_clin_vs_phatPlusClin": {
            "LR": float(lr_phat),
            "df": int(df_phat),
            "P": float(p_phat),
        },
        "RCS_Cox": {
            "knots": [float(k) for k in knots],
            "C_index": float(c_rcs),
            "partial_loglik": float(pll_rcs),
            "LRT_linear_vs_RCS_LR": float(lr_rcs),
            "LRT_linear_vs_RCS_P": float(p_rcs),
        },
        "bootstrap": {
            "delta_C_phat_minus_clin_mean": float(
                delta_c_phat_clin.mean()),
            "delta_C_phat_minus_clin_95_CI": [
                float(np.percentile(delta_c_phat_clin, 2.5)),
                float(np.percentile(delta_c_phat_clin, 97.5))],
            "delta_C_phat_minus_clin_P_le_0": float(
                (delta_c_phat_clin <= 0).mean()),
            "delta_C_Vk_minus_clin_mean": float(
                delta_c_vk_clin.mean()),
            "delta_C_Vk_minus_clin_95_CI": [
                float(np.percentile(delta_c_vk_clin, 2.5)),
                float(np.percentile(delta_c_vk_clin, 97.5))],
            "delta_C_phat_minus_Vk_mean": float(
                delta_c_phat_vk_clin.mean()),
            "delta_C_phat_minus_Vk_95_CI": [
                float(np.percentile(delta_c_phat_vk_clin, 2.5)),
                float(np.percentile(delta_c_phat_vk_clin, 97.5))],
            "delta_C_phat_minus_Vk_P_le_0": float(
                (delta_c_phat_vk_clin <= 0).mean()),
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
