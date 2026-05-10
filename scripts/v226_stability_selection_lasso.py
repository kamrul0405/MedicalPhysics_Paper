"""v226: L1-regularized logistic + STABILITY SELECTION for σ-
feature subset identification — round 51 (CPU).

Round 47-50 used FIXED multi-σ subset {2, 3, 4, 5}. Beyond-NMI
methodological question: which σ values are most STABLE under
bootstrap resampling?

Method:
  1. L1-regularized logistic on extended σ-sweep [1, 2, 3, 4, 5,
     7, 10] (10 features = 3 clinical + 7 σ).
  2. λ-path: sweep regularization strength to identify a sparse
     subset.
  3. STABILITY SELECTION: bootstrap-resample 200 times, fit L1
     logistic, count fraction of bootstraps where each feature
     is selected (non-zero coefficient).
  4. Identify σ features stable in ≥ 80% of bootstraps.
  5. Compare AUC of stable subset to multi-σ {2, 3, 4, 5}.

Outputs:
  Nature_project/05_results/v226_stability_selection.json
"""
from __future__ import annotations

import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v226_stability_selection.json"

HORIZON = 365
N_BOOT = 200
LAMBDAS = [0.001, 0.005, 0.01, 0.025, 0.05, 0.1, 0.25,
            0.5, 1.0, 2.0]
STABILITY_THRESHOLD = 0.80
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def l1_logistic_fit(X, y, lam):
    """Soft-threshold L1-regularized logistic via proximal
    gradient descent. Includes intercept (column 0) which is NOT
    regularized."""
    n, p = X.shape
    beta = np.zeros(p)
    lr = 0.01
    for it in range(2000):
        eta = X @ beta
        prob = 1.0 / (1.0 + np.exp(-np.clip(eta, -50, 50)))
        grad = X.T @ (prob - y) / n
        beta_new = beta - lr * grad
        # Soft-threshold non-intercept components
        for j in range(1, p):
            beta_new[j] = np.sign(beta_new[j]) * max(
                0, abs(beta_new[j]) - lr * lam)
        if np.max(np.abs(beta_new - beta)) < 1e-6:
            break
        beta = beta_new
    return beta


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X(rows, feats, means=None, stds=None):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    if means is None:
        means = np.nanmean(arr, axis=0)
    if stds is None:
        stds = np.nanstd(arr, axis=0)
        stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return (np.column_stack([np.ones(len(arr_z)), arr_z]),
            means, stds)


def label_binary(rows, X):
    out = []
    for r in rows:
        pfs = r["pfs_days"]
        prog = r["progress"]
        if pfs is None or prog is None:
            continue
        if prog == 1 and pfs < X:
            y = 1
        elif (prog == 0 and pfs >= X) or (prog == 1 and pfs >= X):
            y = 0
        else:
            continue
        out.append({**r, "y": y})
    return out


def main():
    print("=" * 78, flush=True)
    print("v226 L1-LASSO + STABILITY SELECTION (round 51 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # Load MU
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }

    # Compute extended-σ features
    print("\nLoading masks + extended σ kernels...",
          flush=True)
    sigma_list = [1.0, 2.0, 3.0, 4.0, 5.0, 7.0, 10.0]
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if any(v is None for v in c.values()):
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        kern_feats = {}
        for sg in sigma_list:
            kern_feats[f"v_kernel_s{int(sg)}"] = float(
                kernel_outgrowth_volume(m, sg))
        rows.append({
            "pid": pid, **kern_feats,
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "pfs_days": c["pfs_days"],
            "progress": c["progress"],
        })
    bin_lab = label_binary(rows, HORIZON)
    n_bin = len(bin_lab)
    n_pos = sum(1 for r in bin_lab if r["y"] == 1)
    print(f"  {n_bin} labelled MU patients (pos={n_pos})",
          flush=True)

    feats_clin = ["age", "idh1", "mgmt"]
    feats_kernel = [f"v_kernel_s{int(sg)}"
                    for sg in sigma_list]
    feats_all = feats_clin + feats_kernel  # 10 features
    feat_names = ["intercept"] + feats_all

    y = np.array([r["y"] for r in bin_lab], dtype=float)

    # ============================================================
    # PART A: L1 path
    # ============================================================
    print("\n=== PART A: L1 regularization path ===",
          flush=True)
    X_full, m, s = build_X(bin_lab, feats_all)
    path_results = {}
    for lam in LAMBDAS:
        beta = l1_logistic_fit(X_full, y, lam)
        n_nonzero = int(np.sum(np.abs(beta[1:]) > 1e-3))
        scores = X_full @ beta
        auc = auroc(scores, y)
        nonzero_feats = [feats_all[j]
                          for j in range(len(feats_all))
                          if abs(beta[j + 1]) > 1e-3]
        print(f"  λ={lam:.4f}  n_nonzero={n_nonzero:2d}  "
              f"AUC={auc:.4f}  selected={nonzero_feats}",
              flush=True)
        path_results[f"{lam:.4f}"] = {
            "lambda": lam,
            "n_nonzero": n_nonzero,
            "auc": float(auc),
            "selected_features": nonzero_feats,
            "beta": [float(b) for b in beta],
        }

    # ============================================================
    # PART B: Stability selection
    # ============================================================
    print(f"\n=== PART B: Stability selection ({N_BOOT} bootstraps "
          f"at λ=0.05) ===", flush=True)
    lam_stable = 0.05
    selection_counts = np.zeros(len(feats_all), dtype=int)
    for b in range(N_BOOT):
        idx = rng.integers(0, n_bin, size=n_bin)
        rb = [bin_lab[i] for i in idx]
        yb = np.array([r["y"] for r in rb], dtype=float)
        if int(yb.sum()) < 5 or int(len(yb) - yb.sum()) < 5:
            continue
        Xb, _, _ = build_X(rb, feats_all)
        beta_b = l1_logistic_fit(Xb, yb, lam_stable)
        for j in range(len(feats_all)):
            if abs(beta_b[j + 1]) > 1e-3:
                selection_counts[j] += 1
    selection_freq = selection_counts / N_BOOT
    print(f"  Selection frequency over {N_BOOT} bootstraps "
          f"(threshold={STABILITY_THRESHOLD:.0%}):",
          flush=True)
    stable_feats = []
    for j, fname in enumerate(feats_all):
        marker = "✓ STABLE" if selection_freq[j] >= \
            STABILITY_THRESHOLD else ""
        print(f"    {fname:18s}  freq={selection_freq[j]:.3f}  "
              f"{marker}", flush=True)
        if selection_freq[j] >= STABILITY_THRESHOLD:
            stable_feats.append(fname)

    # ============================================================
    # PART C: AUC of stable subset vs multi-σ {2,3,4,5}
    # ============================================================
    print(f"\n=== PART C: Stable-subset AUC vs multi-σ ===",
          flush=True)
    feats_multi4 = feats_clin + [
        "v_kernel_s2", "v_kernel_s3", "v_kernel_s4",
        "v_kernel_s5"]
    Xm4, _, _ = build_X(bin_lab, feats_multi4)
    # Ridge logistic baseline
    from scipy.optimize import minimize as mz
    def ridge_logistic(X, y, l2=1e-2):
        n, p = X.shape
        def nll(b):
            eta = X @ b
            log_one = np.where(eta > 0,
                                eta + np.log1p(np.exp(-eta)),
                                np.log1p(np.exp(eta)))
            return -np.sum(y * eta - log_one) + l2 * np.sum(
                b * b)
        return mz(nll, np.zeros(p), method="L-BFGS-B").x

    b_m4 = ridge_logistic(Xm4, y)
    auc_m4 = auroc(Xm4 @ b_m4, y)
    print(f"  Multi-σ {{2,3,4,5}} (round 47): AUC={auc_m4:.4f}",
          flush=True)

    if len(stable_feats) > 0:
        Xs, _, _ = build_X(bin_lab, stable_feats)
        b_s = ridge_logistic(Xs, y)
        auc_stable = auroc(Xs @ b_s, y)
        print(f"  Stable subset {stable_feats}: "
              f"AUC={auc_stable:.4f}",
              flush=True)
    else:
        auc_stable = float("nan")
        print(f"  No features stable at "
              f"{STABILITY_THRESHOLD:.0%} threshold", flush=True)

    # Lower thresholds
    for thr in [0.5, 0.6, 0.7]:
        sel = [feats_all[j] for j in range(len(feats_all))
                if selection_freq[j] >= thr]
        if len(sel) > 0:
            Xs, _, _ = build_X(bin_lab, sel)
            b_s = ridge_logistic(Xs, y)
            a = auroc(Xs @ b_s, y)
            print(f"  Threshold {thr:.0%} subset {sel}: "
                  f"AUC={a:.4f}",
                  flush=True)

    out = {
        "version": "v226",
        "experiment": ("L1 lasso + stability selection for "
                       "σ-feature subset"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "n_total": n_bin,
        "n_pos": n_pos,
        "n_bootstraps": N_BOOT,
        "stability_threshold": STABILITY_THRESHOLD,
        "lambda_stable": lam_stable,
        "lasso_path": path_results,
        "stability_selection": {
            "selection_frequencies": {
                fn: float(f)
                for fn, f in zip(feats_all, selection_freq)
            },
            "stable_features_at_80_pct": stable_feats,
            "lower_threshold_subsets": {
                f"thr_{int(thr * 100)}_pct": [
                    feats_all[j]
                    for j in range(len(feats_all))
                    if selection_freq[j] >= thr
                ]
                for thr in [0.5, 0.6, 0.7, 0.8]
            },
        },
        "auc_comparison": {
            "multi_sigma_2_3_4_5": float(auc_m4),
            "stable_subset_at_80_pct": (
                float(auc_stable)
                if not np.isnan(auc_stable)
                else None),
        },
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
