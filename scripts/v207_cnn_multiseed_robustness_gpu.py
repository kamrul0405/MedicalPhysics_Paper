"""v207: Multi-seed bootstrap of v205 3D CNN ablation — round 41 (GPU).

Round 40 v205 reported pooled OOF AUC = 0.528 (mask-only) and 0.607
(mask+kernel) under a single RNG seed. Per-fold AUCs showed substantial
variance (e.g., fold 3 hit 1.000 by chance). For Nature/Lancet, we
need multi-seed bootstrap of the entire 5-fold CV procedure to verify
the kernel-rescue effect is not seed-dependent.

Method: 5 RNG seeds {42, 123, 999, 31415, 271828} x 2 variants
{mask-only, mask+kernel} x 5-fold stratified CV = 50 model trainings.
Per-(seed, variant): pooled OOF AUC, fold-mean AUC. Report mean +/-
std across the 5 seeds for both quantities.

If the mask+kernel CNN has consistently higher pooled OOF AUC than
mask-only across all 5 seeds (paired comparison), the kernel-rescue
effect is robust.

Outputs:
  Nature_project/05_results/v207_cnn_multiseed_robustness.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v207_cnn_multiseed_robustness.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
EPOCHS = 30  # reduce vs v205 to keep total time manageable
LR = 5e-4
SEEDS = [42, 123, 999, 31415, 271828]


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class BinaryPFSCNN(nn.Module):
    def __init__(self, in_ch=1, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(32, 1),
        )

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.head(self.global_pool(e3)).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def main():
    print("=" * 78, flush=True)
    print(f"v207 MULTI-SEED CNN ROBUSTNESS (round 41 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    # ---- Load clinical PFS ----
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    age_col = header.index("Age at diagnosis")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "progress": prog, "pfs_days": pfs_d,
            "age": age, "idh1": idh1, "mgmt": mgmt,
        }
    print(f"  Loaded clinical for {len(clinical)} MU patients",
          flush=True)

    # ---- Load masks + kernels + binary 365-d labels ----
    print("\nLoading + binary 365-day labels...", flush=True)
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace("_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["progress"] is None or c["pfs_days"] is None
                or c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None):
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    n = len(data)
    y_all = np.array([d["y"] for d in data], dtype=np.float32)
    n_pos = int(y_all.sum())
    n_neg = int(n - n_pos)
    print(f"  {n} patients, n_pos={n_pos}, n_neg={n_neg}", flush=True)

    seed_results = {}
    for seed in SEEDS:
        print(f"\n{'='*70}\nSEED = {seed}\n{'='*70}", flush=True)
        folds = stratified_kfold(y_all, N_FOLDS, seed=seed)
        variant_data = {}
        for variant_name, in_channels in [("mask_only", 1),
                                           ("mask_plus_kernel", 2)]:
            print(f"\n  --- variant {variant_name} (in_ch={in_channels}) "
                  f"---", flush=True)
            oof_scores = np.zeros(n, dtype=np.float32)
            oof_assigned = np.zeros(n, dtype=bool)
            fold_aucs = []
            for fold_i, test_idx in enumerate(folds):
                train_idx = np.array(sorted(
                    set(range(n)) - set(test_idx.tolist())))
                n_tr_pos = int(y_all[train_idx].sum())
                n_tr_neg = int(len(train_idx) - n_tr_pos)
                if in_channels == 1:
                    Xtr = np.stack(
                        [data[i]["mask"][None]
                         for i in train_idx]).astype(np.float32)
                    Xte = np.stack(
                        [data[i]["mask"][None]
                         for i in test_idx]).astype(np.float32)
                else:
                    Xtr = np.stack(
                        [np.stack([data[i]["mask"],
                                   data[i]["kernel"]], axis=0)
                         for i in train_idx]).astype(np.float32)
                    Xte = np.stack(
                        [np.stack([data[i]["mask"],
                                   data[i]["kernel"]], axis=0)
                         for i in test_idx]).astype(np.float32)
                ytr = y_all[train_idx]
                yte = y_all[test_idx]
                Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
                Xte_t = torch.from_numpy(Xte).to(DEVICE)
                ytr_t = torch.from_numpy(ytr).to(DEVICE)

                torch.manual_seed(seed + fold_i * 1000)
                np.random.seed(seed + fold_i * 1000)
                model = BinaryPFSCNN(in_ch=in_channels,
                                       base=24).to(DEVICE)
                opt = torch.optim.AdamW(model.parameters(), lr=LR,
                                         weight_decay=1e-3)
                pos_w = torch.tensor(
                    [n_tr_neg / max(1, n_tr_pos)],
                    dtype=torch.float32, device=DEVICE)
                criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)

                n_tr = len(train_idx)
                bs = 4
                for ep in range(EPOCHS):
                    model.train()
                    perm = np.random.permutation(n_tr)
                    for i in range(0, n_tr, bs):
                        idx = perm[i:i+bs]
                        loss = criterion(model(Xtr_t[idx]),
                                         ytr_t[idx])
                        opt.zero_grad()
                        loss.backward()
                        opt.step()

                model.eval()
                with torch.no_grad():
                    scr = model(Xte_t).cpu().numpy()
                oof_scores[test_idx] = scr
                oof_assigned[test_idx] = True
                a = auroc(scr, yte)
                fold_aucs.append(float(a))
                del model, opt, Xtr_t, Xte_t, ytr_t
                gc.collect()
                if DEVICE.type == "cuda":
                    torch.cuda.empty_cache()
                print(f"    fold {fold_i+1}/{N_FOLDS}: "
                      f"AUC={a:.4f}", flush=True)

            pooled = auroc(oof_scores[oof_assigned],
                            y_all[oof_assigned])
            print(f"    {variant_name} pooled OOF AUC = "
                  f"{pooled:.4f}; per-fold mean = "
                  f"{np.mean(fold_aucs):.4f}", flush=True)
            variant_data[variant_name] = {
                "fold_aucs": fold_aucs,
                "pooled_oof_auc": float(pooled),
                "per_fold_mean": float(np.mean(fold_aucs)),
                "per_fold_std": float(np.std(fold_aucs)),
            }

        seed_results[seed] = variant_data

    # ---- Aggregate across seeds ----
    print(f"\n{'='*70}\nAGGREGATE ACROSS {len(SEEDS)} SEEDS\n"
          f"{'='*70}", flush=True)
    summary = {}
    for variant_name in ["mask_only", "mask_plus_kernel"]:
        pooled = np.array([seed_results[s][variant_name][
            "pooled_oof_auc"] for s in SEEDS])
        per_fold = np.array([seed_results[s][variant_name][
            "per_fold_mean"] for s in SEEDS])
        print(f"  {variant_name}:")
        print(f"    Pooled OOF AUC across seeds: mean={pooled.mean():.4f}"
              f" +/- {pooled.std():.4f}  range=[{pooled.min():.4f}, "
              f"{pooled.max():.4f}]", flush=True)
        print(f"    Per-fold mean AUC across seeds: mean="
              f"{per_fold.mean():.4f} +/- {per_fold.std():.4f}",
              flush=True)
        summary[variant_name] = {
            "pooled_oof_mean": float(pooled.mean()),
            "pooled_oof_std": float(pooled.std()),
            "pooled_oof_min": float(pooled.min()),
            "pooled_oof_max": float(pooled.max()),
            "per_fold_mean_avg": float(per_fold.mean()),
            "per_fold_mean_std": float(per_fold.std()),
        }

    # Paired comparison: kernel rescue per seed
    rescues = []
    for s in SEEDS:
        d = (seed_results[s]["mask_plus_kernel"]["pooled_oof_auc"]
             - seed_results[s]["mask_only"]["pooled_oof_auc"])
        rescues.append(d)
    rescues = np.array(rescues)
    print(f"\n  Per-seed kernel rescue (Delta pooled OOF AUC):", flush=True)
    for s, r in zip(SEEDS, rescues):
        print(f"    seed={s}: Delta = {r:+.4f}", flush=True)
    print(f"  Mean rescue: {rescues.mean():+.4f} +/- "
          f"{rescues.std():.4f}", flush=True)
    print(f"  All seeds positive rescue: "
          f"{bool((rescues > 0).all())}", flush=True)

    out = {
        "version": "v207",
        "experiment": ("Multi-seed bootstrap of 3D CNN binary 365-d "
                       "PFS classifier mask-only vs mask+kernel"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "epochs_per_fold": EPOCHS,
        "n_total": n,
        "n_pos": n_pos,
        "n_neg": n_neg,
        "seeds": SEEDS,
        "per_seed": {str(k): v for k, v in seed_results.items()},
        "aggregate": summary,
        "kernel_rescue_per_seed": [float(r) for r in rescues],
        "kernel_rescue_mean": float(rescues.mean()),
        "kernel_rescue_std": float(rescues.std()),
        "all_seeds_positive_rescue": bool((rescues > 0).all()),
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
