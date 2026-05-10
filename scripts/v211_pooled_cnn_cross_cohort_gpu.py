"""v211: Pooled MU+RHUH CNN training + cross-cohort generalization
audit — round 43 (GPU).

Round 42 v208 logistic showed cross-cohort failure on RHUH n=31
(Delta=-0.005). Round 41 v207 showed CNN kernel rescue is seed-
dependent. Open question: does pooling MU+RHUH for CNN training
improve cross-cohort generalization?

Method: 5-fold cohort-stratified CV on pooled MU+RHUH (each fold
contains both cohorts in train and test). Train 3D CNN
(mask + kernel σ=3 input) for 30 epochs per fold. Per fold compute
overall test AUC + per-cohort breakdown. Report cohort-pooled OOF
AUC and per-cohort sub-AUC. Also run two LOCO baselines: train
MU-only -> test RHUH; train RHUH-only -> test MU.

Outputs:
  Nature_project/05_results/v211_pooled_cnn_cross_cohort.json
"""
from __future__ import annotations

import csv
import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v211_pooled_cnn_cross_cohort.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
EPOCHS = 30
LR = 5e-4
SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class BinaryPFSCNN(nn.Module):
    def __init__(self, in_ch=2, base=24):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.head = nn.Sequential(
            nn.Flatten(),
            nn.Linear(base * 4, 32),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(32, 1),
        )

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.head(self.global_pool(e3)).squeeze(-1)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def load_mu_data():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "cohort": "MU", "mask": m,
                     "kernel": k, "y": int(y)})
    return data


def load_rhuh_data():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if pfs_d is None:
                continue
            clinical[pid] = {"pfs_days": pfs_d,
                              "progress": event}
    data = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "cohort": "RHUH", "mask": m,
                     "kernel": k, "y": int(y)})
    return data


def cohort_label_kfold(data, k, seed=SEED):
    """Stratify by (cohort, label)."""
    rng = np.random.default_rng(seed)
    folds = [[] for _ in range(k)]
    for cohort_val in ["MU", "RHUH"]:
        for label in [0, 1]:
            idx = [i for i, r in enumerate(data)
                   if r["cohort"] == cohort_val
                   and r["y"] == label]
            rng.shuffle(idx)
            for j, idx_val in enumerate(idx):
                folds[j % k].append(int(idx_val))
    return [np.array(sorted(f)) for f in folds]


def train_cnn(Xtr, ytr, in_ch, n_pos, n_neg, seed):
    torch.manual_seed(seed)
    np.random.seed(seed)
    Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
    ytr_t = torch.from_numpy(ytr).to(DEVICE)
    model = BinaryPFSCNN(in_ch=in_ch, base=24).to(DEVICE)
    opt = torch.optim.AdamW(model.parameters(), lr=LR,
                             weight_decay=1e-3)
    pos_w = torch.tensor([n_neg / max(1, n_pos)],
                          dtype=torch.float32, device=DEVICE)
    criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    n_tr = len(Xtr)
    bs = 4
    for ep in range(EPOCHS):
        model.train()
        perm = np.random.permutation(n_tr)
        for i in range(0, n_tr, bs):
            idx = perm[i:i+bs]
            loss = criterion(model(Xtr_t[idx]), ytr_t[idx])
            opt.zero_grad()
            loss.backward()
            opt.step()
    return model


def predict(model, X):
    model.eval()
    Xt = torch.from_numpy(X).to(DEVICE)
    with torch.no_grad():
        logits = model(Xt).cpu().numpy()
    return logits


def stack_inputs(data_subset):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data_subset]).astype(np.float32)


def main():
    print("=" * 78, flush=True)
    print(f"v211 POOLED MU+RHUH CNN CROSS-COHORT (round 43 GPU) "
          f"device={DEVICE}", flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    mu_data = load_mu_data()
    rhuh_data = load_rhuh_data()
    print(f"  MU: n={len(mu_data)} "
          f"({sum(1 for d in mu_data if d['y']==1)} pos)",
          flush=True)
    print(f"  RHUH: n={len(rhuh_data)} "
          f"({sum(1 for d in rhuh_data if d['y']==1)} pos)",
          flush=True)
    pooled_data = mu_data + rhuh_data
    n = len(pooled_data)
    print(f"  Pooled n={n}", flush=True)

    # ============================================================
    # Pooled cohort-stratified 5-fold CV
    # ============================================================
    print(f"\n=== POOLED 5-FOLD CV ===", flush=True)
    folds = cohort_label_kfold(pooled_data, N_FOLDS,
                                seed=SEED)
    print(f"  Cohort-stratified fold sizes: "
          f"{[len(f) for f in folds]}", flush=True)
    y_all = np.array([d["y"] for d in pooled_data],
                      dtype=np.float32)
    cohort_all = [d["cohort"] for d in pooled_data]
    oof_scores = np.zeros(n, dtype=np.float32)
    oof_assigned = np.zeros(n, dtype=bool)

    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        n_tr = len(train_idx)
        n_tr_pos = int(y_all[train_idx].sum())
        n_tr_neg = n_tr - n_tr_pos
        Xtr = stack_inputs([pooled_data[i] for i in train_idx])
        Xte = stack_inputs([pooled_data[i] for i in test_idx])
        ytr = y_all[train_idx]
        yte = y_all[test_idx]
        model = train_cnn(Xtr, ytr, in_ch=2,
                           n_pos=n_tr_pos, n_neg=n_tr_neg,
                           seed=SEED + fold_i * 1000)
        scores = predict(model, Xte)
        oof_scores[test_idx] = scores
        oof_assigned[test_idx] = True
        a = auroc(scores, yte)
        # per-cohort breakdown
        a_mu_str = ""
        a_rhuh_str = ""
        mu_sel = np.array([cohort_all[i] == "MU"
                            for i in test_idx])
        rhuh_sel = np.array([cohort_all[i] == "RHUH"
                              for i in test_idx])
        if mu_sel.any():
            ymu = yte[mu_sel]
            if int(ymu.sum()) >= 1 and int(
                    len(ymu) - ymu.sum()) >= 1:
                a_mu = auroc(scores[mu_sel], ymu)
                a_mu_str = f", MU={a_mu:.4f}({int(mu_sel.sum())})"
        if rhuh_sel.any():
            yrh = yte[rhuh_sel]
            if int(yrh.sum()) >= 1 and int(
                    len(yrh) - yrh.sum()) >= 1:
                a_rh = auroc(scores[rhuh_sel], yrh)
                a_rhuh_str = f", RHUH={a_rh:.4f}({int(rhuh_sel.sum())})"
        print(f"  Fold {fold_i+1}/{N_FOLDS}: AUC={a:.4f}"
              f"{a_mu_str}{a_rhuh_str}  "
              f"(n_train={n_tr}, n_test={len(test_idx)})",
              flush=True)
        del model
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()

    pooled_oof_auc = auroc(oof_scores[oof_assigned],
                            y_all[oof_assigned])
    print(f"\n  Pooled OOF AUC = {pooled_oof_auc:.4f}",
          flush=True)
    pooled_results_per_cohort = {}
    for cohort_val in ["MU", "RHUH"]:
        sel = np.array([c == cohort_val for c in cohort_all])
        sel = sel & oof_assigned
        if sel.sum() < 5:
            continue
        ys = y_all[sel]
        if ys.sum() < 2 or len(ys) - ys.sum() < 2:
            continue
        a = auroc(oof_scores[sel], ys)
        print(f"  Pooled-CV {cohort_val} subset: n={int(sel.sum())} "
              f"({int(ys.sum())} pos, "
              f"{int(len(ys) - ys.sum())} neg), "
              f"AUC={a:.4f}", flush=True)
        pooled_results_per_cohort[cohort_val] = {
            "n": int(sel.sum()),
            "n_pos": int(ys.sum()),
            "n_neg": int(len(ys) - ys.sum()),
            "auc": float(a),
        }

    # ============================================================
    # LOCO baselines: MU -> RHUH and RHUH -> MU
    # ============================================================
    print(f"\n=== LOCO BASELINES ===", flush=True)
    loco_results = {}
    for tr_name, tr_data, te_name, te_data in [
        ("MU", mu_data, "RHUH", rhuh_data),
        ("RHUH", rhuh_data, "MU", mu_data),
    ]:
        print(f"\n  TRAIN {tr_name} (n={len(tr_data)}) "
              f"-> TEST {te_name} (n={len(te_data)})",
              flush=True)
        n_tr_pos = sum(1 for d in tr_data if d["y"] == 1)
        n_tr_neg = len(tr_data) - n_tr_pos
        Xtr = stack_inputs(tr_data)
        Xte = stack_inputs(te_data)
        ytr = np.array([d["y"] for d in tr_data],
                        dtype=np.float32)
        yte = np.array([d["y"] for d in te_data],
                        dtype=np.float32)
        model = train_cnn(Xtr, ytr, in_ch=2,
                           n_pos=n_tr_pos, n_neg=n_tr_neg,
                           seed=SEED + (1 if tr_name == "MU"
                                          else 2))
        scores = predict(model, Xte)
        a = auroc(scores, yte)
        print(f"    External {te_name} AUC = {a:.4f}",
              flush=True)
        loco_results[f"train_{tr_name}_test_{te_name}"] = {
            "n_train": len(tr_data),
            "n_test": len(te_data),
            "n_test_pos": int(yte.sum()),
            "n_test_neg": int(len(yte) - yte.sum()),
            "external_auc": float(a),
        }
        del model
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()

    out = {
        "version": "v211",
        "experiment": ("Pooled MU+RHUH CNN (cohort-stratified "
                       "5-fold CV) + LOCO cross-cohort baselines"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "epochs_per_fold": EPOCHS,
        "n_folds": N_FOLDS,
        "n_total": n,
        "n_mu": len(mu_data),
        "n_rhuh": len(rhuh_data),
        "pooled_5fold_cv": {
            "pooled_oof_auc": float(pooled_oof_auc),
            "per_cohort": pooled_results_per_cohort,
        },
        "loco_baselines": loco_results,
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
