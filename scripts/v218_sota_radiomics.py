"""v218: SOTA shape/morphological radiomics comparison —
round 47 (CPU).

Compute ~15 hand-crafted shape features from baseline tumor masks
and compare against V_kernel logistic for binary 365-d PFS:

Shape features:
  1. baseline_volume (voxel count)
  2. surface_area (boundary voxel count)
  3. sphericity = pi^(1/3) * (6V)^(2/3) / SA
  4. compactness2 = V^2 / SA^3 (dimensionless)
  5. inertia eigenvalues: elongation = sqrt(lambda2/lambda1),
                            flatness = sqrt(lambda3/lambda1)
  6. bounding-box volume ratio (V / V_bbox = solidity proxy)
  7. distance-to-boundary distribution: mean, std, skew, kurt
  8. connected components: n_components, max_component_fraction
  9. multi-sigma kernel volumes: V_kernel(sigma=2, 3, 4, 5)

Comparison classifiers:
  (a) clinical only (3 features: age + IDH + MGMT) — baseline
  (b) V_kernel only (1 feature)
  (c) clinical + V_kernel (4 features) — round 39 v202
  (d) all shape radiomics (~15 features) logistic
  (e) all radiomics + clinical
  (f) all radiomics + clinical + V_kernel — full kitchen sink

Tests whether V_kernel is SUFFICIENT or whether other hand-crafted
features add incremental value. SOTA radiomics-style comparison.

Outputs:
  Nature_project/05_results/v218_sota_radiomics.json
"""
from __future__ import annotations

import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
from scipy.ndimage import (binary_erosion, distance_transform_edt,
                              gaussian_filter, label as scipy_label)
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v218_sota_radiomics.json"

HORIZON = 365
N_BOOTSTRAPS = 1000
RNG_SEED = 42


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def shape_features(mask):
    """Compute ~15 morphological shape features from binary mask."""
    m = mask.astype(bool)
    V = float(m.sum())
    if V < 5:
        return None
    # Surface area: boundary voxels (not eroded)
    eroded = binary_erosion(m)
    boundary = m & ~eroded
    SA = float(boundary.sum())
    # Sphericity = pi^(1/3) * (6V)^(2/3) / SA  [ideal sphere = 1]
    sphericity = ((np.pi ** (1.0 / 3.0)
                    * (6 * V) ** (2.0 / 3.0)) / SA
                   if SA > 0 else 0.0)
    # Compactness2 = V^2 / SA^3 (dimensionless)
    compactness2 = (V * V) / (SA ** 3) if SA > 0 else 0.0
    # Coordinates of mask voxels
    coords = np.array(np.where(m), dtype=np.float64).T
    # Bounding box
    bbox_vol = float(np.prod(coords.max(axis=0)
                              - coords.min(axis=0) + 1))
    bbox_ratio = V / bbox_vol if bbox_vol > 0 else 0.0
    # Inertia tensor eigenvalues
    centered = coords - coords.mean(axis=0)
    cov = (centered.T @ centered) / max(1, len(coords) - 1)
    eigs = np.linalg.eigvalsh(cov)
    eigs = np.sort(eigs)[::-1]  # descending
    eigs = np.maximum(eigs, 1e-9)
    elongation = np.sqrt(eigs[1] / eigs[0])
    flatness = np.sqrt(eigs[2] / eigs[0])
    # Distance-to-boundary distribution (within mask)
    if eroded.sum() > 0:
        d_inside = distance_transform_edt(m)[m]
        d_mean = float(d_inside.mean())
        d_std = float(d_inside.std())
        # skewness, kurtosis
        d_centered = d_inside - d_mean
        d_skew = (float((d_centered ** 3).mean()
                          / (d_std ** 3 + 1e-9))
                   if d_std > 0 else 0.0)
        d_kurt = (float((d_centered ** 4).mean()
                          / (d_std ** 4 + 1e-9) - 3.0)
                   if d_std > 0 else 0.0)
    else:
        d_mean = d_std = d_skew = d_kurt = 0.0
    # Connected components
    labeled, n_comp = scipy_label(m)
    if n_comp > 0:
        comp_sizes = np.bincount(labeled.flat)[1:]
        max_frac = float(comp_sizes.max() / V)
    else:
        n_comp = 0
        max_frac = 0.0
    return {
        "volume": V,
        "surface_area": SA,
        "sphericity": float(sphericity),
        "compactness2": float(compactness2),
        "elongation": float(elongation),
        "flatness": float(flatness),
        "bbox_ratio": float(bbox_ratio),
        "d_boundary_mean": d_mean,
        "d_boundary_std": d_std,
        "d_boundary_skew": d_skew,
        "d_boundary_kurt": d_kurt,
        "n_components": float(n_comp),
        "max_comp_fraction": max_frac,
    }


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-np.clip(x, -50, 50)))


def build_X(rows, feats):
    arr = np.array([[r[f] for f in feats] for r in rows],
                    dtype=float)
    means = np.nanmean(arr, axis=0)
    stds = np.nanstd(arr, axis=0)
    stds[stds == 0] = 1
    arr_z = (arr - means) / stds
    arr_z = np.nan_to_num(arr_z, nan=0.0)
    return np.column_stack([np.ones(len(arr_z)), arr_z])


def continuous_nri(p_old, p_new, y):
    y = np.asarray(y).astype(int)
    p_old = np.asarray(p_old)
    p_new = np.asarray(p_new)
    pos = y == 1
    neg = y == 0
    if pos.sum() == 0 or neg.sum() == 0:
        return float("nan")
    nri_pos = ((p_new[pos] > p_old[pos]).mean()
                - (p_new[pos] < p_old[pos]).mean())
    nri_neg = ((p_new[neg] < p_old[neg]).mean()
                - (p_new[neg] > p_old[neg]).mean())
    return float(nri_pos + nri_neg)


def evaluate(rows, feats, y_field="y"):
    """Fit logistic with given feats, return AUC and predicted
    probabilities."""
    cc = [r for r in rows if all(
        r[f] is not None and not (isinstance(r[f], float)
                                  and np.isnan(r[f]))
        for f in feats)]
    if len(cc) < 20:
        return None
    y = np.array([r[y_field] for r in cc], dtype=float)
    if y.sum() < 5 or len(y) - y.sum() < 5:
        return None
    X = build_X(cc, feats)
    beta = logistic_fit(X, y)
    p_pred = sigmoid(X @ beta)
    auc = auroc(p_pred, y)
    return {"n": len(cc), "n_features": len(feats),
            "auc": float(auc), "p_pred": p_pred,
            "y": y, "beta": beta.tolist()}


def main():
    print("=" * 78, flush=True)
    print("v218 SOTA RADIOMICS COMPARISON (round 47 CPU)",
          flush=True)
    print("=" * 78, flush=True)
    rng = np.random.default_rng(RNG_SEED)

    # Load clinical
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            progress = (int(row[progress_col])
                        if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": progress,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }

    # Compute shape features for each patient
    print("\nComputing shape features per patient...", flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if (c["age"] is None or c["idh1"] is None
                or c["mgmt"] is None or c["pfs_days"] is None
                or c["progress"] is None):
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        sf = shape_features(m)
        if sf is None:
            continue
        # Multi-sigma kernels
        sf["v_kernel_s2"] = float(kernel_outgrowth_volume(m, 2.0))
        sf["v_kernel_s3"] = float(kernel_outgrowth_volume(m, 3.0))
        sf["v_kernel_s4"] = float(kernel_outgrowth_volume(m, 4.0))
        sf["v_kernel_s5"] = float(kernel_outgrowth_volume(m, 5.0))
        # Binary 365-d PFS label
        pfs = c["pfs_days"]
        prog = int(c["progress"])
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        rows.append({
            "pid": pid, "y": y,
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"], **sf,
        })
    n_total = len(rows)
    n_pos = sum(1 for r in rows if r["y"] == 1)
    n_neg = n_total - n_pos
    print(f"  {n_total} patients (pos={n_pos}, neg={n_neg})",
          flush=True)

    # Define feature sets
    feats_clin = ["age", "idh1", "mgmt"]
    feats_vk_s3 = ["v_kernel_s3"]
    feats_vk_multi = ["v_kernel_s2", "v_kernel_s3",
                       "v_kernel_s4", "v_kernel_s5"]
    feats_shape = [
        "volume", "surface_area", "sphericity", "compactness2",
        "elongation", "flatness", "bbox_ratio",
        "d_boundary_mean", "d_boundary_std", "d_boundary_skew",
        "d_boundary_kurt", "n_components",
        "max_comp_fraction",
    ]

    # Evaluate all model variants
    print("\n=== EVALUATING MODELS ===", flush=True)
    models = [
        ("clinical_only", feats_clin),
        ("Vkernel_s3_only", feats_vk_s3),
        ("Vkernel_multi_sigma", feats_vk_multi),
        ("clin_plus_Vkernel_s3", feats_clin + feats_vk_s3),
        ("clin_plus_Vkernel_multi",
         feats_clin + feats_vk_multi),
        ("shape_only", feats_shape),
        ("clin_plus_shape", feats_clin + feats_shape),
        ("Vkernel_s3_plus_shape", feats_vk_s3 + feats_shape),
        ("clin_plus_shape_plus_Vkernel_s3",
         feats_clin + feats_shape + feats_vk_s3),
        ("clin_plus_shape_plus_Vkernel_multi",
         feats_clin + feats_shape + feats_vk_multi),
    ]

    results = {}
    for name, feats in models:
        r = evaluate(rows, feats)
        if r is None:
            continue
        # Bootstrap CI on AUC
        bs_aucs = []
        n = len(rows)
        for _ in range(N_BOOTSTRAPS):
            idx = rng.integers(0, n, size=n)
            br = [rows[i] for i in idx]
            br_filt = [b for b in br if all(
                b[f] is not None for f in feats)]
            if len(br_filt) < 20:
                continue
            yb = np.array([b["y"] for b in br_filt],
                           dtype=float)
            if yb.sum() < 3 or len(yb) - yb.sum() < 3:
                continue
            Xb = build_X(br_filt, feats)
            betab = logistic_fit(Xb, yb)
            bs_aucs.append(auroc(Xb @ betab, yb))
        bs_aucs = np.array(bs_aucs)
        ci_lo = float(np.percentile(bs_aucs, 2.5))
        ci_hi = float(np.percentile(bs_aucs, 97.5))
        print(f"  {name:42s}  n_feats={r['n_features']:3d}  "
              f"AUC={r['auc']:.4f} [CI {ci_lo:.4f}, {ci_hi:.4f}]",
              flush=True)
        results[name] = {
            "features": feats,
            "n_features": r["n_features"],
            "auc": r["auc"],
            "auc_95_CI": [ci_lo, ci_hi],
            "auc_bootstrap_mean": float(bs_aucs.mean()),
        }

    # Pairwise NRI / IDI vs clinical-only
    print("\n=== RECLASSIFICATION vs clinical-only ===",
          flush=True)
    base_r = evaluate(rows, feats_clin)
    p_old = base_r["p_pred"]
    y_base = base_r["y"]
    nri_idi = {}
    for name, feats in models:
        if name == "clinical_only":
            continue
        r = evaluate(rows, feats)
        if r is None:
            continue
        # only valid if same complete-case set
        if len(r["p_pred"]) != len(p_old):
            continue
        nri = continuous_nri(p_old, r["p_pred"], y_base)
        # IDI
        pos = y_base == 1
        neg = y_base == 0
        idi_v = float((r["p_pred"][pos].mean()
                        - p_old[pos].mean())
                       - (r["p_pred"][neg].mean()
                          - p_old[neg].mean()))
        # Bootstrap CIs
        nri_bs = []
        idi_bs = []
        n = len(rows)
        for _ in range(N_BOOTSTRAPS):
            idx = rng.integers(0, n, size=n)
            br = [rows[i] for i in idx]
            br_filt = [b for b in br if all(
                b[f] is not None for f in feats + feats_clin)]
            if len(br_filt) < 20:
                continue
            yb = np.array([b["y"] for b in br_filt],
                           dtype=float)
            if yb.sum() < 3 or len(yb) - yb.sum() < 3:
                continue
            Xc_b = build_X(br_filt, feats_clin)
            X_b = build_X(br_filt, feats)
            bc = logistic_fit(Xc_b, yb)
            bn = logistic_fit(X_b, yb)
            po_b = sigmoid(Xc_b @ bc)
            pn_b = sigmoid(X_b @ bn)
            n_v = continuous_nri(po_b, pn_b, yb)
            posb = yb == 1
            negb = yb == 0
            if posb.sum() > 0 and negb.sum() > 0:
                ii = float((pn_b[posb].mean() - po_b[posb].mean())
                            - (pn_b[negb].mean()
                               - po_b[negb].mean()))
                if not np.isnan(n_v):
                    nri_bs.append(n_v)
                idi_bs.append(ii)
        nri_bs = np.array(nri_bs)
        idi_bs = np.array(idi_bs)
        nri_p = float((nri_bs <= 0).mean())
        idi_p = float((idi_bs <= 0).mean())
        print(f"  {name:42s}  NRI={nri:+.4f} (P={nri_p:.4f})  "
              f"IDI={idi_v:+.4f} (P={idi_p:.4f})", flush=True)
        nri_idi[name] = {
            "NRI": float(nri),
            "NRI_p_one_sided": nri_p,
            "NRI_95_CI": [float(np.percentile(nri_bs, 2.5)),
                            float(np.percentile(nri_bs, 97.5))],
            "IDI": float(idi_v),
            "IDI_p_one_sided": idi_p,
            "IDI_95_CI": [float(np.percentile(idi_bs, 2.5)),
                            float(np.percentile(idi_bs, 97.5))],
        }

    out = {
        "version": "v218",
        "experiment": ("SOTA shape/morphological radiomics "
                       "comparison vs V_kernel logistic"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "horizon_days": HORIZON,
        "n_total": n_total,
        "n_pos": n_pos,
        "n_neg": n_neg,
        "shape_features": feats_shape,
        "kernel_features_multi_sigma": feats_vk_multi,
        "models": results,
        "reclassification_vs_clin": nri_idi,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
