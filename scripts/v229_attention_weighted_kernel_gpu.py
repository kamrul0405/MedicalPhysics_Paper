"""v229: ATTENTION-WEIGHTED multi-σ kernel — round 52 (GPU).

Truly novel beyond-Lancet method: per-patient attention weighting
over σ values.

Round 47-51 used FIXED concatenation of V_kernel(σ=2,3,4,5) → 4
features → logistic. Round 52 v229: replace concatenation with
per-patient attention. Each patient gets attention scores α_k(x)
that softmax-weight the σ kernels:

  V_attn(x) = Σ_k α_k(x) * V_k(x)
  where α_k(x) = softmax_k(MLP([clinical, V_k]))

This allows the model to emphasize different σ scales for different
patients (e.g., σ=3 for some, σ=4 for others) — patient-specific
multi-scale weighting. Tests if patient-specific σ-emphasis improves
over global multi-σ logistic.

Method: 5-fold stratified CV on MU n=130. Compare:
  (A) Multi-σ fixed logistic (4 σ feats + 3 clinical) — baseline
  (B) Attention-weighted kernel + clinical → MLP head
Bootstrap CI on Δ AUC.

Outputs:
  Nature_project/05_results/v229_attention_weighted_kernel.json
"""
from __future__ import annotations

import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import gaussian_filter
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
OUT_JSON = RESULTS / "v229_attention_weighted_kernel.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

HORIZON = 365
N_FOLDS = 5
EPOCHS = 100
LR = 5e-3
SEED = 42
N_BOOT = 500


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def kernel_outgrowth_volume(mask, sigma):
    K = heat_bimodal(mask, sigma)
    m = mask.astype(bool)
    return int(((K >= 0.5) & ~m).sum())


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


class AttentionWeightedKernelModel(nn.Module):
    """Per-patient attention over σ kernels.

    Inputs:
        v_kernels: (B, K) log V_kernel values for each σ
        clinical: (B, C) clinical features

    Forward:
        Concat [clinical, v_kernels] → MLP → K attention logits
        α = softmax(K logits)
        V_attn = Σ α_k * v_kernels_k  → scalar
        Concat [clinical, V_attn] → small head → logit
    """

    def __init__(self, n_kernels=4, n_clin=3, hidden=16):
        super().__init__()
        self.n_kernels = n_kernels
        # Attention over σ
        self.attn_mlp = nn.Sequential(
            nn.Linear(n_clin + n_kernels, hidden), nn.GELU(),
            nn.Linear(hidden, n_kernels))
        # Output head
        self.head = nn.Sequential(
            nn.Linear(n_clin + 1, hidden), nn.GELU(),
            nn.Dropout(0.2),
            nn.Linear(hidden, 1))

    def forward(self, v_kernels, clinical):
        x_attn = torch.cat([clinical, v_kernels], dim=-1)
        attn_logits = self.attn_mlp(x_attn)
        alpha = F.softmax(attn_logits, dim=-1)
        v_attn = (alpha * v_kernels).sum(dim=-1, keepdim=True)
        x_head = torch.cat([clinical, v_attn], dim=-1)
        return self.head(x_head).squeeze(-1), alpha


def logistic_fit(X, y, l2=1e-2):
    n, p = X.shape

    def neg_log_lik(beta):
        eta = X @ beta
        log_one_plus = np.where(eta > 0,
                                 eta + np.log1p(np.exp(-eta)),
                                 np.log1p(np.exp(eta)))
        return -np.sum(y * eta - log_one_plus) + l2 * np.sum(
            beta * beta)

    return minimize(neg_log_lik, np.zeros(p),
                     method="L-BFGS-B").x


def standardize(X, means=None, stds=None):
    if means is None:
        means = X.mean(axis=0)
    if stds is None:
        stds = X.std(axis=0)
        stds[stds == 0] = 1
    return (X - means) / stds, means, stds


def main():
    print("=" * 78, flush=True)
    print(f"v229 ATTENTION-WEIGHTED KERNEL (round 52 GPU) "
          f"device={DEVICE}", flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    # Load data
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    age_col = header.index("Age at diagnosis")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    idh1_col = header.index("IDH1 mutation")
    mgmt_col = header.index("MGMT methylation")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            age = (float(row[age_col])
                   if row[age_col] is not None else None)
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
            idh1 = (float(row[idh1_col])
                    if row[idh1_col] is not None else None)
            mgmt = (float(row[mgmt_col])
                    if row[mgmt_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {
            "age": age, "progress": prog,
            "pfs_days": pfs_d, "idh1": idh1, "mgmt": mgmt,
        }
    print("\nComputing multi-σ kernels...", flush=True)
    rows = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if any(v is None for v in c.values()):
            continue
        m = (np.load(f) > 0).astype(np.float32)
        if m.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = int(c["progress"])
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        rows.append({
            "pid": pid, "y": y,
            "age": c["age"], "idh1": c["idh1"],
            "mgmt": c["mgmt"],
            "v_k2": float(kernel_outgrowth_volume(m, 2.0)),
            "v_k3": float(kernel_outgrowth_volume(m, 3.0)),
            "v_k4": float(kernel_outgrowth_volume(m, 4.0)),
            "v_k5": float(kernel_outgrowth_volume(m, 5.0)),
        })
    n = len(rows)
    print(f"  {n} labelled MU patients", flush=True)
    y_all = np.array([r["y"] for r in rows], dtype=np.float32)
    folds = stratified_kfold(y_all, N_FOLDS, seed=SEED)
    print(f"  Fold sizes: {[len(f) for f in folds]}", flush=True)

    # Build feature arrays
    clin_arr = np.array(
        [[r["age"], r["idh1"], r["mgmt"]] for r in rows],
        dtype=np.float32)
    vk_arr = np.array(
        [[r["v_k2"], r["v_k3"], r["v_k4"], r["v_k5"]]
         for r in rows], dtype=np.float32)
    vk_log = np.log1p(vk_arr)
    fixed_full = np.concatenate([clin_arr, vk_log], axis=1)

    # ---- Variant A: fixed multi-σ logistic baseline ----
    print(f"\n=== VARIANT A: fixed multi-σ logistic ===",
          flush=True)
    oof_a = np.zeros(n, dtype=np.float32)
    assigned_a = np.zeros(n, dtype=bool)
    fold_aucs_a = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        Xtr_raw = fixed_full[train_idx]
        Xte_raw = fixed_full[test_idx]
        Xtr_z, m, s = standardize(Xtr_raw)
        Xte_z, _, _ = standardize(Xte_raw, means=m, stds=s)
        Xtr = np.column_stack([np.ones(len(Xtr_z)), Xtr_z])
        Xte = np.column_stack([np.ones(len(Xte_z)), Xte_z])
        ytr = y_all[train_idx]
        yte = y_all[test_idx]
        if int(ytr.sum()) < 5 or int(len(ytr) - ytr.sum()) < 5:
            continue
        beta = logistic_fit(Xtr, ytr)
        scores = Xte @ beta
        oof_a[test_idx] = scores
        assigned_a[test_idx] = True
        a = auroc(scores, yte)
        fold_aucs_a.append(float(a))
        print(f"    Fold {fold_i+1}: AUC={a:.4f}", flush=True)
    pooled_a = auroc(oof_a[assigned_a], y_all[assigned_a])
    print(f"  Variant A pooled OOF AUC = {pooled_a:.4f}",
          flush=True)

    # ---- Variant B: Attention-weighted kernel ----
    print(f"\n=== VARIANT B: ATTENTION-WEIGHTED kernel ===",
          flush=True)
    oof_b = np.zeros(n, dtype=np.float32)
    assigned_b = np.zeros(n, dtype=bool)
    fold_aucs_b = []
    learned_attn_per_fold = []
    for fold_i, test_idx in enumerate(folds):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        clin_tr = clin_arr[train_idx]
        clin_te = clin_arr[test_idx]
        vk_tr = vk_log[train_idx]
        vk_te = vk_log[test_idx]
        ytr = y_all[train_idx]
        yte = y_all[test_idx]
        n_pos = int(ytr.sum())
        n_neg = len(ytr) - n_pos
        if n_pos < 2 or n_neg < 2:
            continue
        # Standardize
        clin_mean = clin_tr.mean(axis=0)
        clin_std = clin_tr.std(axis=0)
        clin_std[clin_std == 0] = 1
        clin_tr_z = (clin_tr - clin_mean) / clin_std
        clin_te_z = (clin_te - clin_mean) / clin_std
        vk_mean = vk_tr.mean(axis=0)
        vk_std = vk_tr.std(axis=0)
        vk_std[vk_std == 0] = 1
        vk_tr_z = (vk_tr - vk_mean) / vk_std
        vk_te_z = (vk_te - vk_mean) / vk_std

        clin_tr_t = torch.from_numpy(clin_tr_z.astype(
            np.float32)).to(DEVICE)
        clin_te_t = torch.from_numpy(clin_te_z.astype(
            np.float32)).to(DEVICE)
        vk_tr_t = torch.from_numpy(vk_tr_z.astype(
            np.float32)).to(DEVICE)
        vk_te_t = torch.from_numpy(vk_te_z.astype(
            np.float32)).to(DEVICE)
        ytr_t = torch.from_numpy(ytr).to(DEVICE)

        torch.manual_seed(SEED + fold_i * 100)
        np.random.seed(SEED + fold_i * 100)
        model = AttentionWeightedKernelModel(
            n_kernels=4, n_clin=3, hidden=16).to(DEVICE)
        opt = torch.optim.AdamW(model.parameters(), lr=LR,
                                 weight_decay=1e-3)
        pos_w = torch.tensor([n_neg / max(1, n_pos)],
                              dtype=torch.float32, device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        n_tr = len(train_idx)
        bs = 8
        for ep in range(EPOCHS):
            model.train()
            perm = np.random.permutation(n_tr)
            for i in range(0, n_tr, bs):
                idx = perm[i:i+bs]
                logits, _ = model(vk_tr_t[idx], clin_tr_t[idx])
                loss = criterion(logits, ytr_t[idx])
                opt.zero_grad()
                loss.backward()
                opt.step()
        model.eval()
        with torch.no_grad():
            logits_te, alpha_te = model(vk_te_t, clin_te_t)
            scores = logits_te.cpu().numpy()
            attn_te = alpha_te.cpu().numpy()
        oof_b[test_idx] = scores
        assigned_b[test_idx] = True
        a = auroc(scores, yte)
        fold_aucs_b.append(float(a))
        # Average attention weights on this test fold
        avg_attn = attn_te.mean(axis=0)
        learned_attn_per_fold.append(avg_attn.tolist())
        print(f"    Fold {fold_i+1}: AUC={a:.4f}, mean attn σ_2-5"
              f"={[round(float(x), 3) for x in avg_attn]}",
              flush=True)
        del model, opt
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()
    pooled_b = auroc(oof_b[assigned_b], y_all[assigned_b])
    print(f"  Variant B pooled OOF AUC = {pooled_b:.4f}",
          flush=True)

    # Average attention across all test patients
    attn_arr = np.array(learned_attn_per_fold)
    attn_mean = attn_arr.mean(axis=0)
    print(f"\n  Average learned attention weights across folds:",
          flush=True)
    for k, w in enumerate(attn_mean):
        print(f"    σ_{k+2}: {w:.3f}", flush=True)

    # ---- Bootstrap pairwise Δ ----
    print(f"\n=== Bootstrap pairwise Δ AUC ===", flush=True)
    rng = np.random.default_rng(SEED)
    n_assigned = int(assigned_a.sum())
    deltas_b_minus_a = []
    for _ in range(N_BOOT):
        idx = rng.integers(0, n_assigned, size=n_assigned)
        valid = np.where(assigned_a)[0][idx]
        yb = y_all[valid]
        if int(yb.sum()) < 3 or int(len(yb) - yb.sum()) < 3:
            continue
        a_a = auroc(oof_a[valid], yb)
        a_b = auroc(oof_b[valid], yb)
        deltas_b_minus_a.append(a_b - a_a)
    deltas_b_minus_a = np.array(deltas_b_minus_a)
    ci_lo = float(np.percentile(deltas_b_minus_a, 2.5))
    ci_hi = float(np.percentile(deltas_b_minus_a, 97.5))
    p_one = float((deltas_b_minus_a <= 0).mean())
    print(f"  Δ B - A: mean={deltas_b_minus_a.mean():+.4f}, "
          f"95% CI [{ci_lo:+.4f}, {ci_hi:+.4f}], "
          f"P<=0={p_one:.4f}", flush=True)

    print(f"\n=== SUMMARY ===", flush=True)
    print(f"  Variant A fixed multi-σ logistic OOF: "
          f"{pooled_a:.4f}", flush=True)
    print(f"  Variant B attention-weighted OOF:     "
          f"{pooled_b:.4f}", flush=True)

    out = {
        "version": "v229",
        "experiment": ("Attention-weighted multi-σ kernel "
                       "via per-patient softmax attention"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "epochs_per_fold": EPOCHS,
        "n_total": n,
        "results": {
            "A_fixed_multi_sigma_logistic": {
                "fold_aucs": fold_aucs_a,
                "fold_mean": float(np.mean(fold_aucs_a)),
                "pooled_oof_auc": float(pooled_a),
            },
            "B_attention_weighted": {
                "fold_aucs": fold_aucs_b,
                "fold_mean": float(np.mean(fold_aucs_b)),
                "pooled_oof_auc": float(pooled_b),
                "learned_attention_per_fold":
                    learned_attn_per_fold,
                "learned_attention_mean": [
                    float(x) for x in attn_mean],
            },
        },
        "bootstrap_delta_b_minus_a": {
            "mean": float(deltas_b_minus_a.mean()),
            "95_CI": [ci_lo, ci_hi],
            "p_one_sided": p_one,
        },
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
