"""v217: Comprehensive pretraining-strategy ablation on RHUH transfer
— round 46 (GPU).

Round 44 v213: supervised MU pretrain + frozen-encoder + head-only
RHUH FT -> AUC=0.804.
Round 45 v215: SimCLR multi-cohort label-free pretrain + head FT
on MU -> per-fold 0.706.

Combine these two findings into a definitive 4-way ablation on
RHUH 5-fold CV: which pretraining strategy is best for cross-cohort
PFS deployment?

Variants:
  1. RANDOM INIT: train CNN from scratch on RHUH (5-fold).
  2. SUPERVISED MU: pretrain CNN on MU labelled binary 365-d PFS,
     freeze encoder, head-only FT on RHUH.
  3. SimCLR MULTI-COHORT (LABEL-FREE): SimCLR pretrain on MU + UCSF
     + RHUH + LUMIERE masks (no labels), freeze encoder, head-only
     FT on RHUH.
  4. STACKED (SimCLR + supervised MU): SimCLR pretrain → unfreeze
     and supervised MU FT → freeze encoder → head-only FT on RHUH.
     Tests if combining both pretraining strategies gives
     additional lift.

5-fold stratified CV on RHUH n=31 for each variant. 200-bootstrap
CI on Delta AUC vs random-init baseline.

Outputs:
  Nature_project/05_results/v217_pretrain_ablation.json
"""
from __future__ import annotations

import csv
import gc
import json
import time
import warnings
from pathlib import Path

import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.ndimage import gaussian_filter, zoom

warnings.filterwarnings("ignore")

ROOT = Path(r"C:\Users\kamru\Downloads\Nature_project")
RESULTS = ROOT / "05_results"
CACHE = RESULTS / "cache_3d"
CLINICAL_MU = Path(
    r"C:/Users/kamru/Downloads/MU-Glioma-Post_ClinicalData-July2025.xlsx")
CLINICAL_RHUH = Path(
    r"C:\Users\kamru\Downloads\clinical_data_TCIA_RHUH-GBM.csv")
OUT_JSON = RESULTS / "v217_pretrain_ablation.json"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

SIGMA_KERNEL = 3.0
TARGET_SHAPE = (16, 48, 48)
HORIZON = 365
N_FOLDS = 5
SIMCLR_EPOCHS = 30
MU_SUP_EPOCHS = 30
FT_EPOCHS = 50
SCRATCH_EPOCHS = 30
LR_SIMCLR = 1e-3
LR_SUP = 5e-4
LR_FT = 1e-3
TEMPERATURE = 0.5
PROJ_DIM = 64
BASE = 24
SEED = 42
N_BOOT = 200

COHORT_PREFIXES_PRE = ["MU-Glioma-Post", "RHUH-GBM",
                         "UCSF-POSTOP", "LUMIERE"]


def heat_constant(mask, sigma):
    if mask.sum() == 0:
        return np.zeros_like(mask, dtype=np.float32)
    h = gaussian_filter(mask.astype(np.float32), sigma=sigma)
    if h.max() > 0:
        h = h / h.max()
    return h.astype(np.float32)


def heat_bimodal(mask, sigma):
    persistence = mask.astype(np.float32)
    return np.maximum(persistence, heat_constant(mask, sigma))


def resize_to_target(arr, target_shape):
    factors = [t / s for t, s in zip(target_shape, arr.shape)]
    if arr.dtype == bool or np.array_equal(
            arr, arr.astype(bool).astype(arr.dtype)):
        return zoom(arr.astype(np.float32), factors,
                    order=0).astype(np.float32)
    return zoom(arr, factors, order=1).astype(np.float32)


class Encoder(nn.Module):
    def __init__(self, in_ch=2, base=BASE):
        super().__init__()
        self.enc1 = self._block(in_ch, base)
        self.enc2 = self._block(base, base * 2)
        self.enc3 = self._block(base * 2, base * 4)
        self.pool = nn.MaxPool3d(2)
        self.global_pool = nn.AdaptiveAvgPool3d(1)
        self.feat_dim = base * 4

    def _block(self, in_ch, out_ch):
        return nn.Sequential(
            nn.Conv3d(in_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
            nn.Conv3d(out_ch, out_ch, 3, padding=1),
            nn.GroupNorm(8, out_ch), nn.GELU(),
        )

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        return self.global_pool(e3).flatten(1)


class ProjectionHead(nn.Module):
    def __init__(self, feat_dim=BASE * 4, proj_dim=PROJ_DIM):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, feat_dim), nn.GELU(),
            nn.Linear(feat_dim, proj_dim))

    def forward(self, feat):
        return F.normalize(self.net(feat), dim=-1)


class BinaryHead(nn.Module):
    def __init__(self, feat_dim=BASE * 4):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(feat_dim, 32), nn.GELU(),
            nn.Dropout(0.3), nn.Linear(32, 1))

    def forward(self, feat):
        return self.net(feat).squeeze(-1)


def random_3d_augment(x, rng):
    out = x.clone()
    for axis in [2, 3, 4]:
        if rng.random() > 0.5:
            out = torch.flip(out, dims=[axis])
    scale = float(rng.uniform(0.9, 1.1))
    out = out * scale
    out = out + 0.02 * torch.randn_like(out)
    return torch.clamp(out, 0.0, 1.5)


def nt_xent_loss(z1, z2, temperature=TEMPERATURE):
    n = z1.size(0)
    z = torch.cat([z1, z2], dim=0)
    sim = torch.matmul(z, z.T) / temperature
    mask = torch.eye(2 * n, dtype=torch.bool,
                      device=sim.device)
    sim = sim.masked_fill(mask, float("-inf"))
    targets = torch.cat([torch.arange(n, 2 * n),
                          torch.arange(0, n)],
                          dim=0).to(sim.device)
    return F.cross_entropy(sim, targets)


def auroc(scores, labels):
    s = np.array(scores, dtype=np.float64)
    y = np.array(labels, dtype=np.float64)
    n_pos = int(y.sum())
    n_neg = len(y) - n_pos
    if n_pos == 0 or n_neg == 0:
        return float("nan")
    order = np.argsort(-s)
    y_sorted = y[order]
    cum_pos = np.cumsum(y_sorted)
    fpr = (np.arange(1, len(y) + 1) - cum_pos) / n_neg
    tpr = cum_pos / n_pos
    fpr = np.concatenate([[0.0], fpr])
    tpr = np.concatenate([[0.0], tpr])
    auc = float(np.trapezoid(tpr, fpr) if hasattr(np, "trapezoid")
                 else np.trapz(tpr, fpr))
    if auc < 0.5:
        auc = 1.0 - auc
    return auc


def stratified_kfold(y, k, seed=SEED):
    rng = np.random.default_rng(seed)
    y = np.asarray(y)
    pos_idx = np.where(y == 1)[0]
    neg_idx = np.where(y == 0)[0]
    rng.shuffle(pos_idx)
    rng.shuffle(neg_idx)
    folds = [[] for _ in range(k)]
    for i, idx in enumerate(pos_idx):
        folds[i % k].append(int(idx))
    for i, idx in enumerate(neg_idx):
        folds[i % k].append(int(idx))
    return [np.array(sorted(f)) for f in folds]


def stack_inputs(data):
    return np.stack([np.stack([d["mask"], d["kernel"]], axis=0)
                      for d in data]).astype(np.float32)


def load_all_masks_for_pretrain():
    samples = []
    for prefix in COHORT_PREFIXES_PRE:
        for f in sorted(CACHE.glob(f"{prefix}_*_b.npy")):
            try:
                m_full = (np.load(f) > 0).astype(np.float32)
                if m_full.sum() == 0:
                    continue
                m = resize_to_target(m_full, TARGET_SHAPE)
                k = heat_bimodal(m, SIGMA_KERNEL)
                samples.append({
                    "pid": f.stem, "cohort": prefix,
                    "mask": m, "kernel": k,
                })
            except Exception:
                continue
    return samples


def load_mu_labelled():
    wb = openpyxl.load_workbook(CLINICAL_MU, data_only=True)
    ws = wb["MU Glioma Post"]
    header = [str(h) if h else "" for h in next(
        ws.iter_rows(values_only=True))]
    pid_col = header.index("Patient_ID")
    progress_col = header.index("Progression")
    pfs_col = header.index("Time to First Progression (Days)")
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        try:
            prog = (int(row[progress_col])
                    if row[progress_col] is not None else None)
            pfs_d = (float(row[pfs_col])
                     if row[pfs_col] is not None else None)
        except (ValueError, TypeError):
            continue
        clinical[str(pid)] = {"progress": prog,
                              "pfs_days": pfs_d}
    data = []
    for f in sorted(CACHE.glob("MU-Glioma-Post_PatientID_*_b.npy")):
        pid = f.stem.replace("MU-Glioma-Post_", "").replace(
            "_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        if c["progress"] is None or c["pfs_days"] is None:
            continue
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def load_rhuh_labelled():
    clinical = {}
    with CLINICAL_RHUH.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames
        pfs_col = next((c for c in cols
                         if "Progression free survival" in c
                         and "PFS" in c and "day" in c.lower()),
                        None)
        cens_col = next((c for c in cols
                          if c.strip().lower() == "right censored"),
                         None)
        for row in reader:
            pid = row["Patient ID"].strip()
            try:
                pfs_d = (float(row[pfs_col]) if row[pfs_col]
                         else None)
                cens = row[cens_col].strip().lower()
                event = (0 if cens == "yes" else 1)
            except (ValueError, TypeError, AttributeError):
                continue
            if pfs_d is None:
                continue
            clinical[pid] = {"pfs_days": pfs_d,
                              "progress": event}
    data = []
    for f in sorted(CACHE.glob("RHUH-GBM_*_b.npy")):
        pid = f.stem.replace("RHUH-GBM_", "").replace("_b", "")
        if pid not in clinical:
            continue
        c = clinical[pid]
        m_full = (np.load(f) > 0).astype(np.float32)
        if m_full.sum() == 0:
            continue
        pfs = c["pfs_days"]
        prog = c["progress"]
        if prog == 1 and pfs < HORIZON:
            y = 1
        elif (prog == 0 and pfs >= HORIZON) or (prog == 1
                                                 and pfs >= HORIZON):
            y = 0
        else:
            continue
        m = resize_to_target(m_full, TARGET_SHAPE)
        k = heat_bimodal(m, SIGMA_KERNEL)
        data.append({"pid": pid, "mask": m, "kernel": k,
                     "y": int(y)})
    return data


def simclr_pretrain(samples, epochs, seed=SEED):
    torch.manual_seed(seed)
    np.random.seed(seed)
    rng_np = np.random.default_rng(seed)
    X = stack_inputs(samples)
    Xt = torch.from_numpy(X).to(DEVICE)
    enc = Encoder(in_ch=2, base=BASE).to(DEVICE)
    proj = ProjectionHead(feat_dim=enc.feat_dim,
                            proj_dim=PROJ_DIM).to(DEVICE)
    opt = torch.optim.AdamW(
        list(enc.parameters()) + list(proj.parameters()),
        lr=LR_SIMCLR, weight_decay=1e-4)
    n = len(samples)
    bs = 16
    for ep in range(epochs):
        enc.train()
        proj.train()
        perm = np.random.permutation(n)
        for i in range(0, n, bs):
            idx = perm[i:i+bs]
            if len(idx) < 2:
                continue
            xb = Xt[idx]
            v1 = random_3d_augment(xb, rng_np)
            v2 = random_3d_augment(xb, rng_np)
            z1 = proj(enc(v1))
            z2 = proj(enc(v2))
            loss = nt_xent_loss(z1, z2)
            opt.zero_grad()
            loss.backward()
            opt.step()
    return enc


def supervised_mu_pretrain(enc, mu_data, epochs, seed=SEED,
                              freeze_enc=False):
    """Train (or fine-tune) enc + new head on MU labelled."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    n = len(mu_data)
    n_pos = sum(1 for d in mu_data if d["y"] == 1)
    n_neg = n - n_pos
    X = stack_inputs(mu_data)
    y = np.array([d["y"] for d in mu_data], dtype=np.float32)
    Xt = torch.from_numpy(X).to(DEVICE)
    yt = torch.from_numpy(y).to(DEVICE)
    head = BinaryHead(feat_dim=enc.feat_dim).to(DEVICE)
    if freeze_enc:
        for p in enc.parameters():
            p.requires_grad = False
        params = list(head.parameters())
    else:
        for p in enc.parameters():
            p.requires_grad = True
        params = list(enc.parameters()) + list(
            head.parameters())
    opt = torch.optim.AdamW(params, lr=LR_SUP,
                             weight_decay=1e-3)
    pos_w = torch.tensor([n_neg / max(1, n_pos)],
                          dtype=torch.float32, device=DEVICE)
    criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
    bs = 4
    for ep in range(epochs):
        if not freeze_enc:
            enc.train()
        else:
            enc.eval()
        head.train()
        perm = np.random.permutation(n)
        for i in range(0, n, bs):
            idx = perm[i:i+bs]
            with torch.set_grad_enabled(not freeze_enc):
                feat = enc(Xt[idx])
            logits = head(feat)
            loss = criterion(logits, yt[idx])
            opt.zero_grad()
            loss.backward()
            opt.step()
    return enc


def head_finetune_on_rhuh(enc, rhuh_data, fold_idx_list,
                            seed=SEED):
    """5-fold CV: per fold, freeze enc, train head on RHUH train,
    eval on test. Returns OOF scores."""
    for p in enc.parameters():
        p.requires_grad = False
    enc.eval()
    n = len(rhuh_data)
    y_all = np.array([d["y"] for d in rhuh_data],
                      dtype=np.float32)
    X_all = stack_inputs(rhuh_data)
    Xt_all = torch.from_numpy(X_all).to(DEVICE)
    with torch.no_grad():
        feat_all = enc(Xt_all)
    oof = np.zeros(n, dtype=np.float32)
    assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(fold_idx_list):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        n_tr_pos = int(y_all[train_idx].sum())
        n_tr_neg = len(train_idx) - n_tr_pos
        if n_tr_pos < 2 or n_tr_neg < 2:
            continue
        feat_tr = feat_all[train_idx]
        feat_te = feat_all[test_idx]
        ytr = torch.from_numpy(y_all[train_idx]).to(DEVICE)
        yte = y_all[test_idx]
        torch.manual_seed(seed + fold_i * 100)
        head = BinaryHead(feat_dim=enc.feat_dim).to(DEVICE)
        opt_h = torch.optim.AdamW(head.parameters(), lr=LR_FT,
                                    weight_decay=1e-3)
        pos_w = torch.tensor(
            [n_tr_neg / max(1, n_tr_pos)],
            dtype=torch.float32, device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        n_tr = len(train_idx)
        bs = min(8, n_tr)
        for ep in range(FT_EPOCHS):
            head.train()
            perm = np.random.permutation(n_tr)
            for i in range(0, n_tr, bs):
                idx = perm[i:i+bs]
                logits = head(feat_tr[idx])
                loss = criterion(logits, ytr[idx])
                opt_h.zero_grad()
                loss.backward()
                opt_h.step()
        head.eval()
        with torch.no_grad():
            scores = head(feat_te).cpu().numpy()
        oof[test_idx] = scores
        assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            fold_aucs.append(float(auroc(scores, yte)))
        del head, opt_h
        gc.collect()
    return oof, assigned, fold_aucs


def from_scratch_on_rhuh(rhuh_data, fold_idx_list, seed=SEED):
    """5-fold CV: from-scratch CNN per fold."""
    n = len(rhuh_data)
    y_all = np.array([d["y"] for d in rhuh_data],
                      dtype=np.float32)
    oof = np.zeros(n, dtype=np.float32)
    assigned = np.zeros(n, dtype=bool)
    fold_aucs = []
    for fold_i, test_idx in enumerate(fold_idx_list):
        train_idx = np.array(sorted(
            set(range(n)) - set(test_idx.tolist())))
        tr = [rhuh_data[i] for i in train_idx]
        te = [rhuh_data[i] for i in test_idx]
        n_pos = sum(1 for d in tr if d["y"] == 1)
        n_neg = len(tr) - n_pos
        if n_pos < 2 or n_neg < 2:
            continue
        Xtr = stack_inputs(tr)
        Xte = stack_inputs(te)
        ytr = np.array([d["y"] for d in tr], dtype=np.float32)
        yte = np.array([d["y"] for d in te], dtype=np.float32)
        Xtr_t = torch.from_numpy(Xtr).to(DEVICE)
        Xte_t = torch.from_numpy(Xte).to(DEVICE)
        ytr_t = torch.from_numpy(ytr).to(DEVICE)
        torch.manual_seed(seed + fold_i * 100)
        np.random.seed(seed + fold_i * 100)
        enc = Encoder(in_ch=2, base=BASE).to(DEVICE)
        head = BinaryHead(feat_dim=enc.feat_dim).to(DEVICE)
        opt = torch.optim.AdamW(
            list(enc.parameters()) + list(head.parameters()),
            lr=LR_SUP, weight_decay=1e-3)
        pos_w = torch.tensor([n_neg / max(1, n_pos)],
                              dtype=torch.float32, device=DEVICE)
        criterion = nn.BCEWithLogitsLoss(pos_weight=pos_w)
        bs = min(4, len(tr))
        for ep in range(SCRATCH_EPOCHS):
            enc.train()
            head.train()
            perm = np.random.permutation(len(tr))
            for i in range(0, len(tr), bs):
                idx = perm[i:i+bs]
                feat = enc(Xtr_t[idx])
                logits = head(feat)
                loss = criterion(logits, ytr_t[idx])
                opt.zero_grad()
                loss.backward()
                opt.step()
        enc.eval()
        head.eval()
        with torch.no_grad():
            feat_te = enc(Xte_t)
            scores = head(feat_te).cpu().numpy()
        oof[test_idx] = scores
        assigned[test_idx] = True
        if int(yte.sum()) >= 1 and int(len(yte) - yte.sum()) >= 1:
            fold_aucs.append(float(auroc(scores, yte)))
        del enc, head, opt
        gc.collect()
        if DEVICE.type == "cuda":
            torch.cuda.empty_cache()
    return oof, assigned, fold_aucs


def main():
    print("=" * 78, flush=True)
    print(f"v217 PRETRAIN ABLATION (round 46 GPU) device={DEVICE}",
          flush=True)
    print("=" * 78, flush=True)
    t_start = time.time()

    # Load data
    print("\n=== Loading data ===", flush=True)
    pretrain_samples = load_all_masks_for_pretrain()
    print(f"  Multi-cohort masks for SimCLR: "
          f"{len(pretrain_samples)}", flush=True)
    mu_data = load_mu_labelled()
    print(f"  MU labelled: {len(mu_data)}", flush=True)
    rhuh_data = load_rhuh_labelled()
    print(f"  RHUH labelled: {len(rhuh_data)}", flush=True)
    y_rhuh = np.array([d["y"] for d in rhuh_data],
                       dtype=np.float32)
    folds_rhuh = stratified_kfold(y_rhuh, N_FOLDS, seed=SEED)
    print(f"  RHUH fold sizes: "
          f"{[len(f) for f in folds_rhuh]}", flush=True)

    results = {}

    # ---- Variant 1: random init from-scratch on RHUH ----
    print(f"\n=== VARIANT 1: random init from-scratch RHUH "
          f"5-fold ===", flush=True)
    oof1, assigned1, fold_aucs1 = from_scratch_on_rhuh(
        rhuh_data, folds_rhuh, seed=SEED)
    pooled1 = float(auroc(oof1[assigned1], y_rhuh[assigned1]))
    print(f"  Variant 1 fold AUCs: {fold_aucs1}", flush=True)
    print(f"  Variant 1 pooled OOF AUC = {pooled1:.4f}",
          flush=True)
    results["v1_random_init_scratch"] = {
        "fold_aucs": fold_aucs1,
        "fold_mean": float(np.mean(fold_aucs1)),
        "pooled_oof_auc": pooled1,
    }

    # ---- Variant 2: supervised MU pretrain + frozen enc + RHUH FT ----
    print(f"\n=== VARIANT 2: supervised MU pretrain + frozen enc "
          f"+ RHUH FT ===", flush=True)
    enc2 = Encoder(in_ch=2, base=BASE).to(DEVICE)
    enc2 = supervised_mu_pretrain(enc2, mu_data,
                                     epochs=MU_SUP_EPOCHS,
                                     seed=SEED, freeze_enc=False)
    oof2, assigned2, fold_aucs2 = head_finetune_on_rhuh(
        enc2, rhuh_data, folds_rhuh, seed=SEED)
    pooled2 = float(auroc(oof2[assigned2], y_rhuh[assigned2]))
    print(f"  Variant 2 fold AUCs: {fold_aucs2}", flush=True)
    print(f"  Variant 2 pooled OOF AUC = {pooled2:.4f}",
          flush=True)
    results["v2_supervised_mu_pretrain"] = {
        "fold_aucs": fold_aucs2,
        "fold_mean": float(np.mean(fold_aucs2)),
        "pooled_oof_auc": pooled2,
    }
    del enc2
    gc.collect()
    if DEVICE.type == "cuda":
        torch.cuda.empty_cache()

    # ---- Variant 3: SimCLR multi-cohort + frozen enc + RHUH FT ----
    print(f"\n=== VARIANT 3: SimCLR multi-cohort (LABEL-FREE) + "
          f"frozen enc + RHUH FT ===", flush=True)
    enc3 = simclr_pretrain(pretrain_samples, SIMCLR_EPOCHS,
                              seed=SEED)
    oof3, assigned3, fold_aucs3 = head_finetune_on_rhuh(
        enc3, rhuh_data, folds_rhuh, seed=SEED)
    pooled3 = float(auroc(oof3[assigned3], y_rhuh[assigned3]))
    print(f"  Variant 3 fold AUCs: {fold_aucs3}", flush=True)
    print(f"  Variant 3 pooled OOF AUC = {pooled3:.4f}",
          flush=True)
    results["v3_simclr_multicohort"] = {
        "fold_aucs": fold_aucs3,
        "fold_mean": float(np.mean(fold_aucs3)),
        "pooled_oof_auc": pooled3,
    }

    # ---- Variant 4: stacked (SimCLR -> supervised MU -> RHUH FT) ----
    print(f"\n=== VARIANT 4: STACKED (SimCLR -> supervised MU "
          f"-> RHUH FT) ===", flush=True)
    enc4 = enc3  # reuse SimCLR-pretrained encoder
    enc4 = supervised_mu_pretrain(
        enc4, mu_data, epochs=MU_SUP_EPOCHS, seed=SEED + 1,
        freeze_enc=False)  # supervised FT on MU (encoder fine-
                             # tunable from SimCLR init)
    oof4, assigned4, fold_aucs4 = head_finetune_on_rhuh(
        enc4, rhuh_data, folds_rhuh, seed=SEED)
    pooled4 = float(auroc(oof4[assigned4], y_rhuh[assigned4]))
    print(f"  Variant 4 fold AUCs: {fold_aucs4}", flush=True)
    print(f"  Variant 4 pooled OOF AUC = {pooled4:.4f}",
          flush=True)
    results["v4_stacked_simclr_then_supervised"] = {
        "fold_aucs": fold_aucs4,
        "fold_mean": float(np.mean(fold_aucs4)),
        "pooled_oof_auc": pooled4,
    }

    # ---- Bootstrap CIs vs random-init baseline ----
    print(f"\n=== BOOTSTRAP ({N_BOOT} resamples) ===", flush=True)
    rng = np.random.default_rng(SEED)
    n_assigned = int(assigned1.sum())
    deltas_v2_v1 = []
    deltas_v3_v1 = []
    deltas_v4_v1 = []
    deltas_v4_v2 = []
    for _ in range(N_BOOT):
        idx = rng.integers(0, n_assigned, size=n_assigned)
        valid = np.where(assigned1)[0][idx]
        yb = y_rhuh[valid]
        if int(yb.sum()) < 2 or int(len(yb) - yb.sum()) < 2:
            continue
        a1 = auroc(oof1[valid], yb)
        a2 = auroc(oof2[valid], yb)
        a3 = auroc(oof3[valid], yb)
        a4 = auroc(oof4[valid], yb)
        deltas_v2_v1.append(a2 - a1)
        deltas_v3_v1.append(a3 - a1)
        deltas_v4_v1.append(a4 - a1)
        deltas_v4_v2.append(a4 - a2)
    deltas_v2_v1 = np.array(deltas_v2_v1)
    deltas_v3_v1 = np.array(deltas_v3_v1)
    deltas_v4_v1 = np.array(deltas_v4_v1)
    deltas_v4_v2 = np.array(deltas_v4_v2)
    for name, arr in [
        ("v2 - v1 (supervised - scratch)", deltas_v2_v1),
        ("v3 - v1 (SimCLR - scratch)", deltas_v3_v1),
        ("v4 - v1 (stacked - scratch)", deltas_v4_v1),
        ("v4 - v2 (stacked - supervised)", deltas_v4_v2),
    ]:
        p = float((arr <= 0).mean())
        ci_lo = float(np.percentile(arr, 2.5))
        ci_hi = float(np.percentile(arr, 97.5))
        print(f"  Delta {name}: mean={arr.mean():+.4f}, 95% CI "
              f"[{ci_lo:+.4f}, {ci_hi:+.4f}], P(<=0)={p:.4f}",
              flush=True)
        results[f"bootstrap_{name.split(' ')[0]}_minus_"
                f"{name.split(' ')[2]}"] = {
            "mean": float(arr.mean()),
            "95_CI": [ci_lo, ci_hi],
            "p_one_sided": p,
        }

    print(f"\n=== SUMMARY ===", flush=True)
    print(f"  v1 random init from-scratch:        AUC = {pooled1:.4f}",
          flush=True)
    print(f"  v2 supervised MU pretrain:           AUC = {pooled2:.4f}",
          flush=True)
    print(f"  v3 SimCLR (label-free) pretrain:     AUC = {pooled3:.4f}",
          flush=True)
    print(f"  v4 STACKED (SimCLR + supervised MU): AUC = {pooled4:.4f}",
          flush=True)

    out = {
        "version": "v217",
        "experiment": ("Comprehensive 4-way pretraining-strategy "
                       "ablation on RHUH transfer"),
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "device": str(DEVICE),
        "horizon_days": HORIZON,
        "n_folds": N_FOLDS,
        "n_pretrain_samples": len(pretrain_samples),
        "n_mu_labelled": len(mu_data),
        "n_rhuh_labelled": len(rhuh_data),
        "results": results,
        "elapsed_seconds": time.time() - t_start,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2))
    print(f"\nSaved {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
